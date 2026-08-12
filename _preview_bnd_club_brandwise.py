"""BND club order — brand-wise Excel preview."""
from __future__ import annotations

from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SRC = Path(r"G:\My Drive\2026-2027\Oder Management\AW26 order\Bedsheet\BND.xlsx")
OUT = Path(r"E:\test files\BND_Club_Order_Brandwise.xlsx")
OUT_REPO = Path(r"E:\centralized-db-system\Output\BND_Club_Order_Brandwise.xlsx")


def _key(brand, tc, size):
    return (
        str(brand or "").strip(),
        str(tc if tc is not None else "").strip(),
        str(size or "").strip(),
    )


def _num(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _pretty(n: float):
    return int(n) if float(n).is_integer() else n


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT_REPO.parent.mkdir(parents=True, exist_ok=True)

    wb_src = openpyxl.load_workbook(SRC, data_only=True)
    lines = {}

    ws = wb_src["base order"]
    for r in range(3, (ws.max_row or 0) + 1):
        brand = ws.cell(r, 1).value
        if not brand or not str(brand).strip():
            continue
        k = _key(brand, ws.cell(r, 2).value, ws.cell(r, 3).value)
        qty = _num(ws.cell(r, 23).value)  # W = Qty
        if k not in lines:
            lines[k] = {
                "Brand": str(brand).strip(),
                "TC": ws.cell(r, 2).value,
                "Size": str(ws.cell(r, 3).value or "").strip(),
                "Units": ws.cell(r, 4).value,
                "Product": ws.cell(r, 8).value,
                "Bale Size": ws.cell(r, 10).value,
                "Base Qty (W)": 0.0,
                "Additional Qty (U)": 0.0,
            }
        lines[k]["Base Qty (W)"] += qty

    ws = wb_src["additional order"]
    for r in range(3, (ws.max_row or 0) + 1):
        brand = ws.cell(r, 1).value
        if not brand or not str(brand).strip():
            continue
        k = _key(brand, ws.cell(r, 2).value, ws.cell(r, 3).value)
        qty = _num(ws.cell(r, 21).value)  # U = Additional quantity
        if k not in lines:
            lines[k] = {
                "Brand": str(brand).strip(),
                "TC": ws.cell(r, 2).value,
                "Size": str(ws.cell(r, 3).value or "").strip(),
                "Units": ws.cell(r, 4).value,
                "Product": ws.cell(r, 8).value,
                "Bale Size": ws.cell(r, 10).value,
                "Base Qty (W)": 0.0,
                "Additional Qty (U)": 0.0,
            }
        lines[k]["Additional Qty (U)"] += qty

    rows = list(lines.values())
    for row in rows:
        row["Club Qty"] = row["Base Qty (W)"] + row["Additional Qty (U)"]
    rows.sort(key=lambda r: (str(r["Brand"]).upper(), str(r["Size"]).upper(), str(r.get("TC") or "")))

    brand_tot = defaultdict(
        lambda: {"Base Qty (W)": 0.0, "Additional Qty (U)": 0.0, "Club Qty": 0.0, "Lines": 0}
    )
    for r in rows:
        b = r["Brand"]
        brand_tot[b]["Base Qty (W)"] += r["Base Qty (W)"]
        brand_tot[b]["Additional Qty (U)"] += r["Additional Qty (U)"]
        brand_tot[b]["Club Qty"] += r["Club Qty"]
        brand_tot[b]["Lines"] += 1

    brand_names = sorted(brand_tot.keys(), key=lambda x: x.upper())
    grand = {"Base Qty (W)": 0.0, "Additional Qty (U)": 0.0, "Club Qty": 0.0, "Lines": 0}
    for b in brand_names:
        for k in grand:
            grand[k] += brand_tot[b][k]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    brand_fill = PatternFill("solid", fgColor="E8F0FE")
    thin = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    wb = Workbook()

    # Brand Summary
    ws1 = wb.active
    ws1.title = "Brand Summary"
    headers1 = ["Brand", "Lines", "Base Qty (W)", "Additional Qty (U)", "Club Qty"]
    for i, h in enumerate(headers1, 1):
        cell = ws1.cell(1, i, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    r_i = 2
    for b in brand_names:
        t = brand_tot[b]
        vals = [b, t["Lines"], _pretty(t["Base Qty (W)"]), _pretty(t["Additional Qty (U)"]), _pretty(t["Club Qty"])]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(r_i, c, v)
            cell.border = thin
        r_i += 1

    vals = ["TOTAL", grand["Lines"], _pretty(grand["Base Qty (W)"]), _pretty(grand["Additional Qty (U)"]), _pretty(grand["Club Qty"])]
    for c, v in enumerate(vals, 1):
        cell = ws1.cell(r_i, c, v)
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.border = thin

    ws1.freeze_panes = "A2"
    for i, w in enumerate([28, 10, 14, 18, 12], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Line Detail
    ws2 = wb.create_sheet("Line Detail")
    headers2 = [
        "Brand", "TC", "Size", "Units", "Product", "Bale Size",
        "Base Qty (W)", "Additional Qty (U)", "Club Qty",
    ]
    for i, h in enumerate(headers2, 1):
        cell = ws2.cell(1, i, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    r_i = 2
    for b in brand_names:
        t = brand_tot[b]
        ws2.cell(r_i, 1, b)
        ws2.cell(r_i, 7, _pretty(t["Base Qty (W)"]))
        ws2.cell(r_i, 8, _pretty(t["Additional Qty (U)"]))
        ws2.cell(r_i, 9, _pretty(t["Club Qty"]))
        for c in range(1, 10):
            ws2.cell(r_i, c).fill = brand_fill
            ws2.cell(r_i, c).font = Font(bold=True)
            ws2.cell(r_i, c).border = thin
        r_i += 1
        for x in [row for row in rows if row["Brand"] == b]:
            vals = [
                x["Brand"], x["TC"], x["Size"], x["Units"], x["Product"], x["Bale Size"],
                _pretty(x["Base Qty (W)"]), _pretty(x["Additional Qty (U)"]), _pretty(x["Club Qty"]),
            ]
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(r_i, c, v)
                cell.border = thin
            r_i += 1

    ws2.cell(r_i, 1, "GRAND TOTAL")
    ws2.cell(r_i, 7, _pretty(grand["Base Qty (W)"]))
    ws2.cell(r_i, 8, _pretty(grand["Additional Qty (U)"]))
    ws2.cell(r_i, 9, _pretty(grand["Club Qty"]))
    for c in range(1, 10):
        ws2.cell(r_i, c).fill = total_fill
        ws2.cell(r_i, c).font = Font(bold=True)
        ws2.cell(r_i, c).border = thin

    ws2.freeze_panes = "A2"
    for i, w in enumerate([22, 8, 10, 8, 16, 10, 14, 18, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws3 = wb.create_sheet("Locks")
    ws3["A1"] = "BND club order — teaching preview"
    notes = [
        "Source: BND.xlsx (Bedsheet Booking form Jun-26)",
        "base order Qty = column W (Qty)",
        "additional order Qty = column U (Additional quantity)",
        "Club Qty = Base + Additional per Brand + TC + Size",
        f"Club total pieces = {_pretty(grand['Club Qty'])}",
    ]
    for i, line in enumerate(notes, 3):
        ws3[f"A{i}"] = line
    ws3.column_dimensions["A"].width = 70

    wb.save(OUT)
    wb.save(OUT_REPO)
    print(f"wrote {OUT}")
    print(f"wrote {OUT_REPO}")
    print(f"brands={len(brand_names)} lines={len(rows)} club={_pretty(grand['Club Qty'])}")
    for b in brand_names:
        print(f"  {b}: {_pretty(brand_tot[b]['Club Qty'])}")


if __name__ == "__main__":
    main()
