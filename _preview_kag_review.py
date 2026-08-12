"""Focused review of kag.xlsx (updated layout with Revised Order)."""
from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import subprocess

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SRC = Path(r"g:\My Drive\2026-2027\Oder Management\AW26 order\Bedsheet\kag.xlsx")
OUT = Path(r"E:\test files\KAG_Order_Review.xlsx")


def main():
    wb_f = load_workbook(SRC, data_only=False)
    wb = load_workbook(SRC, data_only=True)
    ws = wb.active
    wsf = wb_f.active

    print("HEADERS R1:")
    for c in range(1, 30):
        v = ws.cell(1, c).value
        if v:
            print(f"  {get_column_letter(c)}: {v}")

    lines_all = []
    order_lines = []
    for r in range(2, (ws.max_row or 2) + 1):
        brand = ws.cell(r, 2).value
        if not brand:
            continue
        bs = float(ws.cell(r, 11).value or 0)
        ex = float(ws.cell(r, 24).value or 0)
        asst = ws.cell(r, 25).value
        qnty = ws.cell(r, 26).value
        revised = ws.cell(r, 27).value
        sheet_val = ws.cell(r, 28).value
        asst = float(asst) if asst not in (None, "") else None
        qnty = float(qnty) if qnty not in (None, "") else None
        revised = float(revised) if revised not in (None, "") else None
        sheet_val = float(sheet_val) if sheet_val not in (None, "") else None
        row = {
            "brand": str(brand).strip(),
            "tc": ws.cell(r, 3).value,
            "size": str(ws.cell(r, 4).value or ""),
            "product": ws.cell(r, 9).value,
            "bs": bs,
            "ex": ex,
            "asst": asst,
            "qnty": qnty,
            "revised": revised,
            "sheet_val_lakhs": sheet_val,
        }
        lines_all.append(row)
        if revised is not None and revised > 0:
            row["qty"] = revised
            row["value"] = revised * ex
            order_lines.append(row)

    tq = sum(x["qty"] for x in order_lines)
    tv = sum(x["value"] for x in order_lines)
    print()
    print("Order lines (Revised Order > 0):", len(order_lines))
    print("TOTAL qty", tq, "value Rs", round(tv, 2))
    print()
    print("Revised vs Qnty:")
    for x in order_lines:
        if x["qnty"] is not None and abs(x["revised"] - x["qnty"]) >= 0.01:
            print(
                f"  DIFF {x['brand']} {x['size']}: "
                f"Revised={x['revised']:g} Qnty={x['qnty']:g} Asst={x['asst']}"
            )

    club = defaultdict(lambda: [0.0, 0.0])
    for x in order_lines:
        club[x["brand"]][0] += x["qty"]
        club[x["brand"]][1] += x["value"]
    print()
    print("BRANDWISE:")
    for b in sorted(club):
        print(f"  {b}: {club[b][0]:g} pcs, Rs {round(club[b][1], 2)}")

    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", "0F766E")
    nf = Font(bold=True)
    tfill = PatternFill("solid", "CCFBF1")
    warn = PatternFill("solid", "FEF3C7")

    wbout = Workbook()
    o = wbout.active
    o.title = "Overview"
    o.append(["Source", "kag.xlsx"])
    o.append(["Distributor (filename)", "KAG"])
    o.append(["Header row", "Row 1"])
    o.append([])
    o.append(["Column", "Meaning (as read)", "Used for order?"])
    o.append(["Y Asst", "Formula = NoOfDesign * BaleSize", "No — reference only"])
    o.append(
        [
            "Z Qnty",
            "Formula = 3 * NoOfDesign * Color",
            "No — not final when Revised exists",
        ]
    )
    o.append(["AA Revised Order", "Typed/final order qty", "YES — order Qty"])
    o.append(
        [
            "AB val",
            "= Revised * ExMill / 100000 (lakhs)",
            "No — we use full Rs = Qty * ExMill",
        ]
    )
    o.append([])
    o.append(["Qty rule", "Revised Order only where Revised > 0"])
    o.append(["Bales", "No bales column — do not invent"])
    o.append(["Value", "Qty * ExMill (full rupees)"])
    o.append([])
    o.append(["Total Qty (pcs)", tq])
    o.append(["Total Value (Rs)", round(tv, 2)])
    o.append(["Order lines", len(order_lines)])
    o.append(
        [
            "Rows skipped (no Revised)",
            sum(1 for x in lines_all if not x["revised"]),
        ]
    )
    o.column_dimensions["A"].width = 34
    o.column_dimensions["B"].width = 55
    o.column_dimensions["C"].width = 40

    ws = wbout.create_sheet("Brandwise")
    ws.append(["Brand", "Qty (pcs)", "Value (=Qty*ExMill)"])
    for cell in ws[1]:
        cell.font = hf
        cell.fill = hfill
        cell.border = thin
    for b in sorted(club):
        ws.append([b, round(club[b][0], 2), round(club[b][1], 2)])
    ws.append(["TOTAL", round(tq, 2), round(tv, 2)])
    for c in ws[ws.max_row]:
        c.font = nf
        c.fill = tfill
    for col, w in zip("ABC", [28, 12, 18]):
        ws.column_dimensions[col].width = w

    ws = wbout.create_sheet("Order Lines")
    ws.append(
        [
            "Brand",
            "Size",
            "Bale Size",
            "ExMill",
            "Asst (ref)",
            "Qnty (ref)",
            "Revised Order (=Qty used)",
            "Value Rs (=Qty*Ex)",
            "Sheet val (lakhs)",
            "Note",
        ]
    )
    for cell in ws[1]:
        cell.font = hf
        cell.fill = hfill
    for x in order_lines:
        note = ""
        if x["qnty"] is not None and abs(x["revised"] - x["qnty"]) >= 0.01:
            note = "Revised differs from Qnty"
        elif x["asst"] is not None and abs(x["revised"] - x["asst"]) < 0.01:
            note = "Revised equals Asst"
        ws.append(
            [
                x["brand"],
                x["size"],
                x["bs"],
                round(x["ex"], 2),
                x["asst"],
                x["qnty"],
                x["revised"],
                round(x["value"], 2),
                round(x["sheet_val_lakhs"], 4) if x["sheet_val_lakhs"] else 0,
                note,
            ]
        )
        if note.startswith("Revised differs"):
            for c in range(1, 11):
                ws.cell(ws.max_row, c).fill = warn
    for i, w in enumerate([22, 10, 10, 10, 12, 12, 22, 16, 14, 28], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws = wbout.create_sheet("Skipped rows")
    ws.append(["These rows have no Revised Order — not counted"])
    ws.append(["Brand", "Size", "Asst", "Qnty", "Revised"])
    for cell in ws[2]:
        cell.font = hf
        cell.fill = hfill
    for x in lines_all:
        if x["revised"]:
            continue
        ws.append([x["brand"], x["size"], x["asst"], x["qnty"], x["revised"]])
    for col, w in zip("ABCDE", [22, 10, 10, 10, 10]):
        ws.column_dimensions[col].width = w

    OUT.parent.mkdir(parents=True, exist_ok=True)
    try:
        wbout.save(OUT)
        saved = OUT
    except PermissionError:
        saved = OUT.with_name("KAG_Order_Review_v2.xlsx")
        wbout.save(saved)

    print("Saved", saved)
    subprocess.Popen(["cmd", "/c", "start", "", str(saved)], shell=False)


if __name__ == "__main__":
    main()
