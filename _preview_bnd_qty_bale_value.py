"""BND club order — brand-wise Qty + Bales + Value."""
from __future__ import annotations

from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SRC = Path(r"G:\My Drive\2026-2027\Oder Management\AW26 order\Bedsheet\BND.xlsx")
OUT = Path(r"E:\test files\BND_Club_Brandwise_Qty_Bale_Value.xlsx")
OUT_REPO = Path(r"E:\centralized-db-system\Output\BND_Club_Brandwise_Qty_Bale_Value.xlsx")


def _key(brand, tc, size):
    return (
        str(brand or "").strip(),
        str(tc if tc is not None else "").strip(),
        str(size or "").strip(),
    )


def _num(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _pretty(n: float, money: bool = False):
    if money:
        return round(float(n), 2)
    if abs(float(n) - round(float(n))) < 1e-9:
        return int(round(float(n)))
    return round(float(n), 2)


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT_REPO.parent.mkdir(parents=True, exist_ok=True)

    wb_src = openpyxl.load_workbook(SRC, data_only=True)
    lines = {}

    def ensure(k, brand, tc, size, units, product, bale_size):
        if k not in lines:
            lines[k] = {
                "Brand": str(brand).strip(),
                "TC": tc,
                "Size": str(size or "").strip(),
                "Units": units,
                "Product": product,
                "Bale Size": bale_size,
                "Base Qty": 0.0,
                "Addl Qty": 0.0,
                "Base Bales": 0.0,
                "Addl Bales": 0.0,
                "Base Value": 0.0,
                "Addl Value": 0.0,
            }

    # base order: Qty=W(23), No of Bales=V(22), AWD Value=X(24), Bale Size=J(10)
    ws = wb_src["base order"]
    for r in range(3, (ws.max_row or 0) + 1):
        brand = ws.cell(r, 1).value
        if not brand or not str(brand).strip():
            continue
        tc, size = ws.cell(r, 2).value, ws.cell(r, 3).value
        k = _key(brand, tc, size)
        bale_size = ws.cell(r, 10).value
        ensure(k, brand, tc, size, ws.cell(r, 4).value, ws.cell(r, 8).value, bale_size)
        lines[k]["Base Qty"] += _num(ws.cell(r, 23).value)
        lines[k]["Base Bales"] += _num(ws.cell(r, 22).value)
        lines[k]["Base Value"] += _num(ws.cell(r, 24).value)
        if bale_size and not lines[k]["Bale Size"]:
            lines[k]["Bale Size"] = bale_size

    # additional: Qty=U(21), Value=V(22), Bales = Qty / Bale Size (no bale column)
    ws = wb_src["additional order"]
    for r in range(3, (ws.max_row or 0) + 1):
        brand = ws.cell(r, 1).value
        if not brand or not str(brand).strip():
            continue
        tc, size = ws.cell(r, 2).value, ws.cell(r, 3).value
        k = _key(brand, tc, size)
        bale_size = ws.cell(r, 10).value
        ensure(k, brand, tc, size, ws.cell(r, 4).value, ws.cell(r, 8).value, bale_size)
        qty = _num(ws.cell(r, 21).value)
        val = _num(ws.cell(r, 22).value)
        bs = _num(bale_size) or _num(lines[k]["Bale Size"])
        lines[k]["Addl Qty"] += qty
        lines[k]["Addl Value"] += val
        lines[k]["Addl Bales"] += (qty / bs) if bs else 0.0
        if bale_size:
            lines[k]["Bale Size"] = bale_size

    rows = list(lines.values())
    for row in rows:
        row["Club Qty"] = row["Base Qty"] + row["Addl Qty"]
        row["Club Bales"] = row["Base Bales"] + row["Addl Bales"]
        row["Club Value"] = row["Base Value"] + row["Addl Value"]

    rows.sort(
        key=lambda r: (str(r["Brand"]).upper(), str(r["Size"]).upper(), str(r.get("TC") or ""))
    )

    brand_tot = defaultdict(
        lambda: {
            "Lines": 0,
            "Base Qty": 0.0,
            "Addl Qty": 0.0,
            "Club Qty": 0.0,
            "Base Bales": 0.0,
            "Addl Bales": 0.0,
            "Club Bales": 0.0,
            "Base Value": 0.0,
            "Addl Value": 0.0,
            "Club Value": 0.0,
        }
    )
    for r in rows:
        b = r["Brand"]
        brand_tot[b]["Lines"] += 1
        for k in (
            "Base Qty", "Addl Qty", "Club Qty",
            "Base Bales", "Addl Bales", "Club Bales",
            "Base Value", "Addl Value", "Club Value",
        ):
            brand_tot[b][k] += r[k]

    brand_names = sorted(brand_tot.keys(), key=lambda x: x.upper())
    grand = {k: 0.0 for k in brand_tot[brand_names[0]]}
    for b in brand_names:
        for k, v in brand_tot[b].items():
            grand[k] += v

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    brand_fill = PatternFill("solid", fgColor="E8F0FE")
    thin = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    wb = Workbook()

    # ---- Brand Summary ----
    ws1 = wb.active
    ws1.title = "Brand Summary"
    headers1 = [
        "Brand", "Lines",
        "Base Qty", "Addl Qty", "Club Qty",
        "Base Bales", "Addl Bales", "Club Bales",
        "Base Value", "Addl Value", "Club Value",
    ]
    for i, h in enumerate(headers1, 1):
        cell = ws1.cell(1, i, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    r_i = 2
    for b in brand_names:
        t = brand_tot[b]
        vals = [
            b, t["Lines"],
            _pretty(t["Base Qty"]), _pretty(t["Addl Qty"]), _pretty(t["Club Qty"]),
            _pretty(t["Base Bales"]), _pretty(t["Addl Bales"]), _pretty(t["Club Bales"]),
            _pretty(t["Base Value"], money=True),
            _pretty(t["Addl Value"], money=True),
            _pretty(t["Club Value"], money=True),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(r_i, c, v)
            cell.border = thin
            if c >= 9:
                cell.number_format = "#,##0.00"
            elif c >= 3:
                cell.number_format = "#,##0.##"
        r_i += 1

    vals = [
        "TOTAL", int(grand["Lines"]),
        _pretty(grand["Base Qty"]), _pretty(grand["Addl Qty"]), _pretty(grand["Club Qty"]),
        _pretty(grand["Base Bales"]), _pretty(grand["Addl Bales"]), _pretty(grand["Club Bales"]),
        _pretty(grand["Base Value"], money=True),
        _pretty(grand["Addl Value"], money=True),
        _pretty(grand["Club Value"], money=True),
    ]
    for c, v in enumerate(vals, 1):
        cell = ws1.cell(r_i, c, v)
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.border = thin
        if c >= 9:
            cell.number_format = "#,##0.00"
        elif c >= 3:
            cell.number_format = "#,##0.##"

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:K{r_i}"
    widths = [22, 8, 11, 11, 11, 11, 11, 11, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ---- Line Detail ----
    ws2 = wb.create_sheet("Line Detail")
    headers2 = [
        "Brand", "TC", "Size", "Units", "Product", "Bale Size",
        "Base Qty", "Addl Qty", "Club Qty",
        "Base Bales", "Addl Bales", "Club Bales",
        "Base Value", "Addl Value", "Club Value",
    ]
    for i, h in enumerate(headers2, 1):
        cell = ws2.cell(1, i, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    r_i = 2
    for b in brand_names:
        t = brand_tot[b]
        # brand subtotal
        sub = [
            b, "", "", "", "", "",
            _pretty(t["Base Qty"]), _pretty(t["Addl Qty"]), _pretty(t["Club Qty"]),
            _pretty(t["Base Bales"]), _pretty(t["Addl Bales"]), _pretty(t["Club Bales"]),
            _pretty(t["Base Value"], money=True),
            _pretty(t["Addl Value"], money=True),
            _pretty(t["Club Value"], money=True),
        ]
        for c, v in enumerate(sub, 1):
            cell = ws2.cell(r_i, c, v)
            cell.fill = brand_fill
            cell.font = Font(bold=True)
            cell.border = thin
            if c >= 13:
                cell.number_format = "#,##0.00"
            elif c >= 7:
                cell.number_format = "#,##0.##"
        r_i += 1
        for x in [row for row in rows if row["Brand"] == b]:
            vals = [
                x["Brand"], x["TC"], x["Size"], x["Units"], x["Product"], x["Bale Size"],
                _pretty(x["Base Qty"]), _pretty(x["Addl Qty"]), _pretty(x["Club Qty"]),
                _pretty(x["Base Bales"]), _pretty(x["Addl Bales"]), _pretty(x["Club Bales"]),
                _pretty(x["Base Value"], money=True),
                _pretty(x["Addl Value"], money=True),
                _pretty(x["Club Value"], money=True),
            ]
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(r_i, c, v)
                cell.border = thin
                if c >= 13:
                    cell.number_format = "#,##0.00"
                elif c >= 7:
                    cell.number_format = "#,##0.##"
            r_i += 1

    grand_row = [
        "GRAND TOTAL", "", "", "", "", "",
        _pretty(grand["Base Qty"]), _pretty(grand["Addl Qty"]), _pretty(grand["Club Qty"]),
        _pretty(grand["Base Bales"]), _pretty(grand["Addl Bales"]), _pretty(grand["Club Bales"]),
        _pretty(grand["Base Value"], money=True),
        _pretty(grand["Addl Value"], money=True),
        _pretty(grand["Club Value"], money=True),
    ]
    for c, v in enumerate(grand_row, 1):
        cell = ws2.cell(r_i, c, v)
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.border = thin
        if c >= 13:
            cell.number_format = "#,##0.00"
        elif c >= 7:
            cell.number_format = "#,##0.##"

    ws2.freeze_panes = "A2"
    for i, w in enumerate([20, 8, 10, 8, 14, 10, 10, 10, 10, 10, 10, 10, 13, 13, 13], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ---- Locks ----
    ws3 = wb.create_sheet("How calculated")
    notes = [
        "BND club order — Qty / Bales / Value",
        "",
        "QTY",
        "  base order column W (Qty)",
        "  additional order column U (Additional quantity)",
        "  Club Qty = Base + Addl",
        "",
        "VALUE",
        "  base order column X (AWD Value) = ExMill x Qty",
        "  additional order column V (Value) = ExMill x Qty",
        "  Club Value = Base + Addl",
        "",
        "BALES",
        "  base order column V (No of Bales) — as written on sheet",
        "  additional order: no bale column → Addl Bales = Addl Qty / Bale Size",
        "  Club Bales = Base Bales + Addl Bales",
        "",
        f"TOTALS: Qty={_pretty(grand['Club Qty'])} | Bales={_pretty(grand['Club Bales'])} | Value={_pretty(grand['Club Value'], money=True)}",
    ]
    for i, line in enumerate(notes, 1):
        ws3[f"A{i}"] = line
    ws3.column_dimensions["A"].width = 90

    wb.save(OUT)
    wb.save(OUT_REPO)
    print(f"wrote {OUT}")
    print(f"wrote {OUT_REPO}")
    print(
        f"TOTAL qty={_pretty(grand['Club Qty'])} "
        f"bales={_pretty(grand['Club Bales'])} "
        f"value={_pretty(grand['Club Value'], money=True)}"
    )
    for b in brand_names:
        t = brand_tot[b]
        print(
            f"  {b}: qty={_pretty(t['Club Qty'])} "
            f"bales={_pretty(t['Club Bales'])} "
            f"value={_pretty(t['Club Value'], money=True)}"
        )


if __name__ == "__main__":
    main()
