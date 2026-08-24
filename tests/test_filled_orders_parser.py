"""
Quantity-column detection + bales/pieces normalization unit tests.

Synthetic fixtures mirror the real-file quirks the feature spec was
validated against (standard column, shifted/unlabeled column, multiple
candidates with a derived-sum relationship) since the actual distributor
files aren't checked into the repo.
"""

import sqlite3
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

import article_master_db as amdb
import article_master_parser as amparser
import filled_orders_parser as foparser


def _write_workbook(tmp_path, header, rows, filename="test.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for row in rows:
        ws.append(row)
    path = tmp_path / filename
    wb.save(path)
    return path


def _load_valid_rows(path):
    raw_df = pd.read_excel(path, sheet_name=0, header=None)
    header_idx = amparser.detect_header_row(raw_df)
    header_row = raw_df.iloc[header_idx].tolist()
    col_mapping = amparser.map_columns_to_core(header_row)
    data_rows = raw_df.iloc[header_idx + 1:]
    valid_rows = [
        row for _, row in data_rows.iterrows()
        if amparser.is_data_row(row.tolist(), col_mapping)
    ]
    return header_row, col_mapping, valid_rows


def test_standard_column_all_pieces(tmp_path):
    """kag.xlsx equivalent: standard 'Qnty' column, values already in pieces."""
    header = ["Brand", "Size", "Product", "MRP", "PTR", "Ex-Mill", "Bale Size", "Qnty"]
    rows = [
        ["ASTER", "DB BS", "Bedsheet SS-26", 999, 450, 400, 18, 108],
        ["ASTER", "KG BS", "Bedsheet SS-26", 1200, 500, 420, 12, 60],
    ]
    path = _write_workbook(tmp_path, header, rows)
    header_row, col_mapping, valid_rows = _load_valid_rows(path)

    detection = foparser.detect_quantity_column(header_row, col_mapping, "Bed", valid_rows)
    assert detection["status"] == "ok"
    assert detection["column_label"] == "Qnty"

    items = foparser.build_filled_order_rows(valid_rows, header_row, col_mapping, detection["column_index"])
    assert len(items) == 2
    for item in items:
        unit, final_qty = foparser.normalize_quantity(
            item["raw_qty_value"], item["core_fields"]["bale_pack_size"],
        )
        assert unit == "pieces"
        assert final_qty == item["raw_qty_value"]


def test_bath_qty_in_bales_large_bale_count(tmp_path):
    """Towel booking sheet: 30 bales × 24 pack must be 720 pcs, not 30."""
    header = ["Brand", "Size", "Product", "MRP", "PTR", "Ex-Mill Per Pcs", "Bale Pack Sizes", "Qty in Bales"]
    rows = [["Flora", "R4", "Terry Towel", 1425, 954.75, 777.17, 24, 30]]
    path = _write_workbook(tmp_path, header, rows)
    header_row, col_mapping, valid_rows = _load_valid_rows(path)

    detection = foparser.detect_quantity_column(header_row, col_mapping, "Bath", valid_rows)
    items = foparser.build_filled_order_rows(valid_rows, header_row, col_mapping, detection["column_index"])
    unit, final = foparser.normalize_quantity(
        items[0]["raw_qty_value"],
        items[0]["core_fields"]["bale_pack_size"],
        qty_column_label=detection["column_label"],
        category="Bath",
    )
    assert unit == "bales"
    assert final == 720


def test_is_bale_count_column():
    assert foparser.is_bale_count_quantity_column("Qty in Bales", "Bath") is True
    assert foparser.is_bale_count_quantity_column("Qnty", "Bed") is False
    assert foparser.is_bale_count_quantity_column("Additional Order Qty", "Bed") is False


def test_standard_column_all_bales(tmp_path):
    """KAG_AGRA.xlsx equivalent: standard 'Qty in Bales' column, all bale counts."""
    header = ["Brand", "Size", "Product", "MRP", "PTR", "Ex-Mill Per Pcs", "Bale Pack Sizes", "Qty in Bales"]
    rows = [
        ["TerryCo", "40x60", "Towel", 300, 150, 120, 96, 1],
        ["TerryCo", "70x140", "Towel", 500, 250, 200, 36, 3],
    ]
    path = _write_workbook(tmp_path, header, rows)
    header_row, col_mapping, valid_rows = _load_valid_rows(path)

    detection = foparser.detect_quantity_column(header_row, col_mapping, "Bath", valid_rows)
    assert detection["status"] == "ok"
    assert detection["column_label"].strip().lower() == "qty in bales"

    items = foparser.build_filled_order_rows(valid_rows, header_row, col_mapping, detection["column_index"])
    unit0, final0 = foparser.normalize_quantity(
        items[0]["raw_qty_value"],
        items[0]["core_fields"]["bale_pack_size"],
        qty_column_label=detection["column_label"],
        category="Bath",
    )
    assert (unit0, final0) == ("bales", 96)
    unit1, final1 = foparser.normalize_quantity(
        items[1]["raw_qty_value"],
        items[1]["core_fields"]["bale_pack_size"],
        qty_column_label=detection["column_label"],
        category="Bath",
    )
    assert (unit1, final1) == ("bales", 108)


def test_shifted_unlabeled_column(tmp_path):
    """DCA_Order.xlsx equivalent: standard 'Qnty' column is 100% empty, real
    quantities live in the next (unlabeled) column for some rows."""
    header = ["Brand", "Size", "Product", "MRP", "PTR", "Ex-Mill", "Bale Size", "Qnty", None]
    rows = [
        ["A1", "S1", "Bedsheet SS-26", 100, 50, 40, 10, None, 50],
        ["A2", "S2", "Bedsheet SS-26", 200, 90, 70, 12, None, None],
        ["A3", "S3", "Bedsheet SS-26", 150, 70, 55, 15, None, 45],
    ]
    path = _write_workbook(tmp_path, header, rows)
    header_row, col_mapping, valid_rows = _load_valid_rows(path)

    detection = foparser.detect_quantity_column(header_row, col_mapping, "Bed", valid_rows)
    assert detection["status"] == "ok"
    assert "unlabeled" in detection["column_label"].lower()

    items = foparser.build_filled_order_rows(valid_rows, header_row, col_mapping, detection["column_index"])
    assert len(items) == 2  # blank-qty rows are skipped entirely
    assert {i["raw_qty_value"] for i in items} == {50, 45}


def test_qty_plus_bales_auto_selects_qty_no_prompt(tmp_path):
    """Choice-style sheet: Qty + No of Bales → never ask; Qty is source of truth."""
    header = [
        "Brand", "Size", "Product", "MRP", "PTR", "Ex-Mill", "Bale Size",
        "No of Bales", "Qty",
    ]
    rows = [
        ["Florentine", "King", "Bedsheet SS-26", 1000, 500, 400, 12, 8, 96],
        ["Florentine", "Queen", "Bedsheet SS-26", 900, 450, 360, 12, 2, 24],
        ["Marigold", "King", "Bedsheet SS-26", 1100, 550, 440, 10, 4, 40],
    ]
    path = _write_workbook(tmp_path, header, rows)
    header_row, col_mapping, valid_rows = _load_valid_rows(path)

    detection = foparser.detect_quantity_column(header_row, col_mapping, "Bed", valid_rows)
    assert detection["status"] == "ok"
    assert detection["column_label"].strip().lower() == "qty"


def test_multiple_candidates_needs_confirmation_with_verified_sum(tmp_path):
    """BND.xlsx equivalent: several populated candidate columns, no single
    standard alias present. Additional Order Qty = Qty + Add for every row —
    verified mathematically and surfaced as a hint, not auto-picked."""
    header = [
        "Brand", "Size", "Product", "MRP", "PTR", "Ex-Mill", "Bale Size",
        "No of Bales", "Qty", "Add", "Additional Order Qty",
    ]
    rows = [
        ["Florentine", "King", "Bedsheet SS-26", 1000, 500, 400, 12, 5, 60, 10, 70],
        ["Florentine", "Queen", "Bedsheet SS-26", 900, 450, 360, 12, 8, 96, 0, 96],
        ["Marigold", "King", "Bedsheet SS-26", 1100, 550, 440, 10, 3, 30, 5, 35],
        ["Marigold", "Queen", "Bedsheet SS-26", 950, 480, 380, 10, 2, 20, 2, 22],
    ]
    path = _write_workbook(tmp_path, header, rows)
    header_row, col_mapping, valid_rows = _load_valid_rows(path)

    detection = foparser.detect_quantity_column(header_row, col_mapping, "Bed", valid_rows)
    assert detection["status"] == "ok"
    assert detection["column_label"] == "Additional Order Qty"
    assert "auto_selected_reason" in detection

    items = foparser.build_filled_order_rows(valid_rows, header_row, col_mapping, detection["column_index"])
    assert [it["raw_qty_value"] for it in items] == [70, 96, 35, 22]

    flags = []
    for it in items:
        unit, final_qty = foparser.normalize_quantity(it["raw_qty_value"], it["core_fields"]["bale_pack_size"])
        assert unit == "pieces"  # every raw value here is >= its bale size
        flags.append(foparser.is_clean_bale_multiple(final_qty, it["core_fields"]["bale_pack_size"]))
    # 70 % 12 != 0 and 35 % 10 != 0 / 22 % 10 != 0 -> flagged; 96 % 12 == 0 -> clean
    assert flags == [False, True, False, False]


def _make_am_conn(tmp_path):
    db_path = tmp_path / "am_match.sqlite3"
    schema_path = Path(__file__).resolve().parent.parent / "article_master_schema.sql"
    conn = sqlite3.connect(db_path)
    with open(schema_path, encoding="utf-8") as f:
        conn.executescript(f.read())
    return conn


def test_match_and_normalize_matched_uses_article_master_values(tmp_path):
    header = ["Brand", "Size", "Product", "MRP", "PTR", "Ex-Mill", "Bale Size", "Qnty"]
    # File's own price/bale columns are intentionally wrong/stale — matched
    # rows must ignore them and use Article Master's current values instead.
    rows = [["ASTER", "DB BS", "Bedsheet SS-26", 111, 55, 44, 6, 30]]
    path = _write_workbook(tmp_path, header, rows)
    header_row, col_mapping, valid_rows = _load_valid_rows(path)
    detection = foparser.detect_quantity_column(header_row, col_mapping, "Bed", valid_rows)
    parsed_rows = foparser.build_filled_order_rows(valid_rows, header_row, col_mapping, detection["column_index"])

    conn = _make_am_conn(tmp_path)
    amdb.create_category(conn, 1, "Bed", ["brand", "size"], is_confirmed=True)
    amdb.upsert_article(conn, 1, {
        "category": "Bed", "product_type": "Bedsheet", "brand": "ASTER", "size": "DB BS",
        "mrp": 999, "ptr": 450, "ex_mill_price": 400, "bale_pack_size": 10,
        "item_key": "ASTER|DB BS", "extra_attributes": {},
    })

    result = foparser.match_and_normalize(conn, amdb, 1, parsed_rows[0], ["brand", "size"])
    conn.close()

    assert result["matched"] is True
    assert result["mrp"] == 999
    assert result["bale_size_used"] == 10
    assert result["detected_unit"] == "pieces"
    assert result["final_piece_qty"] == 30
    assert result["is_clean_bale_multiple"] is True  # 30 % 10 == 0


def test_dbl_bs_matches_db_bs_in_article_master(tmp_path):
    """Locked teaching: distributor 'DBL BS' == Article Master 'DB BS' (Double)."""
    header = ["Brand", "Size", "Product", "MRP", "PTR", "Ex-Mill", "Bale Size", "Qnty"]
    rows = [["Cotton Comforts", "DBL BS", "Sheet Sets", 2499, 1547, 1311, 12, 24]]
    path = _write_workbook(tmp_path, header, rows)
    header_row, col_mapping, valid_rows = _load_valid_rows(path)
    detection = foparser.detect_quantity_column(header_row, col_mapping, "Bed", valid_rows)
    parsed_rows = foparser.build_filled_order_rows(valid_rows, header_row, col_mapping, detection["column_index"])

    conn = _make_am_conn(tmp_path)
    amdb.create_category(conn, 1, "Bed", ["brand", "size"], is_confirmed=True)
    amdb.upsert_article(conn, 1, {
        "category": "Bed", "product_type": "Sheet Sets", "brand": "Cotton Comforts", "size": "DB BS",
        "mrp": 2499, "ptr": 1547, "ex_mill_price": 1311, "bale_pack_size": 12,
        "item_key": "COTTON COMFORTS|DB BS", "extra_attributes": {},
    })

    result = foparser.match_and_normalize(
        conn, amdb, 1, parsed_rows[0], ["brand", "size"], category="Bed",
    )
    conn.close()

    assert result["matched"] is True
    assert result["size"] == "DBL BS"  # file spelling preserved on the line
    assert result["item_key"].endswith("|DB BS")  # match key uses canonical size


def test_match_and_normalize_unmatched_falls_back_to_file_values(tmp_path):
    header = ["Brand", "Size", "Product", "MRP", "PTR", "Ex-Mill", "Bale Size", "Qnty"]
    rows = [["NEWBRAND", "XL", "Bedsheet SS-26", 111, 55, 44, 6, 30]]
    path = _write_workbook(tmp_path, header, rows)
    header_row, col_mapping, valid_rows = _load_valid_rows(path)
    detection = foparser.detect_quantity_column(header_row, col_mapping, "Bed", valid_rows)
    parsed_rows = foparser.build_filled_order_rows(valid_rows, header_row, col_mapping, detection["column_index"])

    conn = _make_am_conn(tmp_path)
    result = foparser.match_and_normalize(conn, amdb, 1, parsed_rows[0], ["brand", "size"])
    conn.close()

    assert result["matched"] is False
    assert result["article_id"] is None
    assert result["mrp"] == 111
    assert result["bale_size_used"] == 6


def test_bath_fo_shade_matches_am_color(tmp_path):
    """Choice Corner style: Shade column + physical size must match AM Color + Bath Towel."""
    header = [
        "Brand", "Size", "Product", "Shade", "MRP", "PTR", "Ex-Mill", "Bale Pack Sizes", "Qty in Bales",
    ]
    rows = [["Tulip", "75x150", "Terry Towel", "WHITE", 999, 500, 400, 24, 2]]
    path = _write_workbook(tmp_path, header, rows)
    header_row, col_mapping, valid_rows = _load_valid_rows(path)
    detection = foparser.detect_quantity_column(header_row, col_mapping, "Bath", valid_rows)
    parsed_rows = foparser.build_filled_order_rows(
        valid_rows, header_row, col_mapping, detection["column_index"],
    )

    conn = _make_am_conn(tmp_path)
    amdb.create_category(
        conn, 1, "Bath", ["brand", "size", "color", "product"], is_confirmed=True,
    )
    amdb.upsert_article(conn, 1, {
        "category": "Bath",
        "product_type": "Terry Towel",
        "brand": "Tulip",
        "size": "Bath Towel",
        "mrp": 999,
        "ptr": 500,
        "ex_mill_price": 400,
        "bale_pack_size": 24,
        "item_key": "TULIP|BATH TOWEL|WHITE|TERRY TOWEL",
        "extra_attributes": {"Color": "White", "BS Size": "75x150"},
    })

    result = foparser.match_and_normalize(
        conn, amdb, 1, parsed_rows[0],
        ["brand", "size", "color", "product"],
        category="Bath",
    )
    conn.close()

    assert result["matched"] is True
    assert result["size"] == "75x150"  # file physical size kept for display
    assert (result["extra_attributes"] or {}).get("Color") == "White"
    assert (result["extra_attributes"] or {}).get("BS Size") == "75x150"


def test_bath_fo_blank_color_matches_am_with_color(tmp_path):
    """FO without Shade/Color still matches when brand+size+product are unique."""
    header = [
        "Brand", "Size", "Product", "MRP", "PTR", "Ex-Mill", "Bale Pack Sizes", "Qty in Bales",
    ]
    rows = [["Santino", "40x60(2pc)", "Terry Towel", 800, 400, 350, 12, 1]]
    path = _write_workbook(tmp_path, header, rows)
    header_row, col_mapping, valid_rows = _load_valid_rows(path)
    detection = foparser.detect_quantity_column(header_row, col_mapping, "Bath", valid_rows)
    parsed_rows = foparser.build_filled_order_rows(
        valid_rows, header_row, col_mapping, detection["column_index"],
    )

    conn = _make_am_conn(tmp_path)
    amdb.create_category(
        conn, 1, "Bath", ["brand", "size", "color", "product"], is_confirmed=True,
    )
    amdb.upsert_article(conn, 1, {
        "category": "Bath",
        "product_type": "Terry Towel",
        "brand": "Santino",
        "size": "Hand Towel Set of 2",
        "mrp": 800,
        "ptr": 400,
        "ex_mill_price": 350,
        "bale_pack_size": 12,
        "item_key": "SANTINO|HAND TOWEL SET OF 2|ASSORTED 01|TERRY TOWEL",
        "extra_attributes": {"Color": "Assorted 01", "BS Size": "40x60(2pc)"},
    })

    result = foparser.match_and_normalize(
        conn, amdb, 1, parsed_rows[0],
        ["brand", "size", "color", "product"],
        category="Bath",
    )
    conn.close()

    assert result["matched"] is True
    assert result["article_id"] is not None


def test_annotate_item_issues_unmatched_with_hint(tmp_path):
    conn = _make_am_conn(tmp_path)
    amdb.create_category(conn, 1, "Bed", ["brand", "size"], is_confirmed=True)
    amdb.upsert_article(conn, 1, {
        "category": "Bed", "product_type": "Sheet Sets", "brand": "Cardinal", "size": "SB BS",
        "mrp": 500, "ptr": 300, "ex_mill_price": 250, "bale_pack_size": 10,
        "item_key": "Cardinal|SB BS", "extra_attributes": {},
    })
    item = {
        "matched": False,
        "is_clean_bale_multiple": True,
        "brand": "Blumen",
        "size": "SB BS",
        "product_type": "Sheet Sets",
        "item_key": "Blumen|SB BS",
        "final_piece_qty": 24,
        "bale_size_used": 10,
    }
    foparser.annotate_item_issues(
        conn, amdb, 1, item, ["brand", "size"], category="Bed",
        core_fields={"brand": "Blumen", "size": "SB BS", "product_type": "Sheet Sets"},
        extra_attributes={},
    )
    conn.close()

    assert item["has_issue"] is True
    assert "Not found in Article Master" in item["issue_summary"]
    assert item["field_comparisons"]
    assert item["recommended_action"]["code"] in {"add_to_article_master", "edit_file_or_master"}
    brand_cmp = next(c for c in item["field_comparisons"] if c["field"] == "brand")
    assert brand_cmp["status"] == "mismatch"


def test_annotate_item_issues_flagged_qty(tmp_path):
    conn = _make_am_conn(tmp_path)
    item = {
        "matched": True,
        "is_clean_bale_multiple": False,
        "brand": "Cardinal",
        "size": "SB BS",
        "product_type": "Sheet Sets",
        "final_piece_qty": 96,
        "bale_size_used": 18,
    }
    foparser.annotate_item_issues(conn, amdb, 1, item, ["brand", "size"], category="Bed")
    conn.close()

    assert "not a clean multiple" in (item["issue_summary"] or "")
    assert "96" in (item["issue_summary"] or "")


def test_blumen_typo_matches_bluemen_in_article_master(tmp_path):
    """Distributor file 'Bluemen' must match Article Master 'Blumen'."""
    conn = _make_am_conn(tmp_path)
    amdb.create_category(conn, 1, "Bed", ["brand", "TC", "size"], is_confirmed=True)
    amdb.upsert_article(conn, 1, {
        "category": "Bed", "product_type": "Sheet Sets", "brand": "Blumen", "size": "DB BS",
        "mrp": 799, "ptr": 450, "ex_mill_price": 400, "bale_pack_size": 5,
        "item_key": "BLUMEN|104 (ONE IN A DENT)|DB BS",
        "extra_attributes": {"TC": "104 (ONE IN A DENT)"},
    })

    core = {"brand": "Blumen", "size": "DB BS", "product_type": "Sheet Sets"}
    extra = {}
    article = amdb.resolve_article_match(conn, 1, "Bed", core, extra, ["brand", "TC", "size"])
    conn.close()

    assert article is not None
    assert article["brand"] == "Blumen"
    assert article["size"] == "DB BS"


def test_tc_one_in_a_dent_matches_article_master(tmp_path):
    """File TC='104' must match AM row stored as '104 (ONE IN A DENT)'."""
    conn = _make_am_conn(tmp_path)
    amdb.create_category(conn, 1, "Bed", ["brand", "TC", "size"], is_confirmed=True)
    amdb.upsert_article(conn, 1, {
        "category": "Bed", "product_type": "Bedsheet", "brand": "ASTER", "size": "DB BS",
        "mrp": 999, "ptr": 450, "ex_mill_price": 400, "bale_pack_size": 12,
        "item_key": "ASTER|104 (ONE IN A DENT)|DB BS",
        "extra_attributes": {"TC": "104 (ONE IN A DENT)"},
    })

    core = {"brand": "ASTER", "size": "DB BS", "product_type": "Bedsheet", "bale_pack_size": 12}
    extra = {"TC": "104"}
    key_fields = ["brand", "TC", "size"]
    assert amparser.build_item_key(core, extra, key_fields) == "ASTER|104|DB BS"

    parsed_row = {"core_fields": core, "extra_attributes": extra, "raw_qty_value": 24}
    result = foparser.match_and_normalize(conn, amdb, 1, parsed_row, key_fields, category="Bed")
    conn.close()
    assert result["matched"] is True
    assert result["mrp"] == 999


def test_noise_qty_headers_ignored(tmp_path):
    """Qnty Per Color / Qnty pre Design must never win as order qty."""
    header = [
        "Brand", "Size", "Product", "MRP", "PTR", "Ex-Mill", "Bale Size",
        "Qnty Per Color", "Qnty pre Design", "Qnty",
    ]
    rows = [
        ["ASTER", "DB BS", "Bedsheet", 999, 450, 400, 18, 750, 2250, 108],
        ["ASTER", "SB BS", "Bedsheet", 799, 400, 350, 24, 250, 750, 96],
    ]
    path = _write_workbook(tmp_path, header, rows)
    header_row, col_mapping, valid_rows = _load_valid_rows(path)
    detection = foparser.detect_quantity_column(header_row, col_mapping, "Bed", valid_rows)
    assert detection["status"] == "ok"
    assert detection["column_label"] == "Qnty"


def test_revised_order_preferred_over_qnty(tmp_path):
    """kag.xlsx style: Revised Order is final qty even when Qnty also has values."""
    header = [
        "Brand", "Size", "Product", "MRP", "PTR", "Ex-Mill", "Bale Size",
        "Qnty", "Revised Order",
    ]
    rows = [
        ["ASTER", "DB BS", "Bedsheet", 999, 450, 400, 18, 108, 216],
        ["BLUMEN", "SB BS", "Bedsheet", 799, 400, 350, 24, 54, None],  # no revised = skip
        ["CARDINAL", "DB BS", "Bedsheet", 900, 420, 380, 18, 54, 108],
    ]
    path = _write_workbook(tmp_path, header, rows)
    header_row, col_mapping, valid_rows = _load_valid_rows(path)
    detection = foparser.detect_quantity_column(header_row, col_mapping, "Bed", valid_rows)
    assert detection["status"] == "ok"
    assert detection["column_label"] == "Revised Order"
    items = foparser.build_filled_order_rows(
        valid_rows, header_row, col_mapping, detection["column_index"],
    )
    assert [it["raw_qty_value"] for it in items] == [216, 108]


def test_order_in_pcs_and_bales_mismatch_highlighted(tmp_path):
    """Both Qty + Bales: focus Qty; mismatch sheet bales vs Qty/BaleSize — no silent fix."""
    header = [
        "Brand", "Size", "Product", "MRP", "PTR", "Ex-Mill", "Bale Size",
        "Order In Bales", "Order In Pc's",
    ]
    rows = [
        # qty 1728 / bs 18 = 96 expected bales, sheet says 8 -> mismatch
        ["ASTER", "DB BS", "Bedsheet", 999, 450, 625, 18, 8, 1728],
        # qty 288 / bs 24 = 12, sheet says 12 -> OK
        ["BLUMEN", "SB BS", "Bedsheet", 799, 400, 451, 24, 12, 288],
    ]
    path = _write_workbook(tmp_path, header, rows)
    header_row, col_mapping, valid_rows = _load_valid_rows(path)
    detection = foparser.detect_quantity_column(header_row, col_mapping, "Bed", valid_rows)
    assert detection["column_label"] == "Order In Pc's"
    bales = foparser.detect_bales_column(header_row, col_mapping, detection["column_index"])
    assert bales is not None
    assert "bales" in bales["column_label"].lower()

    parsed = foparser.build_filled_order_rows(
        valid_rows, header_row, col_mapping, detection["column_index"],
        bales_col_idx=bales["column_index"],
    )
    assert len(parsed) == 2

    r0 = foparser.apply_qty_bales_value_rules(
        raw_qty=parsed[0]["raw_qty_value"],
        sheet_bales=parsed[0]["sheet_bales"],
        bale_size=18,
        ex_mill=625,
        qty_column_label=detection["column_label"],
        category="Bed",
    )
    assert r0["final_piece_qty"] == 1728
    assert r0["bale_qty_mismatch"] is True
    assert r0["expected_bales"] == 96
    assert r0["sheet_bales"] == 8
    assert r0["line_value"] == 1728 * 625

    r1 = foparser.apply_qty_bales_value_rules(
        raw_qty=parsed[1]["raw_qty_value"],
        sheet_bales=parsed[1]["sheet_bales"],
        bale_size=24,
        ex_mill=451,
        qty_column_label=detection["column_label"],
        category="Bed",
    )
    assert r1["final_piece_qty"] == 288
    assert r1["bale_qty_mismatch"] is False
    assert r1["line_value"] == 288 * 451


def test_only_bales_computes_qty(tmp_path):
    """Only bales column -> Qty = Bales × Bale Size; Value = Qty × ExMill."""
    header = [
        "Brand", "Size", "Product", "MRP", "PTR", "Ex-Mill", "Bale Size", "Order In Bales",
    ]
    rows = [["ASTER", "DB BS", "Bedsheet", 999, 450, 625, 18, 8]]
    path = _write_workbook(tmp_path, header, rows)
    header_row, col_mapping, valid_rows = _load_valid_rows(path)
    detection = foparser.detect_quantity_column(header_row, col_mapping, "Bed", valid_rows)
    assert "bales" in detection["column_label"].lower()
    parsed = foparser.build_filled_order_rows(
        valid_rows, header_row, col_mapping, detection["column_index"],
    )
    resolved = foparser.apply_qty_bales_value_rules(
        raw_qty=parsed[0]["raw_qty_value"],
        sheet_bales=None,
        bale_size=18,
        ex_mill=625,
        qty_column_label=detection["column_label"],
        category="Bed",
    )
    assert resolved["detected_unit"] == "bales"
    assert resolved["final_piece_qty"] == 144
    assert resolved["line_value"] == 144 * 625
    assert resolved["bale_qty_mismatch"] is False


def test_shifted_qty_does_not_use_value_column(tmp_path):
    """Empty Qnty must not pick Value column as shifted qty."""
    header = [
        "Brand", "Size", "Product", "MRP", "PTR", "Ex-Mill", "Bale Size", "Qnty", "Value",
    ]
    rows = [
        ["ASTER", "DB BS", "Bedsheet", 999, 450, 400, 18, None, 90070],
        ["BLUMEN", "SB BS", "Bedsheet", 799, 400, 350, 24, 144, 65003],
    ]
    path = _write_workbook(tmp_path, header, rows)
    header_row, col_mapping, valid_rows = _load_valid_rows(path)
    detection = foparser.detect_quantity_column(header_row, col_mapping, "Bed", valid_rows)
    assert detection["status"] == "ok"
    assert detection["column_label"] == "Qnty"
    items = foparser.build_filled_order_rows(
        valid_rows, header_row, col_mapping, detection["column_index"],
    )
    assert [it["raw_qty_value"] for it in items] == [144]


def test_detect_category_sheet_sets_any_distributor(tmp_path):
    """Product 'Sheet Sets' → Bed for any distributor (not filename-specific)."""
    header = ["Brand", "Size", "Product", "Ex-Mill", "Bale Size", "Qty"]
    rows = [
        ["Aster", "DB BS", "Sheet Sets", 625, 18, 216],
        ["Cardinal", "SB BS", "Sheet Sets", 536, 24, 144],
    ]
    path = _write_workbook(tmp_path, header, rows, filename="ANY_DIST.xlsx")
    assert foparser.detect_category_from_order_file(path, filename="ANY_DIST.xlsx") == "Bed"


def test_detect_category_falls_back_to_size_codes(tmp_path):
    """When Product is blank, size codes like DB BS still detect Bed."""
    header = ["Brand", "Size", "Product", "Ex-Mill", "Bale Size", "Qty"]
    rows = [
        ["Aster", "DB BS", None, 625, 18, 108],
        ["Blumen", "KS BS", "", 1100, 12, 72],
    ]
    path = _write_workbook(tmp_path, header, rows, filename="mystery.xlsx")
    assert foparser.detect_category_from_order_file(path, filename="mystery.xlsx") == "Bed"


SAIN_XLS = Path(r"G:\My Drive\2026-2027\Oder Management\AW26 order\Bedsheet\SAIN.xls")
BALAJI_XLSX = Path(r"G:\My Drive\2026-2027\Oder Management\AW26 order\Bedsheet\BALAJI.xlsx")


@pytest.mark.skipif(not SAIN_XLS.exists(), reason="SAIN.xls not on G: drive")
def test_sain_xls_detects_bed_with_xlrd():
    """Legacy .xls (any distributor) must detect Bed once xlrd is installed."""
    import xlrd  # noqa: F401

    assert foparser.detect_category_from_order_file(SAIN_XLS, filename="SAIN.xls") == "Bed"


def test_monthly_split_total_auto_selected_as_qty(tmp_path):
    """Any month columns + TOTAL → TOTAL is Qty (not July/Aug-specific)."""
    header = [
        "Brand", "Size", "Product", "Ex-Mill", "Bale Size",
        "SEP", "OCT", "NOV", "TOTAL",
    ]
    rows = [
        ["Aster", "DB BS", "Sheet Sets", 625, 18, 100, 200, 50, 350],
        ["Epigram", "KS BS", "Sheet Sets", 1100, 12, 0, 120, 0, 120],
        ["Thyme", "KS BS", "Sheet Sets", 900, 12, 40, 40, 40, 120],
    ]
    path = _write_workbook(tmp_path, header, rows, filename="ANY_MONTHS.xlsx")
    header_row, col_mapping, valid_rows = _load_valid_rows(path)
    detection = foparser.detect_quantity_column(
        header_row, col_mapping, "Bed", valid_rows,
    )
    assert detection["status"] == "ok"
    assert detection["column_label"] == "TOTAL"
    assert "SEP" in detection.get("auto_selected_reason", "") or "Monthly" in detection.get("auto_selected_reason", "")
    items = foparser.build_filled_order_rows(
        valid_rows, header_row, col_mapping, detection["column_index"],
    )
    assert sum(it["raw_qty_value"] for it in items) == 350 + 120 + 120


def test_delivery_window_columns_are_not_qty(tmp_path):
    """'Aug - Sep Delivery' must never steal Qty from a real TOTAL column."""
    header = [
        "Brand", "Size", "Product", "Ex-Mill", "Bale Size",
        "Aug - Sep Delivery", "Sep - Oct Delivery", "JULY", "AUG", "TOTAL",
    ]
    rows = [
        ["Aster", "DB BS", "Sheet Sets", 625, 18, 6, 6, 216, 216, 432],
        ["Cardinal", "KS BS", "Sheet Sets", 1100, 12, 6, None, 72, 72, 144],
    ]
    path = _write_workbook(tmp_path, header, rows, filename="delivery_noise.xlsx")
    header_row, col_mapping, valid_rows = _load_valid_rows(path)
    detection = foparser.detect_quantity_column(
        header_row, col_mapping, "Bed", valid_rows,
    )
    assert detection["status"] == "ok"
    assert detection["column_label"] == "TOTAL"


@pytest.mark.skipif(not BALAJI_XLSX.exists(), reason="BALAJI.xlsx not on G: drive")
def test_balaji_xlsx_auto_selects_total_qty():
    wb = foparser.parse_filled_order_workbook(BALAJI_XLSX, "Bed")
    assert wb["status"] == "ok"
    assert wb["quantity_column_used"] == "TOTAL"
    rows = wb["parsed_rows"]
    assert sum(float(r["raw_qty_value"] or 0) for r in rows) == 5316


def test_so_pack_export_rejected_with_clear_guidance(tmp_path):
    """SO Pack download must not be accepted as a distributor filled order."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consolidated"
    ws.append(["SO-wise Product Consolidation"])
    ws.append(["SO Number", "Product Name", "Total Qty"])
    ws.append(["102876117", "ASTER 1+2 DB SET", 216])
    wb.create_sheet("SO Summary")
    wb.create_sheet("Line Item Detail")
    path = tmp_path / "CHOICE CORNER BOMBAY DYEING_SO_Pack.xlsx"
    wb.save(path)

    assert foparser.looks_like_so_pack_workbook(path) is True
    with pytest.raises(ValueError, match="SO Pack export"):
        foparser.detect_category_from_order_file(path, filename=path.name)
    with pytest.raises(ValueError, match="Brand \\+ Size"):
        foparser.parse_filled_order_workbook(path, "Bed")


def test_bath_linen_special_sheet_maps_quality_exmill_and_total_qty(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["QUALITY", "SIZE", "GSM", "NO.SHADES", "SKU", "DELHI", "", "MRP", "EXMILL", "value at exmill"])
    ws.append(["", "", "", "", "", "Per Color", "total qty", "", "", ""])
    ws.append(["FLORA", "40*60", 400, 9, 1, 300, 2700, 142, 77, 207900])
    ws.append(["TULIP", "75*150", 450, 11, 1, 255, 2805, 825, 450, 1262250])
    path = tmp_path / "BND Bath linen special order.xlsx"
    wb.save(path)

    assert foparser.detect_category_from_order_file(path, filename=path.name) == "Bath"
    parsed = foparser.parse_filled_order_workbook(path, "Bath")
    assert parsed["status"] == "ok"
    assert amparser._norm(parsed["quantity_column_used"]) == "total qty"
    rows = parsed["parsed_rows"]
    assert len(rows) == 2
    flora = next(r for r in rows if str(r["core_fields"].get("brand")).upper() == "FLORA")
    assert float(flora["raw_qty_value"]) == 2700
    assert float(flora["core_fields"]["ex_mill_price"]) == 77


def test_addon_filename_detected():
    assert foparser.looks_like_addon_order_filename("BND Bath linen special order.xlsx")
    assert foparser.looks_like_addon_order_filename("bernina additional order.xlsx")
    assert not foparser.looks_like_addon_order_filename("bernina_bed.xlsx")
