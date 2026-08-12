"""Hasta Lavista — multi-distributor orders with corruption-safe qty/value."""
from __future__ import annotations

from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SRC = Path(r"c:\Users\princ\OneDrive\Desktop\hasta lavista.xlsx")
OUT = Path(r"E:\test files\Hasta_Lavista_Distributor_Orders_v2.xlsx")
OUT_REPO = Path(r"E:\centralized-db-system\Output\Hasta_Lavista_Distributor_Orders.xlsx")

DISTRIBUTORS = [
    {"name": "BND", "bales_col": 22, "qty_col": 23, "value_col": 24},
    {"name": "KAG", "bales_col": None, "qty_col": 25, "value_col": 26},
    {"name": "Ptj", "bales_col": 27, "qty_col": None, "value_col": 28},
    {"name": "Choice", "bales_col": 30, "qty_col": 29, "value_col": None},
]


def _num(v):
    try:
        return None if v is None else float(v)
    except (TypeError, ValueError):
        return None


def _pretty(n, money=False):
    if n is None:
        return None
    n = float(n)
    if money:
        return round(n, 2)
    if abs(n - round(n)) < 1e-9:
        return int(round(n))
    return round(n, 2)


def _qty_sane(qty, bales, bale_size) -> bool:
    if qty is None:
        return False
    if qty < 0:
        return False
    # obvious formula blow-ups
    if qty > 100_000:
        return False
    if bales is not None and bale_size and bales > 0:
        expected = bales * bale_size
        # if both present, they should roughly agree (BND-style sometimes won't —
        # allow wider band, but reject 100x+ drift)
        if expected > 0 and qty > expected * 50:
            return False
    return True


def _bales_sane(bales) -> bool:
    return bales is not None and 0 < bales <= 5_000


def resolve_line(ex_mill, bale_size, raw_bales, raw_qty, raw_val):
    """
    Resolve qty/bales/value with recovery when sheet formulas are corrupted.
    Returns dict or None if no usable order.
    """
    notes = []
    bales = raw_bales if _bales_sane(raw_bales) else None
    qty = raw_qty if _qty_sane(raw_qty, bales, bale_size) else None

    if qty is None and raw_qty is not None:
        notes.append("qty_col_corrupt")
    if bales is None and raw_bales is not None and raw_bales != 0:
        notes.append("bales_col_corrupt")

    if qty is None and bales is not None and bale_size:
        qty = bales * bale_size
        notes.append("qty=bales×bale_size")
    if bales is None and qty is not None and bale_size:
        bales = qty / bale_size
        notes.append("bales=qty÷bale_size")

    if (qty is None or qty == 0) and (bales is None or bales == 0):
        return None

    expected_val = (ex_mill * qty) if (ex_mill is not None and qty is not None) else None
    value = raw_val
    value_source = "sheet"
    if expected_val is not None:
        if value is None or value > 1e10 or (
            expected_val > 0 and abs(value - expected_val) / expected_val > 0.05 and abs(value - expected_val) > 1000
        ):
            value = expected_val
            value_source = "ExMill×Qty"
            if raw_val is not None:
                notes.append("value_recomputed")
    elif value is None:
        value = 0.0
        value_source = "none"

    return {
        "Qty": qty or 0.0,
        "Bales": bales or 0.0,
        "Value": value or 0.0,
        "Value source": value_source,
        "Notes": ", ".join(notes) if notes else "ok",
    }


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT_REPO.parent.mkdir(parents=True, exist_ok=True)

    wb_src = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb_src["excel"]

    articles = []
    for r in range(3, (ws.max_row or 0) + 1):
        brand = ws.cell(r, 1).value
        if not brand or not str(brand).strip():
            continue
        articles.append(
            {
                "row": r,
                "Brand": str(brand).strip(),
                "TC": ws.cell(r, 2).value,
                "Size": str(ws.cell(r, 3).value or "").strip(),
                "Units": ws.cell(r, 4).value,
                "Product": ws.cell(r, 8).value,
                "Bale Size": _num(ws.cell(r, 10).value),
                "ExMill": _num(ws.cell(r, 20).value),
            }
        )

    by_dist = {}
    for dist in DISTRIBUTORS:
        name = dist["name"]
        lines = []
        for a in articles:
            r = a["row"]
            raw_bales = _num(ws.cell(r, dist["bales_col"]).value) if dist["bales_col"] else None
            raw_qty = _num(ws.cell(r, dist["qty_col"]).value) if dist["qty_col"] else None
            raw_val = _num(ws.cell(r, dist["value_col"]).value) if dist["value_col"] else None
            resolved = resolve_line(a["ExMill"], a["Bale Size"], raw_bales, raw_qty, raw_val)
            if not resolved:
                continue
            lines.append(
                {
                    **{k: a[k] for k in ("Brand", "TC", "Size", "Units", "Product", "Bale Size", "ExMill")},
                    **resolved,
                }
            )
        by_dist[name] = lines

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    brand_fill = PatternFill("solid", fgColor="E8F0FE")
    warn_fill = PatternFill("solid", fgColor="FCE4D6")
    thin = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Distributor Overview"
    headers = ["Distributor", "Lines", "Club Qty", "Club Bales", "Club Value", "Recovered lines", "Notes"]
    for i, h in enumerate(headers, 1):
        c = ws1.cell(1, i, h)
        c.fill = header_fill
        c.font = header_font

    notes_map = {
        "BND": "Clean section (No of Bales + Qty + AWD Value)",
        "KAG": "Qty/Value cols often formula-corrupt → recover via bales if any / ExMill×Qty",
        "Ptj": "Has bales; qty derived = bales×bale_size; value often recomputed",
        "Choice": "Qty often corrupt; bales col usable → qty=bales×bale_size; value=ExMill×Qty",
    }

    r_i = 2
    for name in [d["name"] for d in DISTRIBUTORS]:
        lines = by_dist[name]
        qty = sum(x["Qty"] for x in lines)
        bales = sum(x["Bales"] for x in lines)
        value = sum(x["Value"] for x in lines)
        recovered = sum(1 for x in lines if x["Notes"] != "ok")
        vals = [
            name, len(lines), _pretty(qty), _pretty(bales), _pretty(value, True),
            recovered, notes_map.get(name, ""),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(r_i, c, v)
            cell.border = thin
            if recovered and c == 1:
                cell.fill = warn_fill
            if c == 5:
                cell.number_format = "#,##0.00"
        r_i += 1

    ws1.freeze_panes = "A2"
    for i, w in enumerate([12, 8, 12, 12, 16, 14, 70], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    for name in [d["name"] for d in DISTRIBUTORS]:
        lines = by_dist[name]
        brand_tot = defaultdict(lambda: {"Lines": 0, "Qty": 0.0, "Bales": 0.0, "Value": 0.0})
        for x in lines:
            brand_tot[x["Brand"]]["Lines"] += 1
            brand_tot[x["Brand"]]["Qty"] += x["Qty"]
            brand_tot[x["Brand"]]["Bales"] += x["Bales"]
            brand_tot[x["Brand"]]["Value"] += x["Value"]
        brand_names = sorted(brand_tot.keys(), key=lambda s: s.upper())

        ws_b = wb.create_sheet(f"{name} Brand")
        for i, h in enumerate(["Brand", "Lines", "Qty", "Bales", "Value"], 1):
            cell = ws_b.cell(1, i, h)
            cell.fill = header_fill
            cell.font = header_font
        r_i = 2
        tq = tb = tv = 0.0
        for b in brand_names:
            t = brand_tot[b]
            tq += t["Qty"]
            tb += t["Bales"]
            tv += t["Value"]
            for c, v in enumerate(
                [b, t["Lines"], _pretty(t["Qty"]), _pretty(t["Bales"]), _pretty(t["Value"], True)], 1
            ):
                cell = ws_b.cell(r_i, c, v)
                cell.border = thin
                if c == 5:
                    cell.number_format = "#,##0.00"
            r_i += 1
        for c, v in enumerate(["TOTAL", len(lines), _pretty(tq), _pretty(tb), _pretty(tv, True)], 1):
            cell = ws_b.cell(r_i, c, v)
            cell.fill = total_fill
            cell.font = Font(bold=True)
            cell.border = thin
            if c == 5:
                cell.number_format = "#,##0.00"
        for i, w in enumerate([24, 8, 12, 12, 16], 1):
            ws_b.column_dimensions[get_column_letter(i)].width = w

        ws_l = wb.create_sheet(f"{name} Lines")
        hdr = [
            "Brand", "TC", "Size", "Product", "Bale Size", "ExMill",
            "Qty", "Bales", "Value", "Value source", "Notes",
        ]
        for i, h in enumerate(hdr, 1):
            cell = ws_l.cell(1, i, h)
            cell.fill = header_fill
            cell.font = header_font
        r_i = 2
        for b in brand_names:
            t = brand_tot[b]
            for c, v in enumerate(
                [b, "", "", "", "", "", _pretty(t["Qty"]), _pretty(t["Bales"]), _pretty(t["Value"], True), "", ""],
                1,
            ):
                cell = ws_l.cell(r_i, c, v)
                cell.fill = brand_fill
                cell.font = Font(bold=True)
                cell.border = thin
                if c == 9:
                    cell.number_format = "#,##0.00"
            r_i += 1
            for x in [ln for ln in lines if ln["Brand"] == b]:
                vals = [
                    x["Brand"], x["TC"], x["Size"], x["Product"],
                    _pretty(x["Bale Size"]), _pretty(x["ExMill"], True),
                    _pretty(x["Qty"]), _pretty(x["Bales"]), _pretty(x["Value"], True),
                    x["Value source"], x["Notes"],
                ]
                for c, v in enumerate(vals, 1):
                    cell = ws_l.cell(r_i, c, v)
                    cell.border = thin
                    if x["Notes"] != "ok":
                        cell.fill = warn_fill
                    if c in (6, 9):
                        cell.number_format = "#,##0.00"
                r_i += 1
        for c, v in enumerate(
            ["GRAND TOTAL", "", "", "", "", "", _pretty(tq), _pretty(tb), _pretty(tv, True), "", ""], 1
        ):
            cell = ws_l.cell(r_i, c, v)
            cell.fill = total_fill
            cell.font = Font(bold=True)
            cell.border = thin
            if c == 9:
                cell.number_format = "#,##0.00"
        for i, w in enumerate([20, 8, 10, 14, 10, 10, 10, 10, 14, 12, 28], 1):
            ws_l.column_dimensions[get_column_letter(i)].width = w

    ws_n = wb.create_sheet("How read")
    for i, line in enumerate(
        [
            "Source: hasta lavista.xlsx",
            "Filename ≠ distributor. Distributors in header row: BND | KAG | Ptj | Choice (+ SUP label on last col).",
            "Teaching match: nicknames inside sheet → separate distributor orders from one workbook.",
            "",
            "Column map:",
            "  BND: No of Bales + Qty + AWD Value (clean — matches BND.xlsx style)",
            "  KAG: Qty + AWD Value (many cells formula-corrupt)",
            "  Ptj: No of Bales + AWD Value",
            "  Choice: Qty + No of Bales",
            "",
            "Recovery rules:",
            "  qty > 100000 → corrupt → if bales ok: qty = bales × bale_size",
            "  missing qty + bales ok → qty = bales × bale_size",
            "  missing bales + qty ok → bales = qty ÷ bale_size",
            "  value absurd / missing → ExMill × Qty",
            "Orange rows = recovered / recomputed.",
        ],
        1,
    ):
        ws_n[f"A{i}"] = line
    ws_n.column_dimensions["A"].width = 100

    wb.save(OUT)
    wb.save(OUT_REPO)
    print(f"wrote {OUT}")
    for name in [d["name"] for d in DISTRIBUTORS]:
        lines = by_dist[name]
        rec = sum(1 for x in lines if x["Notes"] != "ok")
        print(
            name,
            "lines", len(lines),
            "recovered", rec,
            "qty", _pretty(sum(x["Qty"] for x in lines)),
            "bales", _pretty(sum(x["Bales"] for x in lines)),
            "value", _pretty(sum(x["Value"] for x in lines), True),
        )


if __name__ == "__main__":
    main()
