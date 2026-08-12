"""Hasta Lavista preview — strict teaching rules (no fake fixes)."""
from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import subprocess

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SRC = r"c:\Users\princ\OneDrive\Desktop\hasta lavista.xlsx"
OUT = Path(r"E:\test files\Hasta_Lavista_Rules_v1.xlsx")

# LOCKED RULES:
# 1. Qty + Bales both given -> focus QTY; check bales == qty/bale_size;
#    match=OK else HIGHLIGHT (do not fake-fix typed bales).
# 2. Only Bales given -> auto Qty = Bales * Bale Size.
# 3. Value always from Qty: Value = Qty * ExMill.
#    If only-bales path, first make Qty then * ExMill.
# Note: Only Qty (no bales) -> keep Qty, do NOT invent bales; Value = Qty * ExMill.


def fnum(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def process(has_q, has_b, sheet_v, bs, ex):
    q = has_q
    b = has_b
    notes = []

    if q is not None and b is not None:
        expected_b = q / bs
        if abs(b - expected_b) < 0.01:
            bale_status = "OK"
        else:
            bale_status = "MISMATCH"
            notes.append(f"typed bales={b}, expected from qty={round(expected_b, 2)}")
    elif b is not None and q is None:
        q = b * bs
        bale_status = "ONLY_BALES"
        notes.append("qty auto = bales * bale_size")
    elif q is not None and b is None:
        bale_status = "NO_BALES_COL"
    else:
        return None

    value = q * ex if q is not None else None
    if sheet_v is not None and value is not None and abs(sheet_v - value) >= 1:
        value_status = "SHEET_VALUE_DIFFERS"
        notes.append(f"sheet value={round(sheet_v, 2)}, qty*exmill={round(value, 2)}")
    elif sheet_v is not None:
        value_status = "SHEET_MATCHES_QTY"
    else:
        value_status = "FROM_QTY"

    return {
        "qty": q,
        "bales": b,
        "value": value,
        "bale_status": bale_status,
        "value_status": value_status,
        "sheet_value": sheet_v,
        "notes": "; ".join(notes),
    }


def main():
    wb0 = load_workbook(SRC, data_only=True)
    ws0 = wb0.active

    lines = []
    for r in range(3, 200):
        brand = ws0.cell(r, 1).value
        if not brand:
            continue
        bs = fnum(ws0.cell(r, 10).value)
        ex = fnum(ws0.cell(r, 20).value)
        if not bs or not ex:
            continue
        lines.append(
            {
                "brand": str(brand).strip(),
                "tc": ws0.cell(r, 2).value,
                "size": str(ws0.cell(r, 3).value or ""),
                "product": ws0.cell(r, 8).value,
                "bs": bs,
                "ex": ex,
                "bnd_b": fnum(ws0.cell(r, 22).value),
                "bnd_q": fnum(ws0.cell(r, 23).value),
                "bnd_v_sheet": fnum(ws0.cell(r, 24).value),
                "kag_q": fnum(ws0.cell(r, 25).value),
                "kag_v_sheet": fnum(ws0.cell(r, 26).value),
                "ptj_b": fnum(ws0.cell(r, 27).value),
                "ptj_v_sheet": fnum(ws0.cell(r, 28).value),
                "ch_q": fnum(ws0.cell(r, 29).value),
                "sup_b": fnum(ws0.cell(r, 30).value),
            }
        )

    dists_cfg = [
        ("BND", lambda r: process(r["bnd_q"], r["bnd_b"], r["bnd_v_sheet"], r["bs"], r["ex"])),
        ("KAG", lambda r: process(r["kag_q"], None, r["kag_v_sheet"], r["bs"], r["ex"])),
        ("Ptj", lambda r: process(None, r["ptj_b"], r["ptj_v_sheet"], r["bs"], r["ex"])),
        ("Choice", lambda r: process(r["ch_q"], None, None, r["bs"], r["ex"])),
        ("SUP", lambda r: process(None, r["sup_b"], None, r["bs"], r["ex"])),
    ]

    results = {d: [] for d, _ in dists_cfg}
    for line in lines:
        for d, fn in dists_cfg:
            out_row = fn(line)
            if out_row is None:
                continue
            results[d].append({**line, **out_row})

    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", "0F766E")
    bad = PatternFill("solid", "FEE2E2")
    ok = PatternFill("solid", "DCFCE7")
    info = PatternFill("solid", "FEF3C7")
    nf = Font(bold=True)
    tfill = PatternFill("solid", "CCFBF1")

    sheet_fields = {
        "BND": "No of Bales + Qty + AWD Value",
        "KAG": "Qty + AWD Value",
        "Ptj": "No of Bales + AWD Value",
        "Choice": "Qty only",
        "SUP": "No of Bales only",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Rules Locked"
    ws.append(["Rule", "Meaning"])
    ws.append(
        [
            "1",
            "Qty + Bales both given -> focus QTY. Check bales match qty (bales == qty/bale_size). Match=OK, else HIGHLIGHT. Do not fake-fix bales.",
        ]
    )
    ws.append(["2", "Only Bales given -> auto Qty = Bales * Bale Size"])
    ws.append(
        [
            "3",
            "Value always from Qty: Value = Qty * ExMill. If only-bales path, first make Qty then * ExMill.",
        ]
    )
    ws.append(
        [
            "Note",
            "Only Qty (no bales) -> keep Qty, do NOT invent bales; Value = Qty * ExMill",
        ]
    )
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 110

    ws = wb.create_sheet("Overview")
    ws.append(
        [
            "Distributor",
            "Sheet fields",
            "Qty (focus/calc)",
            "Bales (as in sheet)",
            "Value (=Qty*ExMill)",
            "Bale mismatches",
            "Sheet value differs",
        ]
    )
    for cell in ws[1]:
        cell.font = hf
        cell.fill = hfill
        cell.border = thin

    for d, _ in dists_cfg:
        rows = results[d]
        q = sum(x["qty"] or 0 for x in rows)
        has_b = any(x["bales"] is not None for x in rows)
        b = sum(x["bales"] or 0 for x in rows) if has_b else None
        v = sum(x["value"] or 0 for x in rows)
        mm = sum(1 for x in rows if x["bale_status"] == "MISMATCH")
        vd = sum(1 for x in rows if x["value_status"] == "SHEET_VALUE_DIFFERS")
        ws.append(
            [
                d,
                sheet_fields[d],
                round(q, 2),
                (round(b, 2) if b is not None else "(no bales col)"),
                round(v, 2),
                mm,
                vd,
            ]
        )
    for col, w in zip("ABCDEFG", [12, 34, 16, 18, 18, 16, 18]):
        ws.column_dimensions[col].width = w

    ws = wb.create_sheet("Highlights")
    ws.append(
        [
            "Distributor",
            "Brand",
            "Size",
            "Issue",
            "Sheet bales",
            "Qty used",
            "Expected bales (=Qty/BS)",
            "Sheet value",
            "Value used (=Qty*Ex)",
            "Notes",
        ]
    )
    for cell in ws[1]:
        cell.font = hf
        cell.fill = hfill
        cell.border = thin

    for d, _ in dists_cfg:
        for x in results[d]:
            if x["bale_status"] == "MISMATCH" or x["value_status"] == "SHEET_VALUE_DIFFERS":
                exp_b = round((x["qty"] / x["bs"]), 2) if x["qty"] and x["bs"] else None
                ws.append(
                    [
                        d,
                        x["brand"],
                        x["size"],
                        x["bale_status"]
                        if x["bale_status"] == "MISMATCH"
                        else x["value_status"],
                        x["bales"],
                        round(x["qty"], 2) if x["qty"] is not None else None,
                        exp_b,
                        round(x["sheet_value"], 2) if x["sheet_value"] is not None else None,
                        round(x["value"], 2) if x["value"] is not None else None,
                        x["notes"],
                    ]
                )
                for c in range(1, 11):
                    ws.cell(ws.max_row, c).fill = bad
    for i, w in enumerate([12, 22, 10, 22, 12, 12, 20, 14, 18, 50], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for d, _ in dists_cfg:
        ws = wb.create_sheet(f"{d} Brandwise")
        ws.append(["Distributor", d])
        ws.append(["Sheet fields", sheet_fields[d]])
        ws.append(["Brand", "Qty", "Bales (sheet)", "Value (=Qty*ExMill)", "Bale check"])
        for cell in ws[3]:
            cell.font = hf
            cell.fill = hfill
            cell.border = thin
        club = defaultdict(lambda: {"q": 0.0, "b": 0.0, "v": 0.0, "mm": 0, "has_b": False})
        for x in results[d]:
            club[x["brand"]]["q"] += x["qty"] or 0
            if x["bales"] is not None:
                club[x["brand"]]["b"] += x["bales"]
                club[x["brand"]]["has_b"] = True
            club[x["brand"]]["v"] += x["value"] or 0
            if x["bale_status"] == "MISMATCH":
                club[x["brand"]]["mm"] += 1
        tq = tb = tv = 0.0
        any_b = False
        for brand in sorted(club):
            c = club[brand]
            bale_cell = round(c["b"], 2) if c["has_b"] else ""
            check = "MISMATCH" if c["mm"] else ("OK" if c["has_b"] else "")
            ws.append([brand, round(c["q"], 2), bale_cell, round(c["v"], 2), check])
            if check == "MISMATCH":
                for col in range(1, 6):
                    ws.cell(ws.max_row, col).fill = bad
            tq += c["q"]
            tv += c["v"]
            if c["has_b"]:
                tb += c["b"]
                any_b = True
        ws.append(["TOTAL", round(tq, 2), (round(tb, 2) if any_b else ""), round(tv, 2), ""])
        for c in ws[ws.max_row]:
            c.font = nf
            c.fill = tfill
        for col, w in zip("ABCDE", [28, 12, 14, 18, 12]):
            ws.column_dimensions[col].width = w

    ws = wb.create_sheet("All Lines")
    ws.append(
        [
            "Distributor",
            "Brand",
            "TC",
            "Size",
            "BaleSize",
            "ExMill",
            "Qty",
            "Bales(sheet)",
            "Value(=Qty*Ex)",
            "Bale check",
            "Value check",
            "Notes",
        ]
    )
    for cell in ws[1]:
        cell.font = hf
        cell.fill = hfill
        cell.border = thin
    for d, _ in dists_cfg:
        for x in results[d]:
            ws.append(
                [
                    d,
                    x["brand"],
                    x["tc"],
                    x["size"],
                    x["bs"],
                    round(x["ex"], 2),
                    round(x["qty"], 2) if x["qty"] is not None else None,
                    x["bales"] if x["bales"] is not None else "",
                    round(x["value"], 2) if x["value"] is not None else None,
                    x["bale_status"],
                    x["value_status"],
                    x["notes"],
                ]
            )
            if x["bale_status"] == "MISMATCH":
                ws.cell(ws.max_row, 10).fill = bad
                ws.cell(ws.max_row, 8).fill = bad
            elif x["bale_status"] == "OK":
                ws.cell(ws.max_row, 10).fill = ok
            elif x["bale_status"] == "ONLY_BALES":
                ws.cell(ws.max_row, 10).fill = info
            if x["value_status"] == "SHEET_VALUE_DIFFERS":
                ws.cell(ws.max_row, 11).fill = bad
    for i, w in enumerate([12, 22, 6, 10, 10, 10, 10, 12, 14, 14, 18, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print("Saved", OUT)
    print("=== TOTALS (rules applied, no fake bale fixes) ===")
    for d, _ in dists_cfg:
        rows = results[d]
        q = sum(x["qty"] or 0 for x in rows)
        has_b = any(x["bales"] is not None for x in rows)
        b = sum(x["bales"] or 0 for x in rows) if has_b else None
        v = sum(x["value"] or 0 for x in rows)
        mm = sum(1 for x in rows if x["bale_status"] == "MISMATCH")
        print(
            d,
            "qty",
            round(q, 2),
            "bales",
            (round(b, 2) if b is not None else "n/a"),
            "value",
            round(v, 2),
            "bale_mismatches",
            mm,
        )

    subprocess.Popen(["cmd", "/c", "start", "", str(OUT)], shell=False)


if __name__ == "__main__":
    main()
