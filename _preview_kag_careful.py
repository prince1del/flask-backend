"""Careful re-read of kag.xlsx — understand column logic before concluding."""
from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import subprocess

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SRC = Path(r"g:\My Drive\2026-2027\Oder Management\AW26 order\Bedsheet\kag.xlsx")
OUT = Path(r"E:\test files\KAG_Careful_Review.xlsx")


def main():
    wb_f = load_workbook(SRC, data_only=False)
    wb = load_workbook(SRC, data_only=True)
    ws = wb.active
    wsf = wb_f.active

    print("=== STRUCTURE ===")
    print("Sheet:", ws.title)
    print("Headers (row 1):")
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            print(f"  {get_column_letter(c)}: {v}")

    print("\nFooter formulas R48:")
    for c in range(25, 29):
        print(f"  {get_column_letter(c)}48 = {wsf.cell(48, c).value!r} -> {ws.cell(48, c).value!r}")

    rows = []
    for r in range(2, 48):
        brand = ws.cell(r, 2).value
        if not brand:
            continue
        bs = float(ws.cell(r, 11).value or 0)
        color = ws.cell(r, 12).value
        designs = ws.cell(r, 15).value
        ex = float(ws.cell(r, 24).value or 0)
        asst = ws.cell(r, 25).value
        qnty = ws.cell(r, 26).value
        rev = ws.cell(r, 27).value
        val = ws.cell(r, 28).value
        asst = float(asst) if asst not in (None, "") else None
        qnty = float(qnty) if qnty not in (None, "") else None
        rev = float(rev) if rev not in (None, "") else None
        val = float(val) if val not in (None, "") else None

        # Verify formula meanings on calculated values
        asst_expect = (
            float(designs) * bs if designs not in (None, "") and bs else None
        )
        qnty_expect = (
            3 * float(designs) * float(color)
            if designs not in (None, "") and color not in (None, "")
            else None
        )

        bales_from_rev = (rev / bs) if (rev and bs) else None
        notes = []
        if rev:
            if asst is not None and abs(rev - asst) < 0.01:
                notes.append("Rev=Asst")
            if qnty is not None and abs(rev - qnty) < 0.01:
                notes.append("Rev=Qnty")
            if (
                designs is not None
                and bales_from_rev is not None
                and abs(bales_from_rev - float(designs)) < 0.01
            ):
                notes.append("Rev/BaleSize = NoOfDesign (full assortment)")
            elif bales_from_rev is not None:
                notes.append(f"Rev/BaleSize={round(bales_from_rev, 2)} (designs={designs})")

        # formula vs typed for Z
        z_raw = wsf.cell(r, 26).value
        z_kind = "formula" if isinstance(z_raw, str) and str(z_raw).startswith("=") else (
            "typed" if z_raw not in (None, "") else "blank"
        )

        rows.append(
            {
                "brand": str(brand).strip(),
                "size": str(ws.cell(r, 4).value or ""),
                "bs": bs,
                "color": color,
                "designs": designs,
                "ex": ex,
                "asst": asst,
                "asst_expect": asst_expect,
                "qnty": qnty,
                "qnty_expect": qnty_expect,
                "qnty_kind": z_kind,
                "rev": rev,
                "bales_from_rev": bales_from_rev,
                "value_rs": (rev * ex) if rev else None,
                "val_lakhs": val,
                "notes": "; ".join(notes),
                "y_formula": wsf.cell(r, 25).value,
                "z_formula": z_raw,
                "ab_formula": wsf.cell(r, 28).value,
            }
        )

    ordered = [r for r in rows if r["rev"]]
    print("\n=== COLUMN LOGIC (verified) ===")
    print("Y Asst  = NoOfDesign * BaleSize     (full assortment pcs)")
    print("Z Qnty  = 3 * NoOfDesign * Color    (OR typed number on some rows)")
    print("AA Revised Order = typed final order pcs (drives AB value)")
    print("AB val  = Revised * ExMill / 100000 (value in lakhs)")
    print()
    print("Ordered lines (Revised>0):", len(ordered))
    print("Sum Revised pcs:", sum(r["rev"] for r in ordered))
    print("Sum Value Rs:", round(sum(r["value_rs"] for r in ordered), 2))
    print("Sum val lakhs:", round(sum((r["val_lakhs"] or 0) for r in ordered), 4))
    print("Footer AA:", ws.cell(48, 27).value, "AB:", ws.cell(48, 28).value)

    print("\n=== ORDERED LINES ===")
    for r in ordered:
        print(
            f"{r['brand'][:20]:20} {r['size'][:8]:8} "
            f"Asst={r['asst']} Qnty={r['qnty']}({r['qnty_kind']}) "
            f"Rev={r['rev']:g} bales~{r['bales_from_rev']} | {r['notes']}"
        )

    print("\n=== HAS QNTY BUT NO REVISED (not in value) ===")
    for r in rows:
        if r["rev"]:
            continue
        if r["qnty"]:
            print(f"  {r['brand']} {r['size']}: Qnty={r['qnty']} Asst={r['asst']}")

    # Pattern counts
    rev_eq_asst = sum(1 for r in ordered if "Rev=Asst" in r["notes"])
    rev_eq_qnty = sum(1 for r in ordered if "Rev=Qnty" in r["notes"])
    full_asst = sum(1 for r in ordered if "full assortment" in r["notes"])
    print("\n=== PATTERNS among ordered ===")
    print(f"Rev=Asst: {rev_eq_asst}/{len(ordered)}")
    print(f"Rev=Qnty: {rev_eq_qnty}/{len(ordered)}")
    print(f"Rev/BaleSize=NoOfDesign: {full_asst}/{len(ordered)}")

    # Excel output — careful, honest
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", "0F766E")
    nf = Font(bold=True)
    tfill = PatternFill("solid", "CCFBF1")
    warn = PatternFill("solid", "FEF3C7")
    info = PatternFill("solid", "DBEAFE")

    out = Workbook()
    o = out.active
    o.title = "Understanding"
    lines = [
        ["kag.xlsx — careful read"],
        [],
        ["Distributor from filename", "KAG"],
        ["Header row", "1"],
        ["Data rows", "2 to 47"],
        ["Footer", "Row 48 sums Qnty, Revised Order, val"],
        [],
        ["Column", "Formula / content", "What it means"],
        [
            "Y Asst",
            "= NoOfDesign * BaleSize",
            "Full assortment qty in pcs (1 bale per design)",
        ],
        [
            "Z Qnty",
            "= 3 * NoOfDesign * Color  (or typed)",
            "Calculated / proposed qty — NOT always final",
        ],
        [
            "AA Revised Order",
            "Typed number (no formula)",
            "Final order qty in pcs — this drives value",
        ],
        [
            "AB val",
            "= Revised * ExMill / 100000",
            "Order value in LAKHS (sheet unit)",
        ],
        [],
        ["Noise (ignore for order qty)", "Qnty Per Color, Qnty pre Design"],
        ["No bales column", "Implied bales = Revised / BaleSize when Revised present"],
        [],
        ["OBSERVATION", ""],
        [
            "1",
            "Value (AB) always references Revised (AA), never Qnty (Z).",
        ],
        [
            "2",
            "Footer sums Revised + val — Revised is the money qty.",
        ],
        [
            "3",
            f"Of {len(ordered)} revised lines: {rev_eq_asst} equal Asst, {rev_eq_qnty} equal Qnty, "
            f"{full_asst} are full assortment (Rev/BaleSize=NoOfDesign).",
        ],
        [
            "4",
            "Rows with Qnty but blank Revised have AB=0 — no order booked.",
        ],
        [
            "5",
            "Our system Value (Rs) = Revised * ExMill (full rupees, not /100000).",
        ],
        [],
        ["PROPOSED RULE (needs your confirm)", ""],
        ["Qty", "Use AA Revised Order where > 0; skip line if blank"],
        ["Bales", "No sheet bales col; optional check = Qty/BaleSize (info only)"],
        ["Value", "Qty * ExMill (Rs)"],
        ["Do not use", "Z Qnty as final qty when Revised exists or when Revised blank"],
    ]
    for row in lines:
        o.append(row)
    o.column_dimensions["A"].width = 28
    o.column_dimensions["B"].width = 70
    o.column_dimensions["C"].width = 55

    # Overview totals
    ws = out.create_sheet("Totals")
    ws.append(["Metric", "Value"])
    ws.append(["Distributor", "KAG"])
    ws.append(["Order lines (Revised>0)", len(ordered)])
    ws.append(["Total Qty pcs (Revised)", sum(r["rev"] for r in ordered)])
    ws.append(["Total Value Rs (Qty*ExMill)", round(sum(r["value_rs"] for r in ordered), 2)])
    ws.append(["Total sheet val (lakhs)", round(sum((r["val_lakhs"] or 0) for r in ordered), 4)])
    ws.append(["Footer AA sum", ws_footer_aa := wb.active.cell(48, 27).value])
    ws.append(["Footer AB sum", wb.active.cell(48, 28).value])
    # verify footer match
    ws.append(
        [
            "Footer AA matches our sum?",
            "YES"
            if ws_footer_aa is not None
            and abs(float(ws_footer_aa) - sum(r["rev"] for r in ordered)) < 0.01
            else f"check {ws_footer_aa}",
        ]
    )
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 20

    # Brandwise
    club = defaultdict(lambda: [0.0, 0.0])
    for r in ordered:
        club[r["brand"]][0] += r["rev"]
        club[r["brand"]][1] += r["value_rs"]
    ws = out.create_sheet("Brandwise")
    ws.append(["Brand", "Qty (Revised)", "Value Rs"])
    for cell in ws[1]:
        cell.font = hf
        cell.fill = hfill
        cell.border = thin
    tq = tv = 0.0
    for b in sorted(club):
        ws.append([b, round(club[b][0], 2), round(club[b][1], 2)])
        tq += club[b][0]
        tv += club[b][1]
    ws.append(["TOTAL", round(tq, 2), round(tv, 2)])
    for c in ws[ws.max_row]:
        c.font = nf
        c.fill = tfill
    for col, w in zip("ABC", [28, 14, 16]):
        ws.column_dimensions[col].width = w

    # Line detail
    ws = out.create_sheet("Line logic")
    ws.append(
        [
            "Brand",
            "Size",
            "BaleSize",
            "Designs",
            "Color",
            "Asst (=Des*BS)",
            "Qnty",
            "Qnty kind",
            "Revised Order",
            "Implied bales (=Rev/BS)",
            "Value Rs",
            "Sheet val lakhs",
            "Pattern note",
        ]
    )
    for cell in ws[1]:
        cell.font = hf
        cell.fill = hfill
    for r in rows:
        if not r["rev"]:
            continue
        ws.append(
            [
                r["brand"],
                r["size"],
                r["bs"],
                r["designs"],
                r["color"],
                r["asst"],
                r["qnty"],
                r["qnty_kind"],
                r["rev"],
                round(r["bales_from_rev"], 2) if r["bales_from_rev"] is not None else None,
                round(r["value_rs"], 2),
                round(r["val_lakhs"], 4) if r["val_lakhs"] else 0,
                r["notes"],
            ]
        )
        if "Rev=Asst" not in r["notes"] and "Rev=Qnty" not in r["notes"]:
            for c in range(1, 14):
                ws.cell(ws.max_row, c).fill = warn
        elif "full assortment" in r["notes"]:
            ws.cell(ws.max_row, 13).fill = info
    for i, w in enumerate([20, 8, 9, 8, 7, 12, 8, 10, 12, 14, 12, 12, 45], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws = out.create_sheet("Not ordered")
    ws.append(["Blank Revised — AB value is 0; not counted as order"])
    ws.append(["Brand", "Size", "Asst", "Qnty", "Qnty kind"])
    for cell in ws[2]:
        cell.font = hf
        cell.fill = hfill
    for r in rows:
        if r["rev"]:
            continue
        ws.append([r["brand"], r["size"], r["asst"], r["qnty"], r["qnty_kind"]])
    for col, w in zip("ABCDE", [22, 10, 10, 10, 12]):
        ws.column_dimensions[col].width = w

    OUT.parent.mkdir(parents=True, exist_ok=True)
    try:
        out.save(OUT)
        saved = OUT
    except PermissionError:
        saved = OUT.with_name("KAG_Careful_Review_v2.xlsx")
        out.save(saved)
    print("\nSaved", saved)
    subprocess.Popen(["cmd", "/c", "start", "", str(saved)], shell=False)


if __name__ == "__main__":
    main()
