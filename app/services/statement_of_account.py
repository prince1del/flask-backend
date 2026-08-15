"""Build a Statement of Account Excel from a SAP-style party GL export (.xls/.xlsx).

Output matches the BD template used for Choice Corner:
  header + summary + Date | Particulars | Type | Debit | Credit | Balance
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

SKIP_TYPE1 = {"mapping", ""}


def _norm_header(h: Any) -> str:
    return re.sub(r"\s+", " ", str(h or "").strip()).lower()


def _find_col(headers: list[str], *candidates: str) -> int | None:
    norms = [_norm_header(h) for h in headers]
    for cand in candidates:
        c = _norm_header(cand)
        for i, h in enumerate(norms):
            if h == c:
                return i
    for cand in candidates:
        c = _norm_header(cand)
        for i, h in enumerate(norms):
            if c and c in h:
                return i
    return None


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if abs(v - int(v)) < 1e-9 and abs(v) < 1e15:
            return str(int(v))
        return str(v).rstrip("0").rstrip(".") if "." in str(v) else str(v)
    return str(v).strip()


def _parse_date(v: Any, *, xl_datemode: int | None = None) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, float) and xl_datemode is not None:
        try:
            import xlrd

            t = xlrd.xldate_as_tuple(v, xl_datemode)
            return date(t[0], t[1], t[2])
        except Exception:
            pass
    s = str(v).strip()
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%y", "%d/%m/%y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _read_sheet_rows(file_bytes: bytes, filename: str) -> tuple[list[str], list[list[Any]], int | None]:
    """Return (headers, data_rows, xlrd_datemode_or_None). Picks densest sheet."""
    name = (filename or "").lower()
    if name.endswith(".xls") and not name.endswith(".xlsx"):
        import xlrd

        wb = xlrd.open_workbook(file_contents=file_bytes)
        best = max(wb.sheet_names(), key=lambda n: wb.sheet_by_name(n).nrows)
        sh = wb.sheet_by_name(best)
        headers = [_cell_str(sh.cell_value(0, c)) for c in range(sh.ncols)]
        rows: list[list[Any]] = []
        for r in range(1, sh.nrows):
            rows.append([sh.cell_value(r, c) for c in range(sh.ncols)])
        return headers, rows, wb.datemode

    bio = io.BytesIO(file_bytes)
    wb = load_workbook(bio, data_only=True, read_only=True)
    best = max(wb.sheetnames, key=lambda n: wb[n].max_row or 0)
    ws = wb[best]
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        raise ValueError("Ledger file is empty")
    headers = [_cell_str(v) for v in all_rows[0]]
    data = [list(r) for r in all_rows[1:] if any(c is not None and str(c).strip() != "" for c in r)]
    return headers, data, None


def classify_type1(type1: str) -> str | None:
    t = (type1 or "").strip().lower()
    if t in SKIP_TYPE1:
        return None
    if "invoice" in t:
        return "Sale Invoice"
    if "bank" in t or "receipt" in t:
        return "Payment Received"
    if "credit" in t:
        return "Credit Note"
    return type1.strip() or None


def build_particulars(
    typ: str,
    *,
    reference: str,
    assignment: str,
    text: str,
) -> str:
    if typ == "Sale Invoice":
        head = f"Sale Invoice No. {reference}" if reference else "Sale Invoice"
        if assignment:
            return f"{head} (Order Ref {assignment})"
        return head
    if typ == "Payment Received":
        if reference:
            return f"Payment Received (RTGS Ref: {reference})"
        return "Payment Received"
    if typ == "Credit Note":
        return text or reference or "Credit Note"
    return text or reference or typ


def build_statement_rows(
    headers: list[str],
    data_rows: list[list[Any]],
    *,
    xl_datemode: int | None = None,
) -> dict[str, Any]:
    i_name = _find_col(headers, "Name 1", "Name")
    i_acc = _find_col(headers, "Account")
    i_ref = _find_col(headers, "Reference")
    i_assign = _find_col(headers, "Assignment")
    i_type1 = _find_col(headers, "Type1")
    i_date = _find_col(headers, "Doc. Date", "Doc Date", "Pstng Date")
    i_amt = _find_col(headers, "LC amnt", "LC amount", "Amount")
    i_text = _find_col(headers, "Text")

    if i_type1 is None or i_amt is None or i_date is None:
        raise ValueError(
            "Ledger needs Type1, Doc. Date, and LC amnt columns (SAP GL export format)"
        )

    party = ""
    account = ""
    lines: list[dict[str, Any]] = []

    for row in data_rows:
        def get(i: int | None) -> Any:
            if i is None or i >= len(row):
                return None
            return row[i]

        typ = classify_type1(_cell_str(get(i_type1)))
        if typ is None:
            continue
        try:
            amt = float(get(i_amt) or 0)
        except (TypeError, ValueError):
            continue
        if abs(amt) < 1e-9:
            continue

        d = _parse_date(get(i_date), xl_datemode=xl_datemode)
        ref = _cell_str(get(i_ref))
        assign = _cell_str(get(i_assign))
        text = _cell_str(get(i_text))
        name = _cell_str(get(i_name))
        acc = _cell_str(get(i_acc))
        if not party and name:
            party = name
        if not account and acc:
            account = acc

        particulars = build_particulars(typ, reference=ref, assignment=assign, text=text)
        debit = round(amt, 2) if amt > 0 else None
        credit = round(abs(amt), 2) if amt < 0 else None
        lines.append(
            {
                "date": d,
                "particulars": particulars,
                "type": typ,
                "debit": debit,
                "credit": credit,
                "signed": round(amt, 2),
            }
        )

    lines.sort(
        key=lambda x: (
            x["date"] or date(1900, 1, 1),
            0 if x["type"] == "Payment Received" else 1 if x["type"] == "Credit Note" else 2,
            x["particulars"],
        )
    )

    balance = 0.0
    sales_signed = 0.0
    payments_signed = 0.0
    credit_notes_signed = 0.0
    for line in lines:
        balance = round(balance + float(line["signed"]), 2)
        line["balance"] = balance
        signed = float(line["signed"])
        if line["type"] == "Sale Invoice":
            sales_signed = round(sales_signed + signed, 2)
        elif line["type"] == "Payment Received":
            payments_signed = round(payments_signed + signed, 2)
        elif line["type"] == "Credit Note":
            credit_notes_signed = round(credit_notes_signed + signed, 2)

    sales = round(abs(sales_signed), 2)
    payments = round(abs(payments_signed), 2)
    credit_notes = round(abs(credit_notes_signed), 2)

    dates = [ln["date"] for ln in lines if ln["date"]]
    period_from = min(dates) if dates else None
    period_to = max(dates) if dates else None

    return {
        "party_name": party or "Party",
        "account_no": account,
        "period_from": period_from,
        "period_to": period_to,
        "total_sales": sales,
        "total_payments": payments,
        "total_credit_notes": credit_notes,
        "closing_balance": balance,
        "lines": lines,
    }


def _fmt_date(d: date | None) -> str:
    if not d:
        return "—"
    return d.strftime("%d-%b-%Y")


def statement_to_xlsx_bytes(statement: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Statement of Account"

    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    money = '#,##0.00'

    party = statement["party_name"]
    ws["A1"] = f"Statement of Account — {party}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:F1")

    period = (
        f"Period: {_fmt_date(statement.get('period_from'))} to "
        f"{_fmt_date(statement.get('period_to'))}"
    )
    ws["A2"] = (
        f"Account No. {statement.get('account_no') or '—'}   |   "
        f"{period}   |   Generated: {_fmt_date(date.today())}"
    )
    ws.merge_cells("A2:F2")

    ws["A4"] = "Summary"
    ws["A4"].font = header_font
    ws["A5"] = "Total Sales (Invoices)"
    ws["B5"] = float(statement["total_sales"])
    ws["A6"] = "Total Payments Received"
    ws["B6"] = float(statement["total_payments"])
    ws["A7"] = "Total Credit Notes / Deductions"
    ws["B7"] = float(statement["total_credit_notes"])
    ws["A8"] = "Net Closing Balance (Dr. = amount due FROM party)"
    ws["B8"] = float(statement["closing_balance"])
    for r in range(5, 9):
        ws[f"B{r}"].number_format = money

    headers = ["Date", "Particulars", "Type", "Debit (Rs.)", "Credit (Rs.)", "Balance (Rs.)"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(9, c, h)
        cell.font = header_font
        cell.border = thin

    row_i = 10
    for line in statement["lines"]:
        d = line["date"]
        ws.cell(row_i, 1, d.strftime("%Y-%m-%d") if d else "")
        ws.cell(row_i, 2, line["particulars"])
        ws.cell(row_i, 3, line["type"])
        deb = line["debit"]
        cre = line["credit"]
        if deb is not None:
            ws.cell(row_i, 4, float(deb)).number_format = money
        if cre is not None:
            ws.cell(row_i, 5, float(cre)).number_format = money
        ws.cell(row_i, 6, float(line["balance"])).number_format = money
        for c in range(1, 7):
            ws.cell(row_i, c).border = thin
        row_i += 1

    ws.cell(row_i, 1, "CLOSING BALANCE").font = Font(bold=True)
    ws.cell(row_i, 6, float(statement["closing_balance"])).number_format = money
    ws.cell(row_i, 6).font = Font(bold=True)

    widths = [14, 64, 18, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22

    for r in ws.iter_rows(min_row=10, max_row=row_i, min_col=1, max_col=6):
        for cell in r:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def generate_statement_of_account_xlsx(
    file_bytes: bytes,
    filename: str,
) -> tuple[bytes, dict[str, Any]]:
    headers, data_rows, datemode = _read_sheet_rows(file_bytes, filename)
    statement = build_statement_rows(headers, data_rows, xl_datemode=datemode)
    xlsx = statement_to_xlsx_bytes(statement)
    meta = {
        "party_name": statement["party_name"],
        "account_no": statement["account_no"],
        "period_from": statement["period_from"].isoformat() if statement["period_from"] else None,
        "period_to": statement["period_to"].isoformat() if statement["period_to"] else None,
        "total_sales": statement["total_sales"],
        "total_payments": statement["total_payments"],
        "total_credit_notes": statement["total_credit_notes"],
        "closing_balance": statement["closing_balance"],
        "line_count": len(statement["lines"]),
        "filename": _safe_soa_filename(statement["party_name"]),
    }
    return xlsx, meta


def _safe_soa_filename(party: str) -> str:
    slug = re.sub(r"[^\w\-]+", "_", (party or "Party").strip())[:40].strip("_") or "Party"
    return f"{slug}_Statement_of_Account.xlsx"
