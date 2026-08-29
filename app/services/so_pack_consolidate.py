"""SO Pack Consolidate — ZIP/RAR of Bombay Dyeing SO PDFs → consolidated workbook."""

from __future__ import annotations

import io
import re
import tempfile
import zipfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from app.routes.data import (
    _clean_pdf_cell_number,
    _clean_pdf_cell_text,
    _parse_sales_order_header_fields,
)
from app.services.bd_product_catalog import brand_wise_only_label, enrich_bd_product

_PRODUCT_SET_RE = re.compile(r"^(.+?\bSET)\b", re.I)
# Fitted sheet / size glued to dimension: KSFST183X198+30 → KS FST
_GLUED_SIZE_FST_RE = re.compile(
    r"\b(SB|DB|KS|QB|CK)FST(?=\d|X|\b)",
    re.I,
)
# FST183X198+30 → FST (keep type token, drop size)
_FST_DIM_TOKEN_RE = re.compile(r"^FST\d", re.I)
_DIM_IN_TOKEN_RE = re.compile(r"\d+X\d+", re.I)
_DESIGN_CODE_RE = re.compile(r"^\d{3,}[A-Z]", re.I)
_TC_TOKEN_RE = re.compile(r"^\d+TC$", re.I)
_DATE_RE = re.compile(
    r"(?:order\s*date|contract\s*date|date)\s*[:\-]?\s*(\d{1,2}[./\-]\d{1,2}[./\-]\d{2,4})",
    re.I,
)
_ORDER_DATE_RE = re.compile(r"order\s*date\s*[:\-]?\s*(\d{1,2}[./\-]\d{1,2}[./\-]\d{2,4})", re.I)
_PO_RE = re.compile(r"\bPO\s*no\.?\s*[:\-]?\s*(.+?)(?:\n|$)", re.I)
_BUYER_PO_LINE_RE = re.compile(
    r"^(.+?)\s+PO\s*no\.?\s*[:\-]?\s*(.+)$",
    re.I | re.M,
)
_CONTRACT_NO_RE = re.compile(r"contract\s*no\.?\s*[:\-]?\s*([A-Za-z0-9\-]+)", re.I)
_BUYER_LINE_RE = re.compile(
    r"buyer\s*code\s*:\s*\S+\s+order\s*date\s*:\s*\S+\s*\n([A-Z0-9][A-Z0-9 &.,'\-]{3,})",
    re.I,
)


def _parse_bd_date(raw: str | None) -> str | None:
    s = (raw or "").strip()
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s[:10]


def product_short_name(full_description: str) -> str:
    """ASTER 1+2 DB SET 224X244 7985BLU 100TC → ASTER 1+2 DB SET

    Fitted sheets (no SET word) also collapse:
    FLFIEST 1+2 KSFST183X198+30 8106BLU140TC → FLFIEST 1+2 KS FST
    ALLURE 1+2 KS FST183X198+30 8112BGE144TC → ALLURE 1+2 KS FST
    """
    text = _clean_pdf_cell_text(full_description) or ""
    first = text.split("\n", 1)[0].strip()
    m = _PRODUCT_SET_RE.match(first)
    if m:
        return m.group(1).strip()

    normalized = _GLUED_SIZE_FST_RE.sub(r"\1 FST", first)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    parts = normalized.split()
    out: list[str] = []
    for p in parts:
        if _FST_DIM_TOKEN_RE.match(p):
            out.append("FST")
            break
        if _DIM_IN_TOKEN_RE.search(p) or _DESIGN_CODE_RE.match(p) or _TC_TOKEN_RE.match(p):
            break
        out.append(p)
        if len(out) >= 8:
            break
    return " ".join(out) if out else first


_MATERIAL_CODE_RE = re.compile(r"^[A-Z0-9][A-Z0-9 \-_/]{2,}$", re.I)


def _norm_header_cell(cell: Any) -> str:
    return re.sub(r"\s+", " ", _clean_pdf_cell_text(str(cell or "")) or "").strip().lower()


def _map_so_headers(header_row: list[Any]) -> dict[str, int]:
    """Map SO table headers → column index (works across layout variants)."""
    mapping: dict[str, int] = {}
    for i, cell in enumerate(header_row or []):
        h = _norm_header_cell(cell)
        if not h:
            continue
        if "material code" in h or h == "code":
            mapping.setdefault("material_code", i)
        elif "material description" in h or "description" in h:
            mapping.setdefault("description", i)
        elif "hsn" in h:
            mapping.setdefault("hsn", i)
        elif h == "qty" or h.startswith("qty"):
            mapping.setdefault("qty", i)
        elif h == "rate":
            mapping.setdefault("rate", i)
        elif h == "unit":
            mapping.setdefault("unit", i)
        elif "shedule" in h or "schedule" in h or "delivery" in h:
            mapping.setdefault("schedule", i)
        elif "net" in h:
            mapping.setdefault("net", i)
        elif "gst" in h:
            mapping.setdefault("gst", i)
        elif "total" in h:
            mapping.setdefault("total", i)
    return mapping


def _cell_at(row: list[Any], idx: int | None) -> Any:
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _parse_so_table_rich(path: Path) -> list[dict[str, Any]]:
    """SO table cells including GST/Total — header-mapped for any BD buyer layout."""
    default_map = {
        "material_code": 0,
        "description": 1,
        "hsn": 2,
        "qty": 3,
        "rate": 4,
        "unit": 5,
        "schedule": 6,
        "net": 7,
        "gst": 8,
        "total": 9,
    }
    items: list[dict[str, Any]] = []
    try:
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages:
                for table in page.extract_tables() or []:
                    if not table:
                        continue
                    colmap: dict[str, int] | None = None
                    for row in table:
                        if not row:
                            continue
                        joined = " ".join(_norm_header_cell(c) for c in row)
                        if "material" in joined and ("qty" in joined or "net" in joined):
                            mapped = _map_so_headers(row)
                            if "material_code" in mapped and "qty" in mapped and "net" in mapped:
                                colmap = mapped
                            continue

                        use_map = colmap
                        if use_map is None:
                            if len(row) < 10:
                                continue
                            use_map = default_map

                        code = (_clean_pdf_cell_text(_cell_at(row, use_map.get("material_code"))) or "").strip()
                        if not code or code.upper() == "TOTAL" or "material" in code.lower():
                            continue
                        if not _MATERIAL_CODE_RE.match(code):
                            continue

                        desc = _clean_pdf_cell_text(_cell_at(row, use_map.get("description")))
                        qty = _clean_pdf_cell_number(_cell_at(row, use_map.get("qty")))
                        rate = _clean_pdf_cell_number(_cell_at(row, use_map.get("rate")))
                        unit = _clean_pdf_cell_text(_cell_at(row, use_map.get("unit"))) or ""
                        schedule = _clean_pdf_cell_text(_cell_at(row, use_map.get("schedule"))) or ""
                        net = _clean_pdf_cell_number(_cell_at(row, use_map.get("net")))
                        gst = _clean_pdf_cell_number(_cell_at(row, use_map.get("gst")))
                        total = _clean_pdf_cell_number(_cell_at(row, use_map.get("total")))
                        if not desc or qty is None or net is None:
                            continue
                        if gst is None:
                            gst = 0.0
                        if total is None:
                            total = round(net + gst, 2)
                        hsn = _clean_pdf_cell_text(_cell_at(row, use_map.get("hsn"))) or ""
                        variant = parse_bd_variant_meta(code, desc)
                        design_code = (variant or {}).get("design_id")
                        colour_code = (variant or {}).get("colour_id")
                        items.append(
                            {
                                "material_code": code,
                                "item_name": desc,
                                "product_name": product_short_name(desc),
                                "product_detail": desc,
                                "design_code": design_code,
                                "colour_code": colour_code,
                                "variant_kind": (variant or {}).get("kind"),
                                "design_count": (variant or {}).get("design_count") or 0,
                                "colour_count": (variant or {}).get("colour_count") or 0,
                                "hsn": hsn,
                                "qty": float(qty),
                                "rate": float(rate or 0),
                                "unit": unit,
                                "schedule_delivery": schedule,
                                "net_amount": float(net),
                                "gst_amount": float(gst),
                                "total_amount": float(total),
                            }
                        )
    except Exception:
        return []
    return items


def _extract_pdf_text(path: Path) -> str:
    chunks: list[str] = []
    try:
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages:
                chunks.append(page.extract_text() or "")
    except Exception:
        return ""
    return "\n".join(chunks)


def _parse_so_header_rich(text: str, source_name: str) -> dict[str, Any]:
    base = _parse_sales_order_header_fields(text) or {}
    contract = None
    m = _CONTRACT_NO_RE.search(text or "")
    if m:
        contract = m.group(1).strip()
    so_number = (
        contract
        or (base.get("order_ref_no") or "").strip()
        or (base.get("invoice_no") or "").strip()
    )
    if not so_number:
        # BND 102876136.pdf / BND_102876136.pdf
        stem = Path(source_name).stem
        m2 = re.search(r"(\d{6,})", stem)
        if m2:
            so_number = m2.group(1)

    order_date = None
    m = _ORDER_DATE_RE.search(text or "")
    if m:
        order_date = _parse_bd_date(m.group(1))
    if not order_date:
        # Contract line "Date : 25.07.2026" — prefer Order Date when present
        dates = _DATE_RE.findall(text or "")
        if dates:
            order_date = _parse_bd_date(dates[-1] if len(dates) > 1 else dates[0])

    po_number = None
    buyer_name = (base.get("buyer_name") or "").strip()
    m = _BUYER_PO_LINE_RE.search(text or "")
    if m:
        maybe_buyer = m.group(1).strip()
        po_number = m.group(2).strip()
        if not re.search(r"buyer\s*code|order\s*date|gst\s*no|contract", maybe_buyer, re.I):
            if not buyer_name:
                buyer_name = maybe_buyer
    if not po_number:
        m = _PO_RE.search(text or "")
        if m:
            po_number = m.group(1).strip()
            po_number = re.split(r"\s{2,}|\t", po_number)[0].strip()

    if not buyer_name:
        m = _BUYER_LINE_RE.search(text or "")
        if m:
            buyer_name = m.group(1).strip()
            buyer_name = re.sub(r"\s+PO\s*no\.?.*$", "", buyer_name, flags=re.I).strip()
    if not buyer_name:
        # Line after Buyer Code row is usually buyer name
        lines = [ln.strip() for ln in (text or "").splitlines() if ln.strip()]
        for i, ln in enumerate(lines):
            if re.search(r"buyer\s*code", ln, re.I) and i + 1 < len(lines):
                nxt = lines[i + 1]
                if not re.search(r"gst\s*no|we the undersigned", nxt, re.I):
                    buyer_name = re.sub(r"\s+PO\s*no\.?.*$", "", nxt, flags=re.I).strip()
                break

    return {
        "so_number": so_number or "",
        "order_date": order_date,
        "buyer_code": (base.get("buyer_code") or "").strip() or None,
        "buyer_name": buyer_name or None,
        "po_number": po_number,
    }


_DESIGN_COLOR_TRAIL_RE = re.compile(r"(\d{4})([A-Za-z]{3})\s*$")
_DESIGN_COLOR_TOKEN_RE = re.compile(r"(?<![A-Za-z0-9])(\d{4})([A-Za-z]{3})(?![A-Za-z0-9])")
_ASSORT_RE = re.compile(r"ASS(?:ORT|RT|T)\s*(\d+)", re.I)
_WHITE_RE = re.compile(r"\b(?:WHITE|WHITES|WHITS|WHT)\b", re.I)


def parse_bd_design_color(
    material_code: str | None = None,
    product_detail: str | None = None,
) -> tuple[str | None, str | None]:
    """Extract numeric design + colour (e.g. BS03DBASTER8142DPK → 8142, DPK)."""
    mat = (material_code or "").strip()
    if mat:
        m = _DESIGN_COLOR_TRAIL_RE.search(mat)
        if m:
            return m.group(1), m.group(2).upper()
        m = _DESIGN_COLOR_TOKEN_RE.search(mat)
        if m:
            return m.group(1), m.group(2).upper()
        m = re.search(r"(\d{4})([A-Za-z]{3})(?:\d+TC)?$", re.sub(r"\s+", "", mat), re.I)
        if m:
            return m.group(1), m.group(2).upper()

    detail = (product_detail or "").strip()
    if detail:
        tokens = _DESIGN_COLOR_TOKEN_RE.findall(detail)
        if tokens:
            design, colour = tokens[-1]
            return design, colour.upper()
        compact = re.sub(r"\s+", "", detail)
        m = re.search(r"(\d{4})([A-Za-z]{3})(?:\d+TC)?", compact, re.I)
        if m:
            return m.group(1), m.group(2).upper()
    return None, None


def parse_bd_variant_meta(
    material_code: str | None = None,
    product_detail: str | None = None,
) -> dict[str, Any] | None:
    """Resolve design/colour variant for Brand Wise Size Wise counts.

    Supports:
      - numeric design+colour: 8142DPK → 1 design, 1 colour code
      - assortment packs: ASSORT4 / ASST6 → 1 design, N colours
      - solid finishes: WHITE → 1 design, 1 colour
    """
    design, colour = parse_bd_design_color(material_code, product_detail)
    if design and colour:
        return {
            "kind": "numeric",
            "design_id": str(design),
            "colour_id": str(colour).upper(),
            "design_count": 1,
            "colour_count": 1,
        }

    blob = f"{material_code or ''} {product_detail or ''}"
    m = _ASSORT_RE.search(blob)
    if not m:
        m = _ASSORT_RE.search(re.sub(r"\s+", "", blob))
    if m:
        n = int(m.group(1))
        if n > 0:
            label = f"ASSORT{n}"
            return {
                "kind": "assort",
                "design_id": label,
                "colour_id": label,
                "design_count": 1,  # assortment = one design
                "colour_count": n,  # ASSORT4 → 4 colours
            }

    if _WHITE_RE.search(blob):
        return {
            "kind": "solid",
            "design_id": "WHITE",
            "colour_id": "WHITE",
            "design_count": 1,
            "colour_count": 1,
        }
    return None


def _unpack_archive(file_bytes: bytes, filename: str) -> list[tuple[str, bytes]]:
    """Return list of (relative_name, pdf_bytes) in archive member order."""
    name = (filename or "pack.zip").lower()
    pdfs: list[tuple[str, bytes]] = []

    if name.endswith(".rar"):
        try:
            import rarfile  # type: ignore
        except ImportError as exc:
            raise ValueError(
                "RAR support needs the rarfile package. Install rarfile, or upload a ZIP instead."
            ) from exc
        with tempfile.NamedTemporaryFile(suffix=".rar", delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        try:
            with rarfile.RarFile(tmp_path) as rf:
                for info in rf.infolist():
                    fn = info.filename
                    if fn.lower().endswith(".pdf") and not info.is_dir():
                        try:
                            pdfs.append((Path(fn).name, rf.read(info)))
                        except rarfile.RarCannotExec as exc:
                            raise ValueError(
                                "RAR uploaded, but server extractor tool is unavailable. "
                                "Please upload ZIP (same PDFs) or enable unrar/unar/bsdtar on server."
                            ) from exc
        finally:
            try:
                Path(tmp_path).unlink(missing_ok=True)
            except OSError:
                pass
        return pdfs

    # ZIP (default)
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue
                fn = info.filename
                if fn.lower().endswith(".pdf"):
                    pdfs.append((Path(fn).name, zf.read(info)))
    except zipfile.BadZipFile as exc:
        raise ValueError("Not a valid ZIP file") from exc
    return pdfs


def _load_pack_pdfs(file_bytes: bytes, filename: str) -> list[tuple[str, bytes]]:
    """Load PDFs from a ZIP/RAR archive or a single PDF upload."""
    name = (filename or "").lower()
    if name.endswith(".pdf"):
        label = Path(filename or "SO.pdf").name
        return [(label, file_bytes)]
    return _unpack_archive(file_bytes, filename)


def iter_analyze_so_pack(
    file_bytes: bytes,
    filename: str,
):
    """Yield ("progress", message) then ("done", payload)."""
    short_arch = Path(filename or "pack.zip").name
    lower = short_arch.lower()
    if lower.endswith(".pdf"):
        yield ("progress", f"Loading {short_arch}…")
    else:
        yield ("progress", f"Unpacking {short_arch}…")
    pdfs = _load_pack_pdfs(file_bytes, filename)
    if not pdfs:
        raise ValueError("No PDF files found in the upload")
    yield from _iter_analyze_pdf_list(pdfs, filename)


def iter_analyze_so_pack_pdfs(
    pdfs: list[tuple[str, bytes]],
    source_label: str = "SO_PDFs",
):
    """Analyze an explicit list of PDFs (loose uploads, same distributor)."""
    if not pdfs:
        raise ValueError("No PDF files to analyze")
    label = source_label or "SO_PDFs"
    yield ("progress", f"Loading {len(pdfs)} PDF(s)…")
    yield from _iter_analyze_pdf_list(pdfs, label)


def _iter_analyze_pdf_list(
    pdfs: list[tuple[str, bytes]],
    source_filename: str,
):
    """Shared PDF parse → consolidate pipeline with progress yields."""
    total_pdfs = len(pdfs)
    yield ("progress", f"Found {total_pdfs} PDF(s) — starting parse…")

    line_detail: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    so_map: dict[str, dict[str, Any]] = {}

    with tempfile.TemporaryDirectory(prefix="so_pack_") as tmp:
        tmp_path = Path(tmp)
        for idx, (source_name, raw) in enumerate(pdfs, start=1):
            pdf_label = Path(str(source_name).replace("\\", "/")).name
            yield ("progress", f"PDF {idx}/{total_pdfs}: reading {pdf_label}")
            safe_name = source_name.replace("/", "_").replace("\\", "_")
            pdf_path = tmp_path / safe_name
            pdf_path.write_bytes(raw)
            text = _extract_pdf_text(pdf_path)
            yield ("progress", f"PDF {idx}/{total_pdfs}: header — {pdf_label}")
            header = _parse_so_header_rich(text, source_name)
            yield ("progress", f"PDF {idx}/{total_pdfs}: line items — {pdf_label}")
            lines = _parse_so_table_rich(pdf_path)
            if not lines:
                errors.append({"source_pdf": source_name, "error": "No line items parsed"})
                yield ("progress", f"PDF {idx}/{total_pdfs}: no lines in {pdf_label}")
                continue
            so_number = header.get("so_number") or Path(source_name).stem
            buyer = (header.get("buyer_name") or "").strip()
            so_hint = f"SO {so_number}" + (f" · {buyer}" if buyer else "")
            yield (
                "progress",
                f"PDF {idx}/{total_pdfs}: {len(lines)} lines · {so_hint}",
            )
            if so_number not in so_map:
                so_map[so_number] = {
                    "so_number": so_number,
                    "order_date": header.get("order_date"),
                    "buyer_code": header.get("buyer_code"),
                    "buyer_name": header.get("buyer_name"),
                    "po_number": header.get("po_number"),
                    "source_pdf": source_name,
                    "lines": [],
                }
            bucket = so_map[so_number]
            for k in ("order_date", "buyer_code", "buyer_name", "po_number"):
                if not bucket.get(k) and header.get(k):
                    bucket[k] = header.get(k)
            for ln in lines:
                row = {
                    "so_number": so_number,
                    "order_date": bucket.get("order_date"),
                    "buyer_code": bucket.get("buyer_code"),
                    "buyer_name": bucket.get("buyer_name"),
                    "po_number": bucket.get("po_number"),
                    "material_code": ln["material_code"],
                    "product_name": ln["product_name"],
                    "product_detail": ln["product_detail"],
                    "design_code": ln.get("design_code"),
                    "colour_code": ln.get("colour_code"),
                    "variant_kind": ln.get("variant_kind"),
                    "design_count": ln.get("design_count") or 0,
                    "colour_count": ln.get("colour_count") or 0,
                    "hsn": ln.get("hsn"),
                    "qty": ln["qty"],
                    "rate": ln["rate"],
                    "unit": ln.get("unit"),
                    "schedule_delivery": ln.get("schedule_delivery"),
                    "net_amount": ln["net_amount"],
                    "gst_amount": ln["gst_amount"],
                    "total_amount": ln["total_amount"],
                    "source_pdf": source_name,
                }
                line_detail.append(row)
                bucket["lines"].append(row)

    yield ("progress", "Matching product names & rolling up qty / amounts…")
    consol_acc: dict[tuple[str, str], dict[str, Any]] = {}
    for row in line_detail:
        key = (str(row["so_number"]), str(row["product_name"]))
        acc = consol_acc.get(key)
        if not acc:
            consol_acc[key] = {
                "so_number": row["so_number"],
                "order_date": row.get("order_date"),
                "buyer_name": row.get("buyer_name"),
                "po_number": row.get("po_number"),
                "product_name": row["product_name"],
                "sku_lines": 0,
                "total_qty": 0.0,
                "net_amount": 0.0,
                "gst_amount": 0.0,
                "total_amount": 0.0,
            }
            acc = consol_acc[key]
        acc["sku_lines"] += 1
        acc["total_qty"] = round(acc["total_qty"] + float(row["qty"]), 3)
        acc["net_amount"] = round(acc["net_amount"] + float(row["net_amount"]), 2)
        acc["gst_amount"] = round(acc["gst_amount"] + float(row["gst_amount"]), 2)
        acc["total_amount"] = round(acc["total_amount"] + float(row["total_amount"]), 2)

    consolidated = sorted(
        consol_acc.values(),
        key=lambda r: (str(r["so_number"]), str(r["product_name"])),
    )

    yield ("progress", "Building SO summary…")
    so_summary: list[dict[str, Any]] = []
    for so_number, bucket in sorted(so_map.items(), key=lambda x: x[0]):
        lines = bucket.get("lines") or []
        products = {ln["product_name"] for ln in lines}
        so_summary.append(
            {
                "so_number": so_number,
                "order_date": bucket.get("order_date"),
                "buyer_code": bucket.get("buyer_code"),
                "buyer_name": bucket.get("buyer_name"),
                "po_number": bucket.get("po_number"),
                "product_types": len(products),
                "sku_lines": len(lines),
                "total_qty": round(sum(float(ln["qty"]) for ln in lines), 3),
                "net_amount": round(sum(float(ln["net_amount"]) for ln in lines), 2),
                "gst_amount": round(sum(float(ln["gst_amount"]) for ln in lines), 2),
                "total_amount": round(sum(float(ln["total_amount"]) for ln in lines), 2),
                "source_pdf": bucket.get("source_pdf"),
            }
        )

    meta = {
        "source_filename": source_filename,
        "pdf_count": len(pdfs),
        "so_count": len(so_summary),
        "consolidated_rows": len(consolidated),
        "line_rows": len(line_detail),
        "total_qty": round(sum(r["total_qty"] for r in so_summary), 3),
        "net_amount": round(sum(r["net_amount"] for r in so_summary), 2),
        "gst_amount": round(sum(r["gst_amount"] for r in so_summary), 2),
        "total_amount": round(sum(r["total_amount"] for r in so_summary), 2),
        "errors": errors,
    }
    yield (
        "progress",
        f"Done · {len(so_summary)} SO · {len(consolidated)} products · {len(line_detail)} lines",
    )
    yield (
        "done",
        {
            "meta": meta,
            "consolidated": consolidated,
            "so_summary": so_summary,
            "line_detail": line_detail,
        },
    )


def analyze_so_pack(file_bytes: bytes, filename: str) -> dict[str, Any]:
    """Parse all SO PDFs in a zip/rar/pdf → consolidated / summary / line detail."""
    result: dict[str, Any] | None = None
    for kind, payload in iter_analyze_so_pack(file_bytes, filename):
        if kind == "done" and isinstance(payload, dict):
            result = payload
    if result is None:
        raise ValueError("SO pack analyze produced no result")
    return result


def analyze_so_pack_pdfs(
    pdfs: list[tuple[str, bytes]],
    source_label: str = "SO_PDFs",
) -> dict[str, Any]:
    """Parse loose SO PDFs → consolidated payload."""
    result: dict[str, Any] | None = None
    for kind, payload in iter_analyze_so_pack_pdfs(pdfs, source_label):
        if kind == "done" and isinstance(payload, dict):
            result = payload
    if result is None:
        raise ValueError("SO pack analyze produced no result")
    return result


def so_pack_excel_download_name(payload: dict[str, Any] | None) -> str:
    """Filename from distributor/buyer name(s), e.g. 'KALRA AGENCIES_SO_Pack.xlsx'."""
    buyers: list[str] = []
    seen: set[str] = set()
    payload = payload or {}
    for row in (payload.get("so_summary") or []) + (payload.get("consolidated") or []):
        if not isinstance(row, dict):
            continue
        name = str(row.get("buyer_name") or "").strip()
        if not name:
            continue
        key = name.casefold()
        if key in seen:
            continue
        seen.add(key)
        buyers.append(name)

    if len(buyers) == 1:
        label = buyers[0]
    elif len(buyers) > 1:
        label = "_".join(buyers[:3]) if len(buyers) <= 3 else f"{buyers[0]}_and_{len(buyers) - 1}_more"
    else:
        src = str((payload.get("meta") or {}).get("source_filename") or "")
        label = re.sub(r"\.(zip|rar)$", "", src, flags=re.I).strip() or "SO_Pack"

    safe = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "", label)
    safe = re.sub(r"\s+", " ", safe).strip(" .") or "SO_Pack"
    safe = safe[:100]
    return f"{safe}_SO_Pack.xlsx"


def build_batch_excel_zip(packs: list[dict[str, Any]]) -> tuple[bytes, str]:
    """Build one ZIP containing a separate Excel per distributor pack.

    Returns (zip_bytes, download_filename).
    """
    if not packs:
        raise ValueError("No packs to export")

    used: set[str] = set()
    buf = io.BytesIO()
    written = 0
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for payload in packs:
            if not isinstance(payload, dict):
                continue
            if not (
                payload.get("meta")
                or payload.get("consolidated")
                or payload.get("so_summary")
                or payload.get("line_detail")
            ):
                continue
            xlsx_bytes = build_consolidated_xlsx(payload)
            name = so_pack_excel_download_name(payload)
            base, _, ext = name.rpartition(".")
            if not base:
                base, ext = name, "xlsx"
            candidate = name
            n = 2
            while candidate.casefold() in used:
                candidate = f"{base}_{n}.{ext}"
                n += 1
            used.add(candidate.casefold())
            zf.writestr(candidate, xlsx_bytes)
            written += 1

    if written < 1:
        raise ValueError("No valid packs to export")

    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    return buf.getvalue(), f"SO_Pack_Batch_{written}_{stamp}.zip"


def build_consolidated_xlsx(payload: dict[str, Any]) -> bytes:
    """Build workbook matching Consolidated_SO_Product_Qty_Amount.xlsx shape.

    All sheets: AutoFilter on header row.
    Consolidated / SO Summary: bottom TOTAL via SUBTOTAL(109) (filter-aware).
    Line Item Detail: top SUBTOTAL(109) + freeze at B4 (col A + rows 1–3 sticky).
    """
    meta = payload.get("meta") or {}
    wb = Workbook()
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)

    def _subtotal_sum(sheet, row_i: int, col_letter: str, first_data: int, last_data: int) -> None:
        """Filter-aware SUM (109 = ignore rows hidden by AutoFilter)."""
        cell = sheet[f"{col_letter}{row_i}"]
        if last_data >= first_data:
            cell.value = f"=SUBTOTAL(109,{col_letter}{first_data}:{col_letter}{last_data})"
        else:
            cell.value = 0
        cell.font = bold

    # --- Consolidated ---
    ws = wb.active
    ws.title = "Consolidated"
    pdf_n = int(meta.get("pdf_count") or 0)
    ws["A1"] = "SO-wise Product Consolidation"
    ws["A1"].font = title_font
    ws["A2"] = f"Consolidated from {pdf_n} SO PDFs | Product-level summary"
    headers = [
        "SO Number",
        "Order Date",
        "Buyer Name",
        "PO Number",
        "Product Name",
        "SKU Lines",
        "Total Qty",
        "Net Amount",
        "GST Amount",
        "Total Amount",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, h)
        cell.font = bold
    consol_rows = list(payload.get("consolidated") or [])
    for r_i, row in enumerate(consol_rows, 5):
        vals = [
            row.get("so_number"),
            row.get("order_date"),
            row.get("buyer_name"),
            row.get("po_number"),
            row.get("product_name"),
            row.get("sku_lines"),
            row.get("total_qty"),
            row.get("net_amount"),
            row.get("gst_amount"),
            row.get("total_amount"),
        ]
        for c_i, v in enumerate(vals, 1):
            ws.cell(r_i, c_i, v)
    consol_first, consol_last = 5, 4 + len(consol_rows)
    if consol_rows:
        total_r = consol_last + 1
        ws.cell(total_r, 1, "TOTAL").font = bold
        for col_letter in ("G", "H", "I", "J"):
            _subtotal_sum(ws, total_r, col_letter, consol_first, consol_last)
        ws.auto_filter.ref = f"A4:J{consol_last}"
    else:
        ws.auto_filter.ref = "A4:J4"

    # --- SO Summary ---
    ws2 = wb.create_sheet("SO Summary")
    ws2["A1"] = "SO Summary"
    ws2["A1"].font = title_font
    h2 = [
        "SO Number",
        "Order Date",
        "Buyer Code",
        "Buyer Name",
        "PO Number",
        "Product Types",
        "SKU Lines",
        "Total Qty",
        "Net Amount",
        "GST Amount",
        "Total Amount",
        "Source PDF",
    ]
    for col, h in enumerate(h2, 1):
        cell = ws2.cell(3, col, h)
        cell.font = bold
    summary_rows = list(payload.get("so_summary") or [])
    for r_i, row in enumerate(summary_rows, 4):
        vals = [
            row.get("so_number"),
            row.get("order_date"),
            row.get("buyer_code"),
            row.get("buyer_name"),
            row.get("po_number"),
            row.get("product_types"),
            row.get("sku_lines"),
            row.get("total_qty"),
            row.get("net_amount"),
            row.get("gst_amount"),
            row.get("total_amount"),
            row.get("source_pdf"),
        ]
        for c_i, v in enumerate(vals, 1):
            ws2.cell(r_i, c_i, v)
    sum_first, sum_last = 4, 3 + len(summary_rows)
    if summary_rows:
        total_r = sum_last + 1
        ws2.cell(total_r, 1, "TOTAL").font = bold
        for col_letter in ("H", "I", "J", "K"):
            _subtotal_sum(ws2, total_r, col_letter, sum_first, sum_last)
        ws2.auto_filter.ref = f"A3:L{sum_last}"
    else:
        ws2.auto_filter.ref = "A3:L3"

    # --- Line Item Detail ---
    ws3 = wb.create_sheet("Line Item Detail")
    ws3["A1"] = "Line Item Detail"
    ws3["A1"].font = title_font
    h3 = [
        "SO Number",
        "Order Date",
        "Buyer Name",
        "PO Number",
        "Material Code",
        "Product Name",
        "Product Detail/Size",
        "HSN",
        "Qty",
        "Rate",
        "Unit",
        "Schedule Delivery",
        "Net Amount",
        "GST Amount",
        "Total Amount",
        "Source PDF",
    ]
    line_rows = list(payload.get("line_detail") or [])
    # Row 2 = filter-aware SUBTOTAL, Row 3 = headers, data from row 4
    ws3["A2"] = "TOTAL"
    ws3["A2"].font = bold
    for col, h in enumerate(h3, 1):
        cell = ws3.cell(3, col, h)
        cell.font = bold
    for r_i, row in enumerate(line_rows, 4):
        vals = [
            row.get("so_number"),
            row.get("order_date"),
            row.get("buyer_name"),
            row.get("po_number"),
            row.get("material_code"),
            row.get("product_name"),
            row.get("product_detail"),
            row.get("hsn"),
            row.get("qty"),
            row.get("rate"),
            row.get("unit"),
            row.get("schedule_delivery"),
            row.get("net_amount"),
            row.get("gst_amount"),
            row.get("total_amount"),
            row.get("source_pdf"),
        ]
        for c_i, v in enumerate(vals, 1):
            ws3.cell(r_i, c_i, v)
    line_first, line_last = 4, 3 + len(line_rows)
    # Qty=I, Net=M, GST=N, Total=O — SUBTOTAL updates when AutoFilter hides rows
    for col_letter in ("I", "M", "N", "O"):
        _subtotal_sum(ws3, 2, col_letter, line_first, line_last)
    if line_rows:
        ws3.auto_filter.ref = f"A3:P{line_last}"
    else:
        ws3.auto_filter.ref = "A3:P3"

    def _line_short(row: dict[str, Any]) -> str:
        short = str(row.get("product_name") or "").strip()
        if not short and row.get("product_detail"):
            short = product_short_name(str(row.get("product_detail") or ""))
        return short

    def _write_brand_only_sheet(sheet) -> None:
        """Brand Wise Summary — collection only; order = first PO appearance."""
        headers = ["Brand", "Total Qty", "Net Amount", "GST Amount", "Total Amount"]
        sheet["A1"] = "Brand Wise Summary"
        sheet["A1"].font = title_font
        for col, h in enumerate(headers, 1):
            cell = sheet.cell(2, col, h)
            cell.font = bold

        brand_acc: dict[str, dict[str, Any]] = {}
        others = {"qty": 0.0, "net": 0.0, "gst": 0.0, "total": 0.0, "order": 10**9}
        for idx, row in enumerate(line_rows):
            short = _line_short(row)
            label = brand_wise_only_label(short) if short else None
            if label:
                target = brand_acc.get(label)
                if not target:
                    target = {"qty": 0.0, "net": 0.0, "gst": 0.0, "total": 0.0, "order": idx}
                    brand_acc[label] = target
            else:
                target = others
            target["qty"] = round(target["qty"] + float(row.get("qty") or 0), 3)
            target["net"] = round(target["net"] + float(row.get("net_amount") or 0), 2)
            target["gst"] = round(target["gst"] + float(row.get("gst_amount") or 0), 2)
            target["total"] = round(target["total"] + float(row.get("total_amount") or 0), 2)

        brand_sorted = sorted(brand_acc.items(), key=lambda item: item[1]["order"])
        r_i = 3
        for label, totals in brand_sorted:
            sheet.cell(r_i, 1, label)
            sheet.cell(r_i, 2, totals["qty"])
            sheet.cell(r_i, 3, totals["net"])
            sheet.cell(r_i, 4, totals["gst"])
            sheet.cell(r_i, 5, totals["total"])
            r_i += 1

        others_row = r_i
        sheet.cell(others_row, 1, "Others").font = bold
        sheet.cell(others_row, 2, others["qty"]).font = bold
        sheet.cell(others_row, 3, others["net"]).font = bold
        sheet.cell(others_row, 4, others["gst"]).font = bold
        sheet.cell(others_row, 5, others["total"]).font = bold

        total_row = others_row + 1
        data_first, data_last = 3, others_row
        sheet.cell(total_row, 1, "TOTAL").font = bold
        for col_letter in ("B", "C", "D", "E"):
            _subtotal_sum(sheet, total_row, col_letter, data_first, data_last)
        sheet.auto_filter.ref = f"A2:E{data_last}" if data_last >= data_first else "A2:E2"
        sheet.freeze_panes = "A3"
        sheet.column_dimensions["A"].width = 28
        for letter in ("B", "C", "D", "E"):
            sheet.column_dimensions[letter].width = 16

    def _write_brand_size_wise_sheet(sheet) -> None:
        """Brand Wise Size Wise Summary — Brand | Sheet Option | Design | Colour | qty/amounts."""
        headers = [
            "Brand",
            "Sheet Option",
            "Total Design option",
            "Total Colour Option",
            "Total Qty",
            "Net Amount",
            "GST Amount",
            "Total Amount",
        ]
        sheet["A1"] = "Brand Wise Size Wise Summary"
        sheet["A1"].font = title_font
        for col, h in enumerate(headers, 1):
            cell = sheet.cell(2, col, h)
            cell.font = bold

        # key -> rollup; order = first line index in pack (ZIP/PO order)
        groups: dict[tuple[str, str], dict[str, Any]] = {}
        others = {
            "qty": 0.0,
            "net": 0.0,
            "gst": 0.0,
            "total": 0.0,
            "order": 10**9,
            "numeric_design_colors": defaultdict(set),
            "assort_colour_counts": {},  # ASSORT4 → 4 colours (1 design)
            "solids": set(),
        }

        def _blank_group(order_idx: int) -> dict[str, Any]:
            return {
                "qty": 0.0,
                "net": 0.0,
                "gst": 0.0,
                "total": 0.0,
                "order": order_idx,
                "numeric_design_colors": defaultdict(set),
                "assort_colour_counts": {},
                "solids": set(),
            }

        for idx, row in enumerate(line_rows):
            short = _line_short(row)
            enriched = enrich_bd_product(short) if short else {}
            brand = enriched.get("collection")
            sheet_opt = enriched.get("product_type")

            # Always re-parse from material/detail so ASSORT meaning stays correct.
            variant = parse_bd_variant_meta(
                row.get("material_code"),
                row.get("product_detail"),
            )

            if brand and sheet_opt:
                key = (str(brand), str(sheet_opt))
                target = groups.get(key)
                if not target:
                    target = _blank_group(idx)
                    groups[key] = target
            else:
                target = others

            target["qty"] = round(target["qty"] + float(row.get("qty") or 0), 3)
            target["net"] = round(target["net"] + float(row.get("net_amount") or 0), 2)
            target["gst"] = round(target["gst"] + float(row.get("gst_amount") or 0), 2)
            target["total"] = round(target["total"] + float(row.get("total_amount") or 0), 2)

            if variant:
                kind = variant.get("kind")
                design_id = str(variant.get("design_id") or "")
                colour_id = str(variant.get("colour_id") or "").upper()
                if kind == "numeric" and design_id and colour_id:
                    target["numeric_design_colors"][design_id].add(colour_id)
                elif kind == "assort":
                    n = int(variant.get("colour_count") or 0)
                    if n > 0 and design_id:
                        # Keep max colour count if duplicate ASSORT lines appear.
                        prev = int(target["assort_colour_counts"].get(design_id) or 0)
                        target["assort_colour_counts"][design_id] = max(prev, n)
                elif kind == "solid":
                    solid_id = design_id or colour_id or "WHITE"
                    target["solids"].add(solid_id)

        def _design_colour_counts(bucket: dict[str, Any]) -> tuple[int, int]:
            numeric = bucket.get("numeric_design_colors") or {}
            assort = bucket.get("assort_colour_counts") or {}
            solids = bucket.get("solids") or set()

            # ASSORT4 = 1 design + 4 colours; WHITE = 1 design + 1 colour
            design_n = len(numeric) + len(assort) + len(solids)
            colour_n = 0
            if numeric:
                colour_n += max((len(cols) for cols in numeric.values()), default=0)
            colour_n += sum(int(v) for v in assort.values())
            colour_n += len(solids)
            return int(design_n), int(colour_n)
        ordered = sorted(groups.items(), key=lambda item: item[1]["order"])
        r_i = 3
        for (brand, sheet_opt), totals in ordered:
            design_n, colour_n = _design_colour_counts(totals)
            sheet.cell(r_i, 1, brand)
            sheet.cell(r_i, 2, sheet_opt)
            sheet.cell(r_i, 3, design_n)
            sheet.cell(r_i, 4, colour_n)
            sheet.cell(r_i, 5, totals["qty"])
            sheet.cell(r_i, 6, totals["net"])
            sheet.cell(r_i, 7, totals["gst"])
            sheet.cell(r_i, 8, totals["total"])
            r_i += 1

        others_row = r_i
        o_design_n, o_colour_n = _design_colour_counts(others)
        sheet.cell(others_row, 1, "Others").font = bold
        sheet.cell(others_row, 2, "").font = bold
        sheet.cell(others_row, 3, o_design_n).font = bold
        sheet.cell(others_row, 4, o_colour_n).font = bold
        sheet.cell(others_row, 5, others["qty"]).font = bold
        sheet.cell(others_row, 6, others["net"]).font = bold
        sheet.cell(others_row, 7, others["gst"]).font = bold
        sheet.cell(others_row, 8, others["total"]).font = bold

        total_row = others_row + 1
        data_first, data_last = 3, others_row
        sheet.cell(total_row, 1, "TOTAL").font = bold
        for col_letter in ("E", "F", "G", "H"):
            _subtotal_sum(sheet, total_row, col_letter, data_first, data_last)
        sheet.auto_filter.ref = f"A2:H{data_last}" if data_last >= data_first else "A2:H2"
        sheet.freeze_panes = "A3"
        sheet.column_dimensions["A"].width = 22
        sheet.column_dimensions["B"].width = 22
        for letter in ("C", "D", "E", "F", "G", "H"):
            sheet.column_dimensions[letter].width = 16

    # 4) Brand + sheet option + design/colour counts (ZIP/PO order)
    ws4 = wb.create_sheet("Brand Wise Size Wise Summary")
    _write_brand_size_wise_sheet(ws4)

    # 5) Brand only (taught collection), ZIP/PO order
    ws5 = wb.create_sheet("Brand Wise Summary")
    _write_brand_only_sheet(ws5)

    for sheet in (ws, ws2, ws3):
        for col in sheet.columns:
            letter = col[0].column_letter
            sheet.column_dimensions[letter].width = min(28, max(12, len(str(col[0].value or "")) + 4))

    # Freeze last: Consolidated/Summary header sticky; Line Item keeps SO# (col A) + top 3 rows.
    ws.freeze_panes = "A5"
    ws2.freeze_panes = "A4"
    ws3.freeze_panes = "B4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
