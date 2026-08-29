"""Lightweight spreadsheet preview for Drive files (mobile in-app viewer)."""

from __future__ import annotations

import csv
import io
from typing import Any


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return f"{value:.4g}"
    return str(value).strip()


def _trim_rows(rows: list[list[str]], *, max_rows: int, max_cols: int) -> tuple[list[list[str]], bool, bool]:
    truncated_rows = len(rows) > max_rows
    trimmed = rows[:max_rows]
    col_truncated = False
    out: list[list[str]] = []
    for row in trimmed:
        cells = [_cell_str(c) for c in row[:max_cols]]
        if len(row) > max_cols:
            col_truncated = True
        if any(cells):
            out.append(cells)
    return out, truncated_rows, col_truncated


def preview_spreadsheet_bytes(
    content: bytes,
    filename: str,
    *,
    max_rows: int = 200,
    max_cols: int = 24,
    max_sheets: int = 5,
) -> dict[str, Any]:
    """Parse xlsx/xls/csv into JSON-safe rows for mobile preview."""
    name = (filename or "workbook").strip()
    lower = name.lower()
    sheets: list[dict[str, Any]] = []

    if lower.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows_raw = [list(row) for row in reader]
        rows, row_trunc, col_trunc = _trim_rows(rows_raw, max_rows=max_rows, max_cols=max_cols)
        sheets.append(
            {
                "name": "CSV",
                "rows": rows,
                "row_count": len(rows_raw),
                "col_count": max((len(r) for r in rows_raw), default=0),
                "truncated_rows": row_trunc,
                "truncated_cols": col_trunc,
            }
        )
    elif lower.endswith(".xlsx"):
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            for sheet_name in wb.sheetnames[:max_sheets]:
                ws = wb[sheet_name]
                rows_raw: list[list[str]] = []
                for row in ws.iter_rows(values_only=True):
                    rows_raw.append([_cell_str(c) for c in row])
                    if len(rows_raw) >= max_rows + 1:
                        break
                rows, row_trunc, col_trunc = _trim_rows(rows_raw, max_rows=max_rows, max_cols=max_cols)
                sheets.append(
                    {
                        "name": sheet_name,
                        "rows": rows,
                        "row_count": len(rows_raw),
                        "col_count": max((len(r) for r in rows_raw), default=0),
                        "truncated_rows": row_trunc or len(rows_raw) > max_rows,
                        "truncated_cols": col_trunc,
                    }
                )
        finally:
            wb.close()
    elif lower.endswith(".xls"):
        import xlrd

        book = xlrd.open_workbook(file_contents=content)
        for index in range(min(book.nsheets, max_sheets)):
            sh = book.sheet_by_index(index)
            rows_raw: list[list[str]] = []
            for r in range(sh.nrows):
                rows_raw.append([_cell_str(sh.cell_value(r, c)) for c in range(sh.ncols)])
                if len(rows_raw) >= max_rows + 1:
                    break
            rows, row_trunc, col_trunc = _trim_rows(rows_raw, max_rows=max_rows, max_cols=max_cols)
            sheets.append(
                {
                    "name": sh.name,
                    "rows": rows,
                    "row_count": sh.nrows,
                    "col_count": sh.ncols,
                    "truncated_rows": row_trunc or sh.nrows > max_rows,
                    "truncated_cols": col_trunc,
                }
            )
    else:
        raise ValueError("Preview supports .xlsx, .xls, and .csv only")

    if not sheets or not any(s.get("rows") for s in sheets):
        raise ValueError("Spreadsheet is empty or could not be read")

    return {
        "kind": "spreadsheet",
        "file_name": name,
        "sheets": sheets,
    }
