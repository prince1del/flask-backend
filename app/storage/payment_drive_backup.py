"""Backup distributor payment ledgers to Drive/NEXORA/Payment Receiving.

One readable Excel per user (replaced on each update — no duplicate files).
Best-effort: never fails the API call that triggered the backup.
"""

from __future__ import annotations

import logging
import tempfile
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

PAYMENT_RECEIVING_SUBFOLDER = "Payment Receiving"
_BACKUP_COOLDOWN_SEC = 300
_last_drive_backup_at: dict[tuple[int, str, str], float] = {}


def _safe_stem(label: str | None) -> str:
    text = (label or "nexora").strip()
    for ch in '<>:"/\\|?*':
        text = text.replace(ch, " ")
    return " ".join(text.split()) or "nexora"


def _stable_backup_name(username: str | None, label: str) -> str:
    return f"{_safe_stem(username)} {label}.xlsx"


def _should_run_read_backup(user_id: int | None, workspace_id: str | None, kind: str) -> bool:
    if not user_id:
        return False
    key = (int(user_id), str(workspace_id or "default"), kind)
    now = time.time()
    if now - _last_drive_backup_at.get(key, 0.0) < _BACKUP_COOLDOWN_SEC:
        return False
    _last_drive_backup_at[key] = now
    return True


def _mark_backup_done(user_id: int | None, workspace_id: str | None, kind: str) -> None:
    if user_id:
        _last_drive_backup_at[(int(user_id), str(workspace_id or "default"), kind)] = time.time()


def _style_header_row(ws) -> None:
    try:
        from openpyxl.styles import Font

        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold
    except Exception:
        pass


def _write_temp_workbook(wb) -> Path | None:
    try:
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, prefix="nexora_payment_")
        path = Path(tmp.name)
        tmp.close()
        wb.save(path)
        return path
    except OSError:
        logger.exception("Could not write payment backup workbook")
        return None


def _build_so_payment_workbook(
    distributors: list[dict[str, Any]],
    *,
    exported_at: str,
    workspace_id: str | None,
    user_id: int | None,
):
    from openpyxl import Workbook

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Exported at", exported_at])
    summary.append(["Workspace", workspace_id or ""])
    summary.append(["User id", user_id or ""])
    summary.append(["Distributors", len(distributors)])
    so_bill_total = round(sum(float(d.get("so_bill_total") or 0) for d in distributors), 2)
    paid_total = round(sum(float(d.get("paid_total") or 0) for d in distributors), 2)
    outstanding_total = round(sum(float(d.get("outstanding_total") or 0) for d in distributors), 2)
    summary.append(["SO bill total", so_bill_total])
    summary.append(["Paid total", paid_total])
    summary.append(["Outstanding total", outstanding_total])

    ws = wb.create_sheet("Payment lines")
    ws.append(
        [
            "Distributor",
            "SO / Order ref",
            "Payment date",
            "Amount",
            "Note",
            "SO bill",
            "Paid on SO",
            "Outstanding on SO",
            "Recorded at",
        ]
    )
    for dist in distributors:
        dist_name = dist.get("distributor_name") or ""
        for order in dist.get("orders") or []:
            order_ref = order.get("order_ref_no") or ""
            bill = order.get("so_bill_amount")
            paid_on_so = order.get("paid_amount")
            outstanding = order.get("outstanding")
            payments = order.get("payments") or []
            if not payments:
                continue
            for payment in payments:
                ws.append(
                    [
                        dist_name,
                        order_ref or payment.get("order_ref_no") or "",
                        payment.get("payment_date") or "",
                        payment.get("amount"),
                        payment.get("note") or "",
                        bill,
                        paid_on_so,
                        outstanding,
                        payment.get("created_at") or "",
                    ]
                )
    _style_header_row(ws)
    return wb


def _build_category_payment_workbook(
    distributors: list[dict[str, Any]],
    *,
    exported_at: str,
    workspace_id: str | None,
    user_id: int | None,
):
    from openpyxl import Workbook

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Exported at", exported_at])
    summary.append(["Workspace", workspace_id or ""])
    summary.append(["User id", user_id or ""])
    summary.append(["Distributors", len(distributors)])

    ws = wb.create_sheet("Deposit lines")
    ws.append(
        [
            "Distributor",
            "Season",
            "Category",
            "Payment date",
            "Amount",
            "Note",
            "SO total",
            "Bill after CD",
            "Paid total",
            "Outstanding",
            "Recorded at",
        ]
    )
    for dist in distributors:
        dist_name = dist.get("distributor_name") or ""
        for season_entry in dist.get("seasons") or []:
            season = season_entry.get("season") or ""
            for cat in season_entry.get("categories") or []:
                category = cat.get("category") or ""
                so_total = cat.get("so_total")
                bill_after_cd = cat.get("bill_after_cd")
                paid_total = cat.get("paid_total")
                outstanding = cat.get("outstanding")
                deposits = cat.get("deposits") or []
                if not deposits:
                    continue
                for dep in deposits:
                    ws.append(
                        [
                            dist_name,
                            season,
                            category,
                            dep.get("payment_date") or "",
                            dep.get("amount"),
                            dep.get("note") or "",
                            so_total,
                            bill_after_cd,
                            paid_total,
                            outstanding,
                            dep.get("created_at") or "",
                        ]
                    )
    _style_header_row(ws)
    return wb


def _push_snapshot(
    *,
    user_id: int | None,
    workspace_id: str | None,
    local_path: Path,
    display_name: str,
) -> None:
    try:
        from app.storage.nexora_docs import push_file_to_nexora_drive

        push_file_to_nexora_drive(
            user_id=user_id,
            workspace_id=workspace_id,
            local_path=local_path,
            subfolder=PAYMENT_RECEIVING_SUBFOLDER,
            display_name=display_name,
            replace_if_exists=True,
        )
    finally:
        try:
            local_path.unlink(missing_ok=True)
        except OSError:
            pass


def backup_so_payment_collection_to_drive(
    *,
    db: Any,
    user_id: int | None,
    workspace_id: str | None,
    username: str | None = None,
    allow_read_cooldown: bool = False,
) -> None:
    """Order Desk → Distributor's Payment Collection (SO-wise deposits)."""
    if not user_id or not workspace_id:
        return
    if allow_read_cooldown and not _should_run_read_backup(user_id, workspace_id, "so"):
        return
    try:
        distributors = db.list_distributor_payment_collection(
            workspace_id, user_id=user_id,
        )
        exported_at = datetime.now(timezone.utc).isoformat()
        wb = _build_so_payment_workbook(
            distributors,
            exported_at=exported_at,
            workspace_id=workspace_id,
            user_id=user_id,
        )
        path = _write_temp_workbook(wb)
        if not path:
            return
        name = _stable_backup_name(username, "SO Payment Receiving")
        _push_snapshot(
            user_id=user_id,
            workspace_id=workspace_id,
            local_path=path,
            display_name=name,
        )
        _mark_backup_done(user_id, workspace_id, "so")
    except Exception:
        logger.exception("SO payment receiving Drive backup failed")


def backup_category_payment_status_to_drive(
    *,
    db: Any,
    user_id: int | None,
    workspace_id: str | None,
    username: str | None = None,
    allow_read_cooldown: bool = False,
) -> None:
    """Drawer → Distributors Payment Status (season/category deposits)."""
    if not user_id:
        return
    if allow_read_cooldown and not _should_run_read_backup(user_id, workspace_id, "category"):
        return
    try:
        distributors = db.list_distributor_category_payment_status(user_id)
        exported_at = datetime.now(timezone.utc).isoformat()
        wb = _build_category_payment_workbook(
            distributors,
            exported_at=exported_at,
            workspace_id=workspace_id,
            user_id=user_id,
        )
        path = _write_temp_workbook(wb)
        if not path:
            return
        name = _stable_backup_name(username, "Distributor Payment Status")
        _push_snapshot(
            user_id=user_id,
            workspace_id=workspace_id,
            local_path=path,
            display_name=name,
        )
        _mark_backup_done(user_id, workspace_id, "category")
    except Exception:
        logger.exception("Category payment status Drive backup failed")
