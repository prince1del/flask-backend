"""
Call List / MBO outreach — user-scoped prospect sheets.

Standalone from Party Master / Market Visit / Approach. Optional
linked_distributor_id is stored NULL for now (future link).
"""

from __future__ import annotations

import io
import json
import re
import sqlite3
from datetime import datetime, timezone
from typing import Any

from flask import Blueprint, Response, current_app, jsonify, request, send_file
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename

from app.routes.auth import get_request_user_id, get_workspace_id, require_jwt_auth, require_role

call_lists_bp = Blueprint("call_lists", __name__, url_prefix="/api/v1/call-lists")

CALL_STATUSES = {
    "pending",
    "called",
    "currently_working",
    "follow_up",
    "not_interested",
    "wrong_number",
    "deal_done",
    "no_answer",
}

# Prefer non-pending when deriving a single primary for sort / legacy column.
_PRIMARY_STATUS_ORDER = (
    "deal_done",
    "not_interested",
    "wrong_number",
    "follow_up",
    "currently_working",
    "called",
    "no_answer",
    "pending",
)

_STATUS_LABELS = {
    "pending": "Pending",
    "called": "Called",
    "currently_working": "Currently working",
    "follow_up": "Follow up",
    "not_interested": "Not interested",
    "wrong_number": "Wrong number",
    "no_answer": "No answer",
    "deal_done": "Deal done",
}


def _status_labels_joined(statuses: list[str]) -> str:
    return ", ".join(
        _STATUS_LABELS.get(s, s.replace("_", " ").title()) for s in statuses
    )


def _db_path() -> str:
    return current_app.config.get("DATABASE_PATH", "centralized_db.sqlite3")


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _require_user() -> tuple[int | None, str | None, tuple | None]:
    uid = get_request_user_id()
    if uid is None:
        return None, None, (
            jsonify(
                {
                    "success": False,
                    "error": {"code": "NO_USER", "message": "User id required"},
                }
            ),
            401,
        )
    return uid, get_workspace_id(), None


def _ensure_tables(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS call_list_batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            workspace_id TEXT NOT NULL,
            user_id INTEGER NOT NULL,
            title TEXT NOT NULL,
            source_filename TEXT,
            category_hint TEXT,
            row_count INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS call_list_rows (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id INTEGER NOT NULL,
            workspace_id TEXT NOT NULL,
            user_id INTEGER NOT NULL,
            serial_no INTEGER,
            shop_name TEXT,
            address TEXT,
            town_city TEXT,
            district TEXT,
            state TEXT,
            pin_code TEXT,
            phone TEXT,
            email TEXT,
            call_status TEXT NOT NULL DEFAULT 'pending',
            profile_notes TEXT,
            brands_carried TEXT,
            deals_in TEXT,
            feedback_note TEXT,
            last_called_at TEXT,
            -- Future: optional link to Party Master (never required)
            linked_distributor_id INTEGER,
            linked_party_kind TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(batch_id) REFERENCES call_list_batches(id) ON DELETE CASCADE
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_call_list_batches_user "
        "ON call_list_batches(user_id, workspace_id, id DESC)"
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_call_list_rows_batch "
        "ON call_list_rows(batch_id, user_id, call_status)"
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_call_list_rows_geo "
        "ON call_list_rows(batch_id, state, town_city, district)"
    )
    # Multi-select tags (JSON array). Primary call_status kept for sort / legacy.
    cols = {
        r[1]
        for r in conn.execute("PRAGMA table_info(call_list_rows)").fetchall()
    }
    if "call_statuses_json" not in cols:
        conn.execute(
            "ALTER TABLE call_list_rows ADD COLUMN call_statuses_json TEXT"
        )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS call_list_brands (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            workspace_id TEXT NOT NULL,
            user_id INTEGER NOT NULL,
            brand_name TEXT NOT NULL,
            brand_key TEXT NOT NULL,
            created_at TEXT NOT NULL,
            UNIQUE(user_id, workspace_id, brand_key)
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_call_list_brands_user "
        "ON call_list_brands(user_id, workspace_id, brand_name COLLATE NOCASE)"
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS call_list_deals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            workspace_id TEXT NOT NULL,
            user_id INTEGER NOT NULL,
            deal_name TEXT NOT NULL,
            deal_key TEXT NOT NULL,
            created_at TEXT NOT NULL,
            UNIQUE(user_id, workspace_id, deal_key)
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_call_list_deals_user "
        "ON call_list_deals(user_id, workspace_id, deal_name COLLATE NOCASE)"
    )


def _brand_key(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip().lower())


def _split_brand_names(raw: Any) -> list[str]:
    """Parse brands from list or comma/semicolon/newline text."""
    items: list[str] = []
    if raw is None:
        return []
    if isinstance(raw, (list, tuple)):
        items = [str(x).strip() for x in raw if str(x).strip()]
    else:
        text = str(raw).strip()
        if not text:
            return []
        if text.startswith("["):
            try:
                parsed = json.loads(text)
                if isinstance(parsed, list):
                    items = [str(x).strip() for x in parsed if str(x).strip()]
                else:
                    items = [text]
            except (TypeError, ValueError, json.JSONDecodeError):
                items = re.split(r"[,;\n]+", text)
        else:
            items = re.split(r"[,;\n]+", text)
    out: list[str] = []
    seen: set[str] = set()
    for name in items:
        clean = re.sub(r"\s+", " ", str(name).strip())
        if not clean:
            continue
        key = _brand_key(clean)
        if key in seen:
            continue
        seen.add(key)
        out.append(clean)
    return out


def _upsert_call_list_brands(
    conn: sqlite3.Connection,
    user_id: int,
    workspace_id: str,
    names: list[str],
) -> None:
    now = _now_iso()
    for name in names:
        key = _brand_key(name)
        if not key:
            continue
        conn.execute(
            """
            INSERT INTO call_list_brands (
                workspace_id, user_id, brand_name, brand_key, created_at
            ) VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(user_id, workspace_id, brand_key) DO NOTHING
            """,
            (workspace_id, user_id, name.strip(), key, now),
        )


def _user_brand_options(
    conn: sqlite3.Connection, user_id: int, workspace_id: str
) -> list[str]:
    by_key: dict[str, str] = {}
    for (name,) in conn.execute(
        """
        SELECT brand_name FROM call_list_brands
        WHERE user_id = ? AND workspace_id = ?
        ORDER BY brand_name COLLATE NOCASE
        """,
        (user_id, workspace_id),
    ).fetchall():
        clean = str(name or "").strip()
        key = _brand_key(clean)
        if key and key not in by_key:
            by_key[key] = clean
    for (raw,) in conn.execute(
        """
        SELECT brands_carried FROM call_list_rows
        WHERE user_id = ? AND workspace_id = ?
          AND IFNULL(TRIM(brands_carried),'') != ''
        """,
        (user_id, workspace_id),
    ).fetchall():
        for name in _split_brand_names(raw):
            key = _brand_key(name)
            if key and key not in by_key:
                by_key[key] = name
    return sorted(by_key.values(), key=lambda s: s.casefold())


def _upsert_call_list_deals(
    conn: sqlite3.Connection,
    user_id: int,
    workspace_id: str,
    names: list[str],
) -> None:
    now = _now_iso()
    for name in names:
        key = _brand_key(name)
        if not key:
            continue
        conn.execute(
            """
            INSERT INTO call_list_deals (
                workspace_id, user_id, deal_name, deal_key, created_at
            ) VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(user_id, workspace_id, deal_key) DO NOTHING
            """,
            (workspace_id, user_id, name.strip(), key, now),
        )


def _user_deal_options(
    conn: sqlite3.Connection, user_id: int, workspace_id: str
) -> list[str]:
    by_key: dict[str, str] = {}
    for (name,) in conn.execute(
        """
        SELECT deal_name FROM call_list_deals
        WHERE user_id = ? AND workspace_id = ?
        ORDER BY deal_name COLLATE NOCASE
        """,
        (user_id, workspace_id),
    ).fetchall():
        clean = str(name or "").strip()
        key = _brand_key(clean)
        if key and key not in by_key:
            by_key[key] = clean
    for (raw,) in conn.execute(
        """
        SELECT deals_in FROM call_list_rows
        WHERE user_id = ? AND workspace_id = ?
          AND IFNULL(TRIM(deals_in),'') != ''
        """,
        (user_id, workspace_id),
    ).fetchall():
        for name in _split_brand_names(raw):
            key = _brand_key(name)
            if key and key not in by_key:
                by_key[key] = name
    return sorted(by_key.values(), key=lambda s: s.casefold())


def _normalize_statuses(raw: Any) -> list[str]:
    """Accept list, comma string, or single status → unique valid tags."""
    items: list[str] = []
    if raw is None:
        return ["pending"]
    if isinstance(raw, (list, tuple)):
        items = [str(x).strip().lower() for x in raw if str(x).strip()]
    else:
        text = str(raw).strip()
        if not text:
            return ["pending"]
        if text.startswith("["):
            try:
                parsed = json.loads(text)
                if isinstance(parsed, list):
                    items = [str(x).strip().lower() for x in parsed if str(x).strip()]
                else:
                    items = [text.lower()]
            except (TypeError, ValueError, json.JSONDecodeError):
                items = [s.strip().lower() for s in text.split(",") if s.strip()]
        else:
            items = [s.strip().lower() for s in text.split(",") if s.strip()]
    out: list[str] = []
    for s in items:
        if s in CALL_STATUSES and s not in out:
            out.append(s)
    if not out:
        return ["pending"]
    # Pending alone; if mixed with others, drop pending.
    if len(out) > 1 and "pending" in out:
        out = [s for s in out if s != "pending"]
    # Any outcome after outreach implies the shop was called.
    if out != ["pending"] and "called" not in out:
        out.append("called")
    return out or ["pending"]


def _primary_status(statuses: list[str]) -> str:
    for pref in _PRIMARY_STATUS_ORDER:
        if pref in statuses:
            return pref
    return statuses[0] if statuses else "pending"


def _statuses_from_row(row: sqlite3.Row) -> list[str]:
    raw_json = None
    try:
        raw_json = row["call_statuses_json"]
    except (IndexError, KeyError):
        raw_json = None
    if raw_json:
        return _normalize_statuses(raw_json)
    return _normalize_statuses(row["call_status"] if "call_status" in row.keys() else "pending")


def _statuses_json(statuses: list[str]) -> str:
    return json.dumps(_normalize_statuses(statuses), separators=(",", ":"))


def _parse_statuses_payload(raw: Any) -> tuple[list[str] | None, str | None]:
    """Return (statuses, error_message)."""
    if raw is None:
        return None, "Missing status"
    if isinstance(raw, (list, tuple)):
        candidates = [str(x).strip().lower() for x in raw if str(x).strip()]
    else:
        text = str(raw).strip()
        if text.startswith("["):
            try:
                parsed = json.loads(text)
                if isinstance(parsed, list):
                    candidates = [
                        str(x).strip().lower() for x in parsed if str(x).strip()
                    ]
                else:
                    candidates = [text.lower()] if text else []
            except (TypeError, ValueError, json.JSONDecodeError):
                candidates = [s.strip().lower() for s in text.split(",") if s.strip()]
        else:
            candidates = [s.strip().lower() for s in text.split(",") if s.strip()]
    bad = [s for s in candidates if s not in CALL_STATUSES]
    if bad:
        return None, f"Invalid status. Use: {', '.join(sorted(CALL_STATUSES))}"
    return _normalize_statuses(candidates), None


def _norm_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _map_headers(row: tuple[Any, ...]) -> dict[str, int]:
    aliases = {
        "serial_no": {"s no", "sno", "s n", "serial", "sr no", "sr", "no"},
        "shop_name": {
            "shop name",
            "firm name",
            "store name",
            "name",
            "party name",
            "outlet",
            "outlet name",
            "name of retailer",
            "retailer name",
            "retailer",
            "customer name",
            "account name",
            "mbo",
            "mbo name",
            "shop",
        },
        "address": {"address", "addr", "address details for expansion"},
        "town_city": {
            "town city",
            "town",
            "city",
            "city town",
            "area",
            "location",
        },
        "district": {"district", "dist"},
        "state": {"state"},
        "pin_code": {"pin code", "pincode", "pin", "zip"},
        "phone": {
            "phone number",
            "phone",
            "mobile",
            "contact",
            "contact no",
            "contact number",
            "mobile number",
            "contact nos",
        },
        "email": {"email", "e mail", "mail", "email id", "emailid", "e mail id"},
    }
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(row):
        key = _norm_header(cell)
        if not key:
            continue
        for field, names in aliases.items():
            if key in names and field not in mapping:
                mapping[field] = idx
                break
        if "shop_name" not in mapping and _looks_like_shop_header(key):
            mapping["shop_name"] = idx
    return mapping


def _looks_like_shop_header(key: str) -> bool:
    if not key or "owner" in key:
        return False
    has_name = "name" in key
    return has_name and any(
        token in key for token in ("retailer", "shop", "store", "firm", "outlet", "party", "mbo")
    )


def _cell(row: tuple[Any, ...], idx: int | None) -> str | None:
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    text = str(val).strip()
    return text or None


def _normalize_phone_india(raw: str | None) -> str | None:
    """Canonical display for India mobiles and landlines from Excel (+91 …)."""
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) < 7:
        return text

    if digits.startswith("00"):
        digits = digits[2:]

    core: str | None = None
    if digits.startswith("91") and len(digits) >= 11:
        core = digits
    elif len(digits) == 10 and digits[0] in "6789":
        core = "91" + digits
    elif digits.startswith("0") and len(digits) >= 10:
        core = "91" + digits[1:]
    elif len(digits) in (10, 11) and digits[0] in "12345":
        core = "91" + digits

    if not core:
        return text

    national = core[2:]
    if len(national) == 10 and national[0] in "6789":
        return f"+91 {national[:5]} {national[5:]}"
    if len(national) >= 10:
        return f"+91 {national[:3]} {national[3:6]} {national[6:]}"
    return f"+91 {national}"


def _parse_workbook(file_bytes: bytes) -> tuple[str | None, list[dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    category_hint = None
    if rows:
        first = rows[0]
        if first and first[0] and not _norm_header(first[0]).startswith("s no"):
            hint = str(first[0]).strip()
            if len(hint) > 3:
                category_hint = hint[:240]

    header_idx = None
    mapping: dict[str, int] = {}
    for i, row in enumerate(rows[:40]):
        if not row:
            continue
        m = _map_headers(tuple(row))
        if "shop_name" in m or "phone" in m:
            header_idx = i
            mapping = m
            break
    if header_idx is None:
        raise ValueError(
            "Could not find header row (need Shop Name / Phone columns)"
        )

    parsed: list[dict[str, Any]] = []
    for row in rows[header_idx + 1 :]:
        if not row or not any(c is not None and str(c).strip() for c in row):
            continue
        shop = _cell(row, mapping.get("shop_name"))
        phone = _normalize_phone_india(_cell(row, mapping.get("phone")))
        if not shop and not phone:
            continue
        serial_raw = _cell(row, mapping.get("serial_no"))
        serial_no = None
        if serial_raw:
            try:
                serial_no = int(float(serial_raw))
            except (TypeError, ValueError):
                serial_no = None
        parsed.append(
            {
                "serial_no": serial_no,
                "shop_name": shop,
                "address": _cell(row, mapping.get("address")),
                "town_city": _cell(row, mapping.get("town_city")),
                "district": _cell(row, mapping.get("district")),
                "state": _cell(row, mapping.get("state")),
                "pin_code": _cell(row, mapping.get("pin_code")),
                "phone": phone,
                "email": _cell(row, mapping.get("email")),
            }
        )
    if not parsed:
        raise ValueError("No shop rows found in the spreadsheet")
    return category_hint, parsed


def _batch_dict(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "title": row["title"],
        "source_filename": row["source_filename"],
        "category_hint": row["category_hint"],
        "row_count": row["row_count"],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def _row_dict(row: sqlite3.Row) -> dict[str, Any]:
    statuses = _statuses_from_row(row)
    return {
        "id": row["id"],
        "batch_id": row["batch_id"],
        "serial_no": row["serial_no"],
        "shop_name": row["shop_name"],
        "address": row["address"],
        "town_city": row["town_city"],
        "district": row["district"],
        "state": row["state"],
        "pin_code": row["pin_code"],
        "phone": row["phone"],
        "email": row["email"],
        "call_status": _primary_status(statuses),
        "call_statuses": statuses,
        "profile_notes": row["profile_notes"],
        "brands_carried": row["brands_carried"],
        "deals_in": row["deals_in"],
        "feedback_note": row["feedback_note"],
        "last_called_at": row["last_called_at"],
        "linked_distributor_id": row["linked_distributor_id"],
        "linked_party_kind": row["linked_party_kind"],
        "updated_at": row["updated_at"],
    }


def _get_owned_batch(
    conn: sqlite3.Connection, batch_id: int, user_id: int, workspace_id: str
) -> sqlite3.Row | None:
    return conn.execute(
        """
        SELECT * FROM call_list_batches
        WHERE id = ? AND user_id = ? AND workspace_id = ?
        """,
        (batch_id, user_id, workspace_id),
    ).fetchone()


@call_lists_bp.route("", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def list_batches():
    uid, workspace_id, err = _require_user()
    if err:
        return err
    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_tables(conn)
        rows = conn.execute(
            """
            SELECT * FROM call_list_batches
            WHERE user_id = ? AND workspace_id = ?
            ORDER BY id DESC
            """,
            (uid, workspace_id),
        ).fetchall()
    return jsonify({"success": True, "data": [_batch_dict(r) for r in rows]}), 200


@call_lists_bp.route("/upload", methods=["POST"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def upload_batch():
    uid, workspace_id, err = _require_user()
    if err:
        return err
    file = request.files.get("file")
    if file is None or not file.filename:
        return (
            jsonify(
                {
                    "success": False,
                    "error": {"message": "Excel file required (field: file)"},
                }
            ),
            400,
        )
    filename = secure_filename(file.filename) or "call_list.xlsx"
    raw = file.read()
    if not raw:
        return jsonify({"success": False, "error": {"message": "Empty file"}}), 400

    title = (request.form.get("title") or "").strip()
    if not title:
        title = filename.rsplit(".", 1)[0].replace("_", " ").strip() or "Call List"

    try:
        category_hint, parsed = _parse_workbook(raw)
    except Exception as exc:
        return (
            jsonify({"success": False, "error": {"message": str(exc)}}),
            400,
        )

    now = _now_iso()
    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_tables(conn)
        cur = conn.execute(
            """
            INSERT INTO call_list_batches (
                workspace_id, user_id, title, source_filename, category_hint,
                row_count, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                workspace_id,
                uid,
                title,
                filename,
                category_hint,
                len(parsed),
                now,
                now,
            ),
        )
        batch_id = int(cur.lastrowid)
        conn.executemany(
            """
            INSERT INTO call_list_rows (
                batch_id, workspace_id, user_id, serial_no, shop_name, address,
                town_city, district, state, pin_code, phone, email,
                call_status, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?, ?)
            """,
            [
                (
                    batch_id,
                    workspace_id,
                    uid,
                    r["serial_no"],
                    r["shop_name"],
                    r["address"],
                    r["town_city"],
                    r["district"],
                    r["state"],
                    r["pin_code"],
                    r["phone"],
                    r["email"],
                    now,
                    now,
                )
                for r in parsed
            ],
        )
        conn.commit()
        batch = _get_owned_batch(conn, batch_id, uid, workspace_id)

    return jsonify({"success": True, "data": _batch_dict(batch)}), 201


@call_lists_bp.route("/brands", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def list_brands():
    uid, workspace_id, err = _require_user()
    if err:
        return err
    with sqlite3.connect(_db_path()) as conn:
        _ensure_tables(conn)
        brands = _user_brand_options(conn, uid, workspace_id)
    return jsonify({"success": True, "data": brands, "count": len(brands)}), 200


@call_lists_bp.route("/brands", methods=["POST"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def add_brand():
    uid, workspace_id, err = _require_user()
    if err:
        return err
    payload = request.get_json(silent=True) or {}
    name = (payload.get("brand_name") or payload.get("name") or "").strip()
    name = re.sub(r"\s+", " ", name)
    if not name:
        return (
            jsonify({"success": False, "error": {"message": "brand_name is required"}}),
            400,
        )
    with sqlite3.connect(_db_path()) as conn:
        _ensure_tables(conn)
        _upsert_call_list_brands(conn, uid, workspace_id, [name])
        conn.commit()
        brands = _user_brand_options(conn, uid, workspace_id)
    return jsonify({"success": True, "data": brands, "count": len(brands)}), 201


@call_lists_bp.route("/deals", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def list_deals():
    uid, workspace_id, err = _require_user()
    if err:
        return err
    with sqlite3.connect(_db_path()) as conn:
        _ensure_tables(conn)
        deals = _user_deal_options(conn, uid, workspace_id)
    return jsonify({"success": True, "data": deals, "count": len(deals)}), 200


@call_lists_bp.route("/deals", methods=["POST"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def add_deal():
    uid, workspace_id, err = _require_user()
    if err:
        return err
    payload = request.get_json(silent=True) or {}
    name = (payload.get("deal_name") or payload.get("name") or "").strip()
    name = re.sub(r"\s+", " ", name)
    if not name:
        return (
            jsonify({"success": False, "error": {"message": "deal_name is required"}}),
            400,
        )
    with sqlite3.connect(_db_path()) as conn:
        _ensure_tables(conn)
        _upsert_call_list_deals(conn, uid, workspace_id, [name])
        conn.commit()
        deals = _user_deal_options(conn, uid, workspace_id)
    return jsonify({"success": True, "data": deals, "count": len(deals)}), 201


@call_lists_bp.route("/<int:batch_id>", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def get_batch(batch_id: int):
    uid, workspace_id, err = _require_user()
    if err:
        return err
    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_tables(conn)
        batch = _get_owned_batch(conn, batch_id, uid, workspace_id)
        if not batch:
            return (
                jsonify({"success": False, "error": {"message": "List not found"}}),
                404,
            )
        states = [
            r[0]
            for r in conn.execute(
                """
                SELECT DISTINCT state FROM call_list_rows
                WHERE batch_id = ? AND user_id = ? AND IFNULL(TRIM(state),'') != ''
                ORDER BY state COLLATE NOCASE
                """,
                (batch_id, uid),
            ).fetchall()
        ]
        cities = [
            r[0]
            for r in conn.execute(
                """
                SELECT DISTINCT town_city FROM call_list_rows
                WHERE batch_id = ? AND user_id = ?
                  AND IFNULL(TRIM(town_city),'') != ''
                ORDER BY town_city COLLATE NOCASE
                """,
                (batch_id, uid),
            ).fetchall()
        ]
        # Linked State → Cities map for cascading dropdowns on mobile.
        cities_by_state: dict[str, list[str]] = {}
        for st, city in conn.execute(
            """
            SELECT DISTINCT state, town_city FROM call_list_rows
            WHERE batch_id = ? AND user_id = ?
              AND IFNULL(TRIM(state),'') != ''
              AND IFNULL(TRIM(town_city),'') != ''
            ORDER BY state COLLATE NOCASE, town_city COLLATE NOCASE
            """,
            (batch_id, uid),
        ).fetchall():
            key = str(st).strip()
            val = str(city).strip()
            if not key or not val:
                continue
            bucket = cities_by_state.setdefault(key, [])
            if val not in bucket:
                bucket.append(val)
        districts = [
            r[0]
            for r in conn.execute(
                """
                SELECT DISTINCT district FROM call_list_rows
                WHERE batch_id = ? AND user_id = ?
                  AND IFNULL(TRIM(district),'') != ''
                ORDER BY district COLLATE NOCASE
                """,
                (batch_id, uid),
            ).fetchall()
        ]
        status_rows = conn.execute(
            """
            SELECT call_status, call_statuses_json FROM call_list_rows
            WHERE batch_id = ? AND user_id = ?
            """,
            (batch_id, uid),
        ).fetchall()
        status_counts: dict[str, int] = {}
        for r in status_rows:
            for tag in _statuses_from_row(r):
                status_counts[tag] = status_counts.get(tag, 0) + 1
        brand_options = _user_brand_options(conn, uid, workspace_id)
        deal_options = _user_deal_options(conn, uid, workspace_id)

    data = _batch_dict(batch)
    data["filters"] = {
        "states": states,
        "cities": cities,
        "cities_by_state": cities_by_state,
        "districts": districts,
        "status_counts": status_counts,
        "brand_options": brand_options,
        "deal_options": deal_options,
    }
    return jsonify({"success": True, "data": data}), 200


@call_lists_bp.route("/<int:batch_id>/rows", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def list_rows(batch_id: int):
    uid, workspace_id, err = _require_user()
    if err:
        return err
    limit = min(max(request.args.get("limit", 100, type=int) or 100, 1), 500)
    offset = max(request.args.get("offset", 0, type=int) or 0, 0)
    state = (request.args.get("state") or "").strip()
    city = (request.args.get("city") or "").strip()
    district = (request.args.get("district") or "").strip()
    status = (request.args.get("status") or "").strip().lower()
    q = (request.args.get("q") or "").strip()
    status_filters = [
        s.strip() for s in status.split(",") if s.strip() and s.strip() in CALL_STATUSES
    ]

    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_tables(conn)
        batch = _get_owned_batch(conn, batch_id, uid, workspace_id)
        if not batch:
            return (
                jsonify({"success": False, "error": {"message": "List not found"}}),
                404,
            )
        where = ["batch_id = ?", "user_id = ?"]
        params: list[Any] = [batch_id, uid]
        region_where = list(where)
        region_params: list[Any] = list(params)
        if state:
            where.append("LOWER(IFNULL(state,'')) = LOWER(?)")
            params.append(state)
            region_where.append("LOWER(IFNULL(state,'')) = LOWER(?)")
            region_params.append(state)
        if city:
            where.append("LOWER(IFNULL(town_city,'')) = LOWER(?)")
            params.append(city)
            region_where.append("LOWER(IFNULL(town_city,'')) = LOWER(?)")
            region_params.append(city)
        if district:
            where.append("LOWER(IFNULL(district,'')) = LOWER(?)")
            params.append(district)
            region_where.append("LOWER(IFNULL(district,'')) = LOWER(?)")
            region_params.append(district)
        if status_filters:
            # One chip at a time — match primary call_status (outcome), not every tag.
            if len(status_filters) == 1:
                s = status_filters[0]
                where.append("LOWER(IFNULL(call_status,'pending')) = ?")
                params.append(s)
            else:
                clauses: list[str] = []
                for s in status_filters:
                    clauses.append("LOWER(IFNULL(call_status,'pending')) = ?")
                    params.append(s)
                where.append("(" + " OR ".join(clauses) + ")")
        if q:
            where.append(
                "("
                "IFNULL(shop_name,'') LIKE ? OR IFNULL(phone,'') LIKE ? OR "
                "IFNULL(town_city,'') LIKE ? OR IFNULL(address,'') LIKE ?"
                ")"
            )
            like = f"%{q}%"
            params.extend([like, like, like, like])
            region_where.append(
                "("
                "IFNULL(shop_name,'') LIKE ? OR IFNULL(phone,'') LIKE ? OR "
                "IFNULL(town_city,'') LIKE ? OR IFNULL(address,'') LIKE ?"
                ")"
            )
            region_params.extend([like, like, like, like])
        where_sql = " AND ".join(where)
        total = conn.execute(
            f"SELECT COUNT(*) FROM call_list_rows WHERE {where_sql}",
            tuple(params),
        ).fetchone()[0]
        region_total: int | None = None
        if state or city or district or q:
            region_sql = " AND ".join(region_where)
            region_total = int(
                conn.execute(
                    f"SELECT COUNT(*) FROM call_list_rows WHERE {region_sql}",
                    tuple(region_params),
                ).fetchone()[0]
            )
        rows = conn.execute(
            f"""
            SELECT * FROM call_list_rows
            WHERE {where_sql}
            ORDER BY
                CASE WHEN IFNULL(call_status,'pending') = 'pending' THEN 0 ELSE 1 END,
                IFNULL(serial_no, 999999),
                id
            LIMIT ? OFFSET ?
            """,
            tuple(params + [limit, offset]),
        ).fetchall()

    return (
        jsonify(
            {
                "success": True,
                "data": {
                    "total": int(total),
                    "region_total": region_total,
                    "limit": limit,
                    "offset": offset,
                    "rows": [_row_dict(r) for r in rows],
                },
            }
        ),
        200,
    )


@call_lists_bp.route("/rows/<int:row_id>", methods=["PATCH"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def update_row(row_id: int):
    uid, workspace_id, err = _require_user()
    if err:
        return err
    payload = request.get_json(silent=True) or {}
    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_tables(conn)
        row = conn.execute(
            """
            SELECT * FROM call_list_rows
            WHERE id = ? AND user_id = ? AND workspace_id = ?
            """,
            (row_id, uid, workspace_id),
        ).fetchone()
        if not row:
            return (
                jsonify({"success": False, "error": {"message": "Row not found"}}),
                404,
            )

        sets: list[str] = []
        params: list[Any] = []
        statuses_payload = payload.get("call_statuses")
        status_single = payload.get("call_status")
        if statuses_payload is not None or status_single is not None:
            statuses, status_err = _parse_statuses_payload(
                statuses_payload if statuses_payload is not None else status_single
            )
            if status_err or not statuses:
                return (
                    jsonify(
                        {
                            "success": False,
                            "error": {
                                "message": status_err
                                or f"Invalid status. Use: {', '.join(sorted(CALL_STATUSES))}"
                            },
                        }
                    ),
                    400,
                )
            primary = _primary_status(statuses)
            sets.append("call_status = ?")
            params.append(primary)
            sets.append("call_statuses_json = ?")
            params.append(_statuses_json(statuses))
            if any(s != "pending" for s in statuses):
                sets.append("last_called_at = COALESCE(last_called_at, ?)")
                params.append(_now_iso())

        for field in (
            "profile_notes",
            "feedback_note",
        ):
            if field in payload:
                val = payload.get(field)
                text = None if val is None else str(val).strip() or None
                sets.append(f"{field} = ?")
                params.append(text)

        if "brands_carried" in payload:
            brand_names = _split_brand_names(payload.get("brands_carried"))
            brands_text = ", ".join(brand_names) if brand_names else None
            sets.append("brands_carried = ?")
            params.append(brands_text)
            if brand_names:
                _upsert_call_list_brands(conn, uid, workspace_id, brand_names)

        if "deals_in" in payload:
            deal_names = _split_brand_names(payload.get("deals_in"))
            deals_text = ", ".join(deal_names) if deal_names else None
            sets.append("deals_in = ?")
            params.append(deals_text)
            if deal_names:
                _upsert_call_list_deals(conn, uid, workspace_id, deal_names)

        # Future distributor link — accept but never auto-fill from Party Master.
        if "linked_distributor_id" in payload:
            raw = payload.get("linked_distributor_id")
            linked = None
            if raw is not None and str(raw).strip():
                try:
                    linked = int(raw)
                except (TypeError, ValueError):
                    return (
                        jsonify(
                            {
                                "success": False,
                                "error": {"message": "linked_distributor_id must be int"},
                            }
                        ),
                        400,
                    )
            sets.append("linked_distributor_id = ?")
            params.append(linked)
        if "linked_party_kind" in payload:
            kind = payload.get("linked_party_kind")
            sets.append("linked_party_kind = ?")
            params.append(
                None if kind is None else (str(kind).strip() or None)
            )

        if not sets:
            return jsonify({"success": True, "data": _row_dict(row)}), 200

        now = _now_iso()
        sets.append("updated_at = ?")
        params.append(now)
        params.append(row_id)
        conn.execute(
            f"UPDATE call_list_rows SET {', '.join(sets)} WHERE id = ?",
            tuple(params),
        )
        conn.execute(
            "UPDATE call_list_batches SET updated_at = ? WHERE id = ?",
            (now, row["batch_id"]),
        )
        conn.commit()
        updated = conn.execute(
            "SELECT * FROM call_list_rows WHERE id = ?", (row_id,)
        ).fetchone()

    return jsonify({"success": True, "data": _row_dict(updated)}), 200


@call_lists_bp.route("/<int:batch_id>/export", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def export_batch(batch_id: int):
    uid, workspace_id, err = _require_user()
    if err:
        return err
    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_tables(conn)
        batch = _get_owned_batch(conn, batch_id, uid, workspace_id)
        if not batch:
            return (
                jsonify({"success": False, "error": {"message": "List not found"}}),
                404,
            )
        rows = conn.execute(
            """
            SELECT * FROM call_list_rows
            WHERE batch_id = ? AND user_id = ?
            ORDER BY IFNULL(serial_no, 999999), id
            """,
            (batch_id, uid),
        ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Call List Feedback"
    ws.append(
        [
            "S.No",
            "Shop Name",
            "Address",
            "Town/City",
            "District",
            "State",
            "Pin Code",
            "Phone Number",
            "Email",
            "Call Status",
            "Profile / What they do",
            "Brands carried",
            "Deals in",
            "Feedback note",
            "Last called at",
            "Linked distributor id (future)",
        ]
    )
    for r in rows:
        statuses = _statuses_from_row(r)
        ws.append(
            [
                r["serial_no"],
                r["shop_name"],
                r["address"],
                r["town_city"],
                r["district"],
                r["state"],
                r["pin_code"],
                r["phone"],
                r["email"],
                _status_labels_joined(statuses),
                r["profile_notes"],
                r["brands_carried"],
                r["deals_in"],
                r["feedback_note"],
                r["last_called_at"],
                r["linked_distributor_id"],
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_title = re.sub(r"[^\w\-]+", "_", batch["title"] or "call_list")[:60]
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"{safe_title}_feedback.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@call_lists_bp.route("/<int:batch_id>", methods=["DELETE"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def delete_batch(batch_id: int):
    uid, workspace_id, err = _require_user()
    if err:
        return err
    with sqlite3.connect(_db_path()) as conn:
        _ensure_tables(conn)
        batch = conn.execute(
            """
            SELECT id FROM call_list_batches
            WHERE id = ? AND user_id = ? AND workspace_id = ?
            """,
            (batch_id, uid, workspace_id),
        ).fetchone()
        if not batch:
            return (
                jsonify({"success": False, "error": {"message": "List not found"}}),
                404,
            )
        conn.execute("DELETE FROM call_list_rows WHERE batch_id = ?", (batch_id,))
        conn.execute("DELETE FROM call_list_batches WHERE id = ?", (batch_id,))
        conn.commit()
    return jsonify({"success": True, "data": {"deleted": batch_id}}), 200
