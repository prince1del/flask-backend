"""Monthly Permanent Journey Plan (PJP) — matches BD Excel travel-plan format."""

from __future__ import annotations

import calendar
import re
import sqlite3
import tempfile
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from flask import Blueprint, current_app, jsonify, request

from app.routes.auth import get_request_user_id, get_workspace_id, require_jwt_auth, require_role

pjp_bp = Blueprint("pjp", __name__, url_prefix="/api/v1/pjp")

DAY_TYPES = {"work", "holiday", "leave", "weekend", "blank"}


def _db_path() -> str:
    return current_app.config.get("DATABASE_PATH", "centralized_db.sqlite3")


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _ensure_tables(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS monthly_pjp_meta (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            workspace_id TEXT NOT NULL,
            user_id INTEGER NOT NULL,
            year_month TEXT NOT NULL,
            sm_name TEXT,
            zone TEXT,
            title TEXT,
            note TEXT,
            updated_at TEXT NOT NULL,
            UNIQUE(workspace_id, user_id, year_month)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS monthly_pjp_days (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            workspace_id TEXT NOT NULL,
            user_id INTEGER NOT NULL,
            plan_date TEXT NOT NULL,
            day_name TEXT,
            place_to_visit TEXT,
            from_place TEXT,
            to_place TEXT,
            business_activity TEXT,
            particulars TEXT,
            travel_kms REAL,
            night_stay TEXT,
            day_type TEXT NOT NULL DEFAULT 'work',
            updated_at TEXT NOT NULL,
            UNIQUE(workspace_id, user_id, plan_date)
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_monthly_pjp_days_user_ym "
        "ON monthly_pjp_days(workspace_id, user_id, plan_date)"
    )


def _require_user_id():
    uid = get_request_user_id()
    if uid is None:
        return None, (
            jsonify({"success": False, "error": {"message": "User id required"}}),
            401,
        )
    try:
        return int(uid), None
    except (TypeError, ValueError):
        return None, (
            jsonify({"success": False, "error": {"message": "Invalid user id"}}),
            400,
        )


def _parse_ym(raw: str) -> tuple[int, int] | None:
    text = (raw or "").strip()
    try:
        y, m = text.split("-", 1)
        yi, mi = int(y), int(m)
        if yi < 2000 or yi > 2100 or mi < 1 or mi > 12:
            return None
        return yi, mi
    except Exception:
        return None


def _day_name(d: date) -> str:
    return d.strftime("%A")


def _infer_day_type(d: date, place: str | None, explicit: str | None) -> str:
    if explicit and explicit in DAY_TYPES:
        return explicit
    p = (place or "").strip().lower()
    if p in {"holiday", "holidays"}:
        return "holiday"
    if p in {"leave", "off", "personal leave"}:
        return "leave"
    if d.weekday() >= 5 and not p:
        return "weekend"
    if not p:
        return "blank"
    return "work"


def _meta_dict(row: sqlite3.Row | None, year_month: str) -> dict:
    if not row:
        return {
            "year_month": year_month,
            "sm_name": None,
            "zone": None,
            "title": f"Travel Plan for the month {year_month}",
            "note": None,
        }
    return {
        "year_month": row["year_month"],
        "sm_name": row["sm_name"],
        "zone": row["zone"],
        "title": row["title"],
        "note": row["note"],
    }


def _day_dict(row: sqlite3.Row) -> dict:
    return {
        "id": row["id"],
        "plan_date": row["plan_date"],
        "day_name": row["day_name"],
        "place_to_visit": row["place_to_visit"],
        "from_place": row["from_place"],
        "to_place": row["to_place"],
        "business_activity": row["business_activity"],
        "particulars": row["particulars"],
        "travel_kms": row["travel_kms"],
        "night_stay": row["night_stay"],
        "day_type": row["day_type"],
        "updated_at": row["updated_at"],
    }


def _empty_day(d: date) -> dict:
    dtype = "weekend" if d.weekday() >= 5 else "blank"
    return {
        "id": None,
        "plan_date": d.isoformat(),
        "day_name": _day_name(d),
        "place_to_visit": None,
        "from_place": None,
        "to_place": None,
        "business_activity": None,
        "particulars": None,
        "travel_kms": None,
        "night_stay": None,
        "day_type": dtype,
        "updated_at": None,
    }


def _month_payload(conn: sqlite3.Connection, workspace_id: str, user_id: int, year_month: str) -> dict:
    parsed = _parse_ym(year_month)
    if not parsed:
        raise ValueError("year_month must be YYYY-MM")
    year, month = parsed
    meta = conn.execute(
        """
        SELECT * FROM monthly_pjp_meta
        WHERE workspace_id = ? AND user_id = ? AND year_month = ?
        """,
        (workspace_id, user_id, year_month),
    ).fetchone()
    rows = conn.execute(
        """
        SELECT * FROM monthly_pjp_days
        WHERE workspace_id = ? AND user_id = ?
          AND plan_date >= ? AND plan_date < ?
        ORDER BY plan_date ASC
        """,
        (
            workspace_id,
            user_id,
            f"{year_month}-01",
            f"{year + (1 if month == 12 else 0)}-{1 if month == 12 else month + 1:02d}-01",
        ),
    ).fetchall()
    by_date = {r["plan_date"]: _day_dict(r) for r in rows}
    days_in_month = calendar.monthrange(year, month)[1]
    days = []
    planned = 0
    for day_n in range(1, days_in_month + 1):
        d = date(year, month, day_n)
        entry = by_date.get(d.isoformat()) or _empty_day(d)
        if entry.get("place_to_visit") or entry.get("day_type") in {"holiday", "leave"}:
            if entry.get("day_type") == "work" or entry.get("place_to_visit"):
                if (entry.get("place_to_visit") or "").strip().lower() not in {
                    "holiday",
                    "leave",
                    "",
                }:
                    planned += 1
        days.append(entry)
    return {
        "meta": _meta_dict(meta, year_month),
        "year_month": year_month,
        "days": days,
        "stats": {
            "days_in_month": days_in_month,
            "planned_days": planned,
            "outstation_nights": sum(
                1 for d in days if (d.get("night_stay") or "").strip()
            ),
        },
    }


@pjp_bp.route("/months", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def list_months():
    uid, err = _require_user_id()
    if err:
        return err
    workspace_id = get_workspace_id()
    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_tables(conn)
        rows = conn.execute(
            """
            SELECT DISTINCT substr(plan_date, 1, 7) AS ym
            FROM monthly_pjp_days
            WHERE workspace_id = ? AND user_id = ?
            UNION
            SELECT year_month AS ym FROM monthly_pjp_meta
            WHERE workspace_id = ? AND user_id = ?
            ORDER BY ym DESC
            """,
            (workspace_id, uid, workspace_id, uid),
        ).fetchall()
    return jsonify(
        {
            "success": True,
            "data": {"months": [r["ym"] for r in rows]},
        }
    )


@pjp_bp.route("/months/<year_month>", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def get_month(year_month: str):
    uid, err = _require_user_id()
    if err:
        return err
    if not _parse_ym(year_month):
        return jsonify({"success": False, "error": {"message": "Use YYYY-MM"}}), 400
    workspace_id = get_workspace_id()
    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_tables(conn)
        payload = _month_payload(conn, workspace_id, uid, year_month)
    return jsonify({"success": True, "data": payload})


@pjp_bp.route("/months/<year_month>", methods=["PUT"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def upsert_month_meta(year_month: str):
    uid, err = _require_user_id()
    if err:
        return err
    if not _parse_ym(year_month):
        return jsonify({"success": False, "error": {"message": "Use YYYY-MM"}}), 400
    workspace_id = get_workspace_id()
    data = request.get_json(silent=True) or {}
    now = _now_iso()
    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_tables(conn)
        conn.execute(
            """
            INSERT INTO monthly_pjp_meta (
                workspace_id, user_id, year_month, sm_name, zone, title, note, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(workspace_id, user_id, year_month) DO UPDATE SET
                sm_name = excluded.sm_name,
                zone = excluded.zone,
                title = excluded.title,
                note = excluded.note,
                updated_at = excluded.updated_at
            """,
            (
                workspace_id,
                uid,
                year_month,
                (data.get("sm_name") or "").strip() or None,
                (data.get("zone") or "").strip() or None,
                (data.get("title") or "").strip() or None,
                (data.get("note") or "").strip() or None,
                now,
            ),
        )
        conn.commit()
        payload = _month_payload(conn, workspace_id, uid, year_month)
    return jsonify({"success": True, "data": payload})


@pjp_bp.route("/days/<plan_date>", methods=["PUT"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def upsert_day(plan_date: str):
    uid, err = _require_user_id()
    if err:
        return err
    try:
        d = date.fromisoformat(plan_date.strip())
    except ValueError:
        return jsonify({"success": False, "error": {"message": "plan_date must be YYYY-MM-DD"}}), 400
    workspace_id = get_workspace_id()
    data = request.get_json(silent=True) or {}
    place = (data.get("place_to_visit") or "").strip() or None
    day_type = _infer_day_type(d, place, (data.get("day_type") or "").strip().lower() or None)
    kms = data.get("travel_kms")
    try:
        kms_val = float(kms) if kms is not None and str(kms).strip() != "" else None
    except (TypeError, ValueError):
        return jsonify({"success": False, "error": {"message": "travel_kms must be a number"}}), 400
    now = _now_iso()
    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_tables(conn)
        conn.execute(
            """
            INSERT INTO monthly_pjp_days (
                workspace_id, user_id, plan_date, day_name,
                place_to_visit, from_place, to_place,
                business_activity, particulars, travel_kms, night_stay,
                day_type, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(workspace_id, user_id, plan_date) DO UPDATE SET
                day_name = excluded.day_name,
                place_to_visit = excluded.place_to_visit,
                from_place = excluded.from_place,
                to_place = excluded.to_place,
                business_activity = excluded.business_activity,
                particulars = excluded.particulars,
                travel_kms = excluded.travel_kms,
                night_stay = excluded.night_stay,
                day_type = excluded.day_type,
                updated_at = excluded.updated_at
            """,
            (
                workspace_id,
                uid,
                d.isoformat(),
                _day_name(d),
                place,
                (data.get("from_place") or "").strip() or None,
                (data.get("to_place") or "").strip() or None,
                (data.get("business_activity") or data.get("purpose_of_visit") or "").strip()
                or None,
                (data.get("particulars") or "").strip() or None,
                kms_val,
                (data.get("night_stay") or "").strip() or None,
                day_type,
                now,
            ),
        )
        conn.commit()
        row = conn.execute(
            """
            SELECT * FROM monthly_pjp_days
            WHERE workspace_id = ? AND user_id = ? AND plan_date = ?
            """,
            (workspace_id, uid, d.isoformat()),
        ).fetchone()
    return jsonify({"success": True, "data": _day_dict(row)})


@pjp_bp.route("/days/<plan_date>", methods=["DELETE"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def clear_day(plan_date: str):
    uid, err = _require_user_id()
    if err:
        return err
    try:
        d = date.fromisoformat(plan_date.strip())
    except ValueError:
        return jsonify({"success": False, "error": {"message": "plan_date must be YYYY-MM-DD"}}), 400
    workspace_id = get_workspace_id()
    with sqlite3.connect(_db_path()) as conn:
        _ensure_tables(conn)
        conn.execute(
            """
            DELETE FROM monthly_pjp_days
            WHERE workspace_id = ? AND user_id = ? AND plan_date = ?
            """,
            (workspace_id, uid, d.isoformat()),
        )
        conn.commit()
    return jsonify({"success": True, "data": {"plan_date": d.isoformat(), "cleared": True}})


@pjp_bp.route("/months/<year_month>/bulk", methods=["PUT"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def bulk_upsert_month(year_month: str):
    """Replace/upsert many days for a month (Excel-shaped payload)."""
    uid, err = _require_user_id()
    if err:
        return err
    if not _parse_ym(year_month):
        return jsonify({"success": False, "error": {"message": "Use YYYY-MM"}}), 400
    workspace_id = get_workspace_id()
    data = request.get_json(silent=True) or {}
    days = data.get("days") or []
    if not isinstance(days, list):
        return jsonify({"success": False, "error": {"message": "days must be an array"}}), 400
    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_tables(conn)
        _apply_month_bulk(conn, workspace_id, uid, year_month, data, days)
        conn.commit()
        payload = _month_payload(conn, workspace_id, uid, year_month)
    return jsonify({"success": True, "data": payload})


def _apply_month_bulk(
    conn: sqlite3.Connection,
    workspace_id: str,
    user_id: int,
    year_month: str,
    meta: dict,
    days: list,
) -> None:
    now = _now_iso()
    if any(k in meta for k in ("sm_name", "zone", "title", "note")):
        conn.execute(
            """
            INSERT INTO monthly_pjp_meta (
                workspace_id, user_id, year_month, sm_name, zone, title, note, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(workspace_id, user_id, year_month) DO UPDATE SET
                sm_name = COALESCE(excluded.sm_name, monthly_pjp_meta.sm_name),
                zone = COALESCE(excluded.zone, monthly_pjp_meta.zone),
                title = COALESCE(excluded.title, monthly_pjp_meta.title),
                note = COALESCE(excluded.note, monthly_pjp_meta.note),
                updated_at = excluded.updated_at
            """,
            (
                workspace_id,
                user_id,
                year_month,
                (meta.get("sm_name") or "").strip() or None,
                (meta.get("zone") or "").strip() or None,
                (meta.get("title") or "").strip() or None,
                (meta.get("note") or "").strip() or None,
                now,
            ),
        )
    for item in days:
        if not isinstance(item, dict):
            continue
        raw_date = (item.get("plan_date") or item.get("date") or "").strip()
        try:
            d = date.fromisoformat(raw_date)
        except ValueError:
            continue
        if d.strftime("%Y-%m") != year_month:
            continue
        place = (item.get("place_to_visit") or item.get("places_to_visit") or "").strip() or None
        day_type = _infer_day_type(
            d, place, (item.get("day_type") or "").strip().lower() or None
        )
        kms = item.get("travel_kms")
        try:
            kms_val = float(kms) if kms is not None and str(kms).strip() != "" else None
        except (TypeError, ValueError):
            kms_val = None
        conn.execute(
            """
            INSERT INTO monthly_pjp_days (
                workspace_id, user_id, plan_date, day_name,
                place_to_visit, from_place, to_place,
                business_activity, particulars, travel_kms, night_stay,
                day_type, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(workspace_id, user_id, plan_date) DO UPDATE SET
                day_name = excluded.day_name,
                place_to_visit = excluded.place_to_visit,
                from_place = excluded.from_place,
                to_place = excluded.to_place,
                business_activity = excluded.business_activity,
                particulars = excluded.particulars,
                travel_kms = excluded.travel_kms,
                night_stay = excluded.night_stay,
                day_type = excluded.day_type,
                updated_at = excluded.updated_at
            """,
            (
                workspace_id,
                user_id,
                d.isoformat(),
                _day_name(d),
                place,
                (item.get("from_place") or item.get("from") or "").strip() or None,
                (item.get("to_place") or item.get("to") or "").strip() or None,
                (
                    item.get("business_activity")
                    or item.get("purpose_of_visit")
                    or ""
                ).strip()
                or None,
                (item.get("particulars") or "").strip() or None,
                kms_val,
                (item.get("night_stay") or "").strip() or None,
                day_type,
                now,
            ),
        )


def _norm_header(value) -> str:
    text = str(value or "").replace("\xa0", " ").strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def _parse_pjp_excel(file_bytes: bytes, filename: str = "") -> dict:
    """Parse BD Excel travel-plan into {year_month, meta, days}."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for PJP Excel import") from exc

    suffix = Path(filename or "pjp.xlsx").suffix.lower() or ".xlsx"
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    try:
        wb = load_workbook(tmp_path, data_only=True)
    finally:
        try:
            Path(tmp_path).unlink(missing_ok=True)
        except Exception:
            pass

    # Prefer the first sheet that has a Date header
    chosen = None
    header_row = None
    colmap: dict[str, int] = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        for r in range(1, min(20, (ws.max_row or 1) + 1)):
            headers = {}
            for c in range(1, min(16, (ws.max_column or 1) + 1)):
                h = _norm_header(ws.cell(r, c).value)
                if not h:
                    continue
                headers[h] = c
            # map aliases
            date_c = next((c for h, c in headers.items() if h == "date" or h.startswith("date")), None)
            place_c = next(
                (
                    c
                    for h, c in headers.items()
                    if "place" in h or h in {"town", "market", "location"}
                ),
                None,
            )
            if date_c and place_c:
                chosen = ws
                header_row = r
                colmap = {
                    "date": date_c,
                    "place": place_c,
                    "from": next((c for h, c in headers.items() if h.startswith("from")), None),
                    "to": next((c for h, c in headers.items() if h == "to" or h.startswith("to ")), None),
                    "activity": next(
                        (
                            c
                            for h, c in headers.items()
                            if "business" in h or "purpose" in h or "activit" in h
                        ),
                        None,
                    ),
                    "particulars": next(
                        (c for h, c in headers.items() if "particular" in h or h == "remarks"),
                        None,
                    ),
                    "kms": next((c for h, c in headers.items() if "km" in h or "kms" in h), None),
                    "night": next((c for h, c in headers.items() if "night" in h), None),
                }
                break
        if chosen:
            break
    if not chosen or not header_row:
        raise ValueError(
            "Could not find PJP header row (need Date + Places to Visit columns)"
        )

    sm_name = None
    zone = None
    title = None
    note = None
    # Scan top rows for SM / Zone / title / disclaimer
    for r in range(1, header_row):
        for c in range(1, min(10, (chosen.max_column or 1) + 1)):
            raw = chosen.cell(r, c).value
            if raw is None:
                continue
            text = str(raw).replace("\xa0", " ").strip()
            low = text.lower()
            if low in {"sm name", "sm", "sales manager"}:
                sm_name = str(chosen.cell(r, c + 1).value or "").strip() or sm_name
            elif low == "zone":
                zone = str(chosen.cell(r, c + 1).value or "").strip() or zone
            elif "travel" in low and "plan" in low:
                title = text
            elif "subject to change" in low:
                note = text

    days: list[dict] = []
    year_months: dict[str, int] = {}

    def _as_date(value) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            text = value.strip()
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y"):
                try:
                    return datetime.strptime(text, fmt).date()
                except ValueError:
                    continue
        return None

    def _txt(value) -> str | None:
        if value is None:
            return None
        if isinstance(value, str):
            t = value.replace("\xa0", " ").strip()
            return t or None
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return str(value)
        return str(value).strip() or None

    for r in range(header_row + 1, (chosen.max_row or header_row) + 1):
        d = _as_date(chosen.cell(r, colmap["date"]).value)
        if not d:
            continue
        place = _txt(chosen.cell(r, colmap["place"]).value) if colmap.get("place") else None
        frm = _txt(chosen.cell(r, colmap["from"]).value) if colmap.get("from") else None
        to = _txt(chosen.cell(r, colmap["to"]).value) if colmap.get("to") else None
        activity = (
            _txt(chosen.cell(r, colmap["activity"]).value) if colmap.get("activity") else None
        )
        particulars = (
            _txt(chosen.cell(r, colmap["particulars"]).value)
            if colmap.get("particulars")
            else None
        )
        # Older sheets used Purpose column only — already mapped to activity
        night = _txt(chosen.cell(r, colmap["night"]).value) if colmap.get("night") else None
        kms_raw = chosen.cell(r, colmap["kms"]).value if colmap.get("kms") else None
        try:
            kms_val = float(kms_raw) if kms_raw is not None and str(kms_raw).strip() != "" else None
        except (TypeError, ValueError):
            kms_val = None
        day_type = _infer_day_type(d, place, None)
        days.append(
            {
                "plan_date": d.isoformat(),
                "place_to_visit": place,
                "from_place": frm,
                "to_place": to,
                "business_activity": activity,
                "particulars": particulars,
                "travel_kms": kms_val,
                "night_stay": night,
                "day_type": day_type,
            }
        )
        ym = d.strftime("%Y-%m")
        year_months[ym] = year_months.get(ym, 0) + 1

    if not days:
        raise ValueError("No dated rows found in the Excel file")

    # Primary month = most common year-month in the sheet
    year_month = max(year_months.items(), key=lambda kv: kv[1])[0]
    days = [d for d in days if d["plan_date"].startswith(year_month)]
    if not title:
        title = f"Travel Plan for the month {year_month}"
    return {
        "year_month": year_month,
        "sm_name": sm_name,
        "zone": zone,
        "title": title,
        "note": note,
        "days": days,
        "sheet": chosen.title,
    }


@pjp_bp.route("/import", methods=["POST"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def import_pjp_excel():
    """Upload monthly PJP Excel (.xlsx) and upsert the full month."""
    uid, err = _require_user_id()
    if err:
        return err
    uploaded = request.files.get("file") or request.files.get("excel")
    if not uploaded or not uploaded.filename:
        return jsonify({"success": False, "error": {"message": "Upload an Excel file (.xlsx)"}}), 400
    filename = uploaded.filename
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify(
            {"success": False, "error": {"message": "Only .xlsx / .xlsm PJP files are supported"}}
        ), 400
    file_bytes = uploaded.read()
    if not file_bytes:
        return jsonify({"success": False, "error": {"message": "Empty file"}}), 400
    try:
        parsed = _parse_pjp_excel(file_bytes, filename)
    except ValueError as exc:
        return jsonify({"success": False, "error": {"message": str(exc)}}), 400
    except Exception as exc:
        return jsonify({"success": False, "error": {"message": f"Unable to read Excel: {exc}"}}), 400

    # Optional override: force year_month from form (rare)
    force_ym = (request.form.get("year_month") or "").strip()
    year_month = force_ym if _parse_ym(force_ym) else parsed["year_month"]
    if force_ym and _parse_ym(force_ym):
        parsed["days"] = [
            d for d in parsed["days"] if str(d.get("plan_date") or "").startswith(year_month)
        ]

    workspace_id = get_workspace_id()
    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_tables(conn)
        _apply_month_bulk(
            conn,
            workspace_id,
            uid,
            year_month,
            {
                "sm_name": parsed.get("sm_name"),
                "zone": parsed.get("zone"),
                "title": parsed.get("title"),
                "note": parsed.get("note"),
            },
            parsed.get("days") or [],
        )
        conn.commit()
        payload = _month_payload(conn, workspace_id, uid, year_month)
    payload["import"] = {
        "filename": filename,
        "sheet": parsed.get("sheet"),
        "imported_days": len(parsed.get("days") or []),
        "planned_days": payload.get("stats", {}).get("planned_days"),
    }
    return jsonify({"success": True, "data": payload})


@pjp_bp.route("/today", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def today_plan():
    uid, err = _require_user_id()
    if err:
        return err
    workspace_id = get_workspace_id()
    today = date.today().isoformat()
    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_tables(conn)
        row = conn.execute(
            """
            SELECT * FROM monthly_pjp_days
            WHERE workspace_id = ? AND user_id = ? AND plan_date = ?
            """,
            (workspace_id, uid, today),
        ).fetchone()
    return jsonify(
        {
            "success": True,
            "data": {
                "plan_date": today,
                "day": _day_dict(row) if row else _empty_day(date.today()),
            },
        }
    )


@pjp_bp.route("/week", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def week_plan():
    """Current calendar week Monday→Sunday — for My Day weekly glance."""
    uid, err = _require_user_id()
    if err:
        return err
    workspace_id = get_workspace_id()
    today = date.today()
    # Monday = 0 … Sunday = 6
    start = today - timedelta(days=today.weekday())
    end = start + timedelta(days=6)
    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_tables(conn)
        rows = conn.execute(
            """
            SELECT * FROM monthly_pjp_days
            WHERE workspace_id = ? AND user_id = ?
              AND plan_date >= ? AND plan_date <= ?
            ORDER BY plan_date ASC
            """,
            (workspace_id, uid, start.isoformat(), end.isoformat()),
        ).fetchall()
    by_date = {r["plan_date"]: _day_dict(r) for r in rows}
    days = []
    planned = 0
    for i in range(7):
        d = start + timedelta(days=i)
        entry = by_date.get(d.isoformat()) or _empty_day(d)
        place = (entry.get("place_to_visit") or "").strip()
        if place and place.lower() not in {"holiday", "leave"}:
            planned += 1
        days.append(entry)
    return jsonify(
        {
            "success": True,
            "data": {
                "start_date": start.isoformat(),
                "end_date": end.isoformat(),
                "days": days,
                "planned_days": planned,
            },
        }
    )
