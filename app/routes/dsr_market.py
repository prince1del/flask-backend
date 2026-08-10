"""DSR Market Visit — form data + Excel export matching field DSR format."""

from __future__ import annotations

import io
import json
import sqlite3
from calendar import month_name
from datetime import datetime, timezone

from flask import Blueprint, current_app, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from app.routes.auth import get_workspace_id, require_jwt_auth, require_role

dsr_market_bp = Blueprint("dsr_market", __name__, url_prefix="/api/v1/dsr-market")

DEFAULT_COMPETITOR_BRANDS = [
    "Bombay Dyeing",
    "Ddecor",
    "Portico",
    "Raymonds",
    "Sansar",
    "Spaces",
    "Swayam",
    "Welspun",
]

DEFAULT_MAIN_CATEGORIES = [
    "Bedsheet",
    "Towel",
    "Comforter",
    "Blanket",
    "Others",
]

DEFAULT_LOW_STOCK_REASONS = [
    "Price",
    "Margin",
    "Designs",
    "Availability",
    "Distributor service",
    "Credit",
    "Slow movement",
    "Quality issue",
    "Lack of schemes",
    "Other",
]

DEFAULT_VISIT_ISSUES = [
    "Stock availability",
    "Delivery delay",
    "Damaged/defective product",
    "Packing issue",
    "Replacement/claim pending",
    "Billing issue",
    "Distributor issue",
    "Scheme/claim issue",
    "Price issue",
    "No sales support",
    "Other",
]

DEFAULT_PLACEMENT_CATEGORIES = [
    "Bedsheets",
    "Towels",
    "Comforters",
    "Blankets",
    "Pillows",
    "Other",
]


def _excel_headers(include_owner: bool) -> list[str]:
    headers = [
        "Sr. No.",
        "Date",
        "Day",
        "Name of Customer",
    ]
    if include_owner:
        headers.append("Owner's Name")
    headers.extend(
        [
            "ContactNos.",
            "MBO / ARS",
            "Type (A / B / C)",
            "Complete Address",
            "City and Area",
            "Existing OR New",
            "Order Recd. in Lacs",
            "BED",
            "BATH",
            "TOB",
            "OTHERS",
            "Other Competitor Brands available in store",
            "Branding in Store - Yes / No",
            "Feed Back from Retailer",
            "Remarks from SM",
        ]
    )
    return headers


def _db_path() -> str:
    return current_app.config.get("DATABASE_PATH", "centralized_db.sqlite3")


def _ensure_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS dsr_market_visits (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            workspace_id TEXT NOT NULL,
            user_id INTEGER,
            username TEXT,
            visit_date TEXT NOT NULL,
            customer_name TEXT NOT NULL,
            location TEXT,
            owner_name TEXT,
            contact_nos TEXT,
            channel_type TEXT,
            customer_type TEXT,
            address TEXT,
            city_area TEXT,
            existing_or_new TEXT,
            order_lacs REAL,
            bed TEXT,
            bath TEXT,
            tob TEXT,
            others TEXT,
            competitor_brands TEXT,
            branding_yn TEXT,
            retailer_feedback TEXT,
            sm_remarks TEXT,
            created_at TEXT NOT NULL
        )
        """
    )
    cols = {row[1] for row in conn.execute("PRAGMA table_info(dsr_market_visits)")}
    if "owner_name" not in cols:
        conn.execute("ALTER TABLE dsr_market_visits ADD COLUMN owner_name TEXT")
    if "location" not in cols:
        conn.execute("ALTER TABLE dsr_market_visits ADD COLUMN location TEXT")
    # Full retailer questionnaire (app intelligence). Never exported to HO Excel.
    if "visit_intel_json" not in cols:
        conn.execute("ALTER TABLE dsr_market_visits ADD COLUMN visit_intel_json TEXT")
    # Draft / day-close / party linkage
    if "is_draft" not in cols:
        conn.execute("ALTER TABLE dsr_market_visits ADD COLUMN is_draft INTEGER NOT NULL DEFAULT 0")
    if "draft_party_kind" not in cols:
        conn.execute("ALTER TABLE dsr_market_visits ADD COLUMN draft_party_kind TEXT")
    if "linked_distributor_id" not in cols:
        conn.execute("ALTER TABLE dsr_market_visits ADD COLUMN linked_distributor_id INTEGER")
    if "linked_retailer_id" not in cols:
        conn.execute("ALTER TABLE dsr_market_visits ADD COLUMN linked_retailer_id INTEGER")
    if "area_text" not in cols:
        conn.execute("ALTER TABLE dsr_market_visits ADD COLUMN area_text TEXT")
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_dsr_market_ws_date "
        "ON dsr_market_visits(workspace_id, visit_date)"
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS dsr_competitor_brands (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            workspace_id TEXT NOT NULL,
            brand_name TEXT NOT NULL,
            brand_key TEXT NOT NULL,
            created_by INTEGER,
            created_at TEXT NOT NULL,
            UNIQUE(workspace_id, brand_key)
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_dsr_competitor_ws "
        "ON dsr_competitor_brands(workspace_id)"
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS dsr_main_categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            workspace_id TEXT NOT NULL,
            category_name TEXT NOT NULL,
            category_key TEXT NOT NULL,
            created_by INTEGER,
            created_at TEXT NOT NULL,
            UNIQUE(workspace_id, category_key)
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_dsr_main_cat_ws "
        "ON dsr_main_categories(workspace_id)"
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS dsr_low_stock_reasons (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            workspace_id TEXT NOT NULL,
            reason_name TEXT NOT NULL,
            reason_key TEXT NOT NULL,
            created_by INTEGER,
            created_at TEXT NOT NULL,
            UNIQUE(workspace_id, reason_key)
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_dsr_low_stock_ws "
        "ON dsr_low_stock_reasons(workspace_id)"
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS dsr_visit_issues (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            workspace_id TEXT NOT NULL,
            issue_name TEXT NOT NULL,
            issue_key TEXT NOT NULL,
            created_by INTEGER,
            created_at TEXT NOT NULL,
            UNIQUE(workspace_id, issue_key)
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_dsr_visit_issues_ws "
        "ON dsr_visit_issues(workspace_id)"
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS dsr_placement_categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            workspace_id TEXT NOT NULL,
            category_name TEXT NOT NULL,
            category_key TEXT NOT NULL,
            created_by INTEGER,
            created_at TEXT NOT NULL,
            UNIQUE(workspace_id, category_key)
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_dsr_placement_cat_ws "
        "ON dsr_placement_categories(workspace_id)"
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS dsr_day_closures (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            workspace_id TEXT NOT NULL,
            user_id INTEGER NOT NULL,
            visit_date TEXT NOT NULL,
            closed_at TEXT NOT NULL,
            UNIQUE(workspace_id, user_id, visit_date)
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_dsr_day_close "
        "ON dsr_day_closures(workspace_id, user_id, visit_date)"
    )
    # Approach Distributor — prospect pipeline (Customers tab + New distributor visits)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS dsr_approach_distributors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            workspace_id TEXT NOT NULL,
            user_id INTEGER,
            username TEXT,
            firm_name TEXT NOT NULL,
            owner_name TEXT,
            contact_nos TEXT,
            city_area TEXT,
            location TEXT,
            address TEXT,
            monthly_ht TEXT,
            main_categories TEXT,
            channel_type TEXT,
            existing_or_new TEXT,
            customer_type TEXT,
            source_visit_id INTEGER,
            notes_json TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_dsr_approach_ws "
        "ON dsr_approach_distributors(workspace_id, firm_name COLLATE NOCASE)"
    )


def _brand_key(name: str) -> str:
    return " ".join((name or "").strip().lower().split())


def _merged_competitor_brands(conn: sqlite3.Connection, workspace_id: str) -> list[str]:
    rows = conn.execute(
        "SELECT brand_name FROM dsr_competitor_brands WHERE workspace_id = ? "
        "ORDER BY brand_name COLLATE NOCASE ASC",
        (workspace_id,),
    ).fetchall()
    custom = [str(r[0]).strip() for r in rows if r and r[0] and str(r[0]).strip()]
    by_key: dict[str, str] = {}
    for name in DEFAULT_COMPETITOR_BRANDS + custom:
        key = _brand_key(name)
        if key and key not in by_key:
            by_key[key] = name.strip()
    return sorted(by_key.values(), key=lambda s: s.lower())


def _merged_main_categories(conn: sqlite3.Connection, workspace_id: str) -> list[str]:
    rows = conn.execute(
        "SELECT category_name FROM dsr_main_categories WHERE workspace_id = ? "
        "ORDER BY category_name COLLATE NOCASE ASC",
        (workspace_id,),
    ).fetchall()
    custom = [str(r[0]).strip() for r in rows if r and r[0] and str(r[0]).strip()]
    by_key: dict[str, str] = {}
    # Keep defaults first, then extras alphabetically among customs only —
    # overall list still sorted for stable checkboxes.
    for name in DEFAULT_MAIN_CATEGORIES + custom:
        key = _brand_key(name)
        if key and key not in by_key:
            by_key[key] = name.strip()
    defaults_keys = {_brand_key(n) for n in DEFAULT_MAIN_CATEGORIES}
    default_ordered = [
        by_key[k] for k in (_brand_key(n) for n in DEFAULT_MAIN_CATEGORIES) if k in by_key
    ]
    extras = sorted(
        (v for k, v in by_key.items() if k not in defaults_keys),
        key=lambda s: s.lower(),
    )
    return default_ordered + extras


def _merged_low_stock_reasons(conn: sqlite3.Connection, workspace_id: str) -> list[str]:
    rows = conn.execute(
        "SELECT reason_name FROM dsr_low_stock_reasons WHERE workspace_id = ? "
        "ORDER BY reason_name COLLATE NOCASE ASC",
        (workspace_id,),
    ).fetchall()
    custom = [str(r[0]).strip() for r in rows if r and r[0] and str(r[0]).strip()]
    by_key: dict[str, str] = {}
    for name in DEFAULT_LOW_STOCK_REASONS + custom:
        key = _brand_key(name)
        if key and key not in by_key:
            by_key[key] = name.strip()
    defaults_keys = {_brand_key(n) for n in DEFAULT_LOW_STOCK_REASONS}
    default_ordered = [
        by_key[k]
        for k in (_brand_key(n) for n in DEFAULT_LOW_STOCK_REASONS)
        if k in by_key
    ]
    extras = sorted(
        (v for k, v in by_key.items() if k not in defaults_keys),
        key=lambda s: s.lower(),
    )
    # Keep "Other" last for the free-text box.
    other_key = _brand_key("Other")
    without_other = [n for n in default_ordered + extras if _brand_key(n) != other_key]
    if other_key in by_key:
        without_other.append(by_key[other_key])
    return without_other


def _merged_visit_issues(conn: sqlite3.Connection, workspace_id: str) -> list[str]:
    rows = conn.execute(
        "SELECT issue_name FROM dsr_visit_issues WHERE workspace_id = ? "
        "ORDER BY issue_name COLLATE NOCASE ASC",
        (workspace_id,),
    ).fetchall()
    custom = [str(r[0]).strip() for r in rows if r and r[0] and str(r[0]).strip()]
    by_key: dict[str, str] = {}
    for name in DEFAULT_VISIT_ISSUES + custom:
        key = _brand_key(name)
        if key and key not in by_key:
            by_key[key] = name.strip()
    defaults_keys = {_brand_key(n) for n in DEFAULT_VISIT_ISSUES}
    default_ordered = [
        by_key[k]
        for k in (_brand_key(n) for n in DEFAULT_VISIT_ISSUES)
        if k in by_key
    ]
    extras = sorted(
        (v for k, v in by_key.items() if k not in defaults_keys),
        key=lambda s: s.lower(),
    )
    other_key = _brand_key("Other")
    without_other = [n for n in default_ordered + extras if _brand_key(n) != other_key]
    if other_key in by_key:
        without_other.append(by_key[other_key])
    return without_other


def _merged_placement_categories(conn: sqlite3.Connection, workspace_id: str) -> list[str]:
    rows = conn.execute(
        "SELECT category_name FROM dsr_placement_categories WHERE workspace_id = ? "
        "ORDER BY category_name COLLATE NOCASE ASC",
        (workspace_id,),
    ).fetchall()
    custom = [str(r[0]).strip() for r in rows if r and r[0] and str(r[0]).strip()]
    by_key: dict[str, str] = {}
    for name in DEFAULT_PLACEMENT_CATEGORIES + custom:
        key = _brand_key(name)
        if key and key not in by_key:
            by_key[key] = name.strip()
    defaults_keys = {_brand_key(n) for n in DEFAULT_PLACEMENT_CATEGORIES}
    default_ordered = [
        by_key[k]
        for k in (_brand_key(n) for n in DEFAULT_PLACEMENT_CATEGORIES)
        if k in by_key
    ]
    extras = sorted(
        (v for k, v in by_key.items() if k not in defaults_keys),
        key=lambda s: s.lower(),
    )
    other_key = _brand_key("Other")
    without_other = [n for n in default_ordered + extras if _brand_key(n) != other_key]
    if other_key in by_key:
        without_other.append(by_key[other_key])
    return without_other


def _current_user() -> dict:
    return getattr(request, "user", None) or {}


def _user_id() -> int | None:
    raw = _current_user().get("user_id")
    try:
        return int(raw) if raw is not None else None
    except (TypeError, ValueError):
        return None


def _row_to_dict(row: sqlite3.Row) -> dict:
    d = dict(row)
    if "is_draft" in d:
        d["is_draft"] = bool(d.get("is_draft"))
    return _enrich_row_narratives(d)


def _looks_like_legacy_narrative(text: str | None) -> bool:
    """Pipe dumps / label stubs from older app builds — not engine prose."""
    t = (text or "").strip()
    if not t:
        return True
    if " | " in t or t.count("|") >= 1:
        return True
    low = t.lower()
    if low.startswith("response:") or "opp:" in low or low.startswith("expects:"):
        return True
    if "visit focus:" in low and "response:" in low:
        return True
    # Chip-ish fragment without a full sentence
    if "the " not in low and "further " not in low and len(t) < 160 and t.count(",") >= 2:
        return True
    return False


def _enrich_row_narratives(d: dict) -> dict:
    """Prefer deterministic engine text whenever visit_intel_json exists.

    Heals older rows that still store pipe-joined dumps so app + Excel match.
    """
    intel = d.get("visit_intel_json")
    if not intel:
        return d
    intel_s = intel if isinstance(intel, str) else json.dumps(intel, ensure_ascii=False)
    feedback, remarks = _resolve_visit_narratives(d, intel_s)
    # Force engine output when it produced text (do not keep legacy dumps).
    if feedback:
        d["retailer_feedback"] = feedback
    elif _looks_like_legacy_narrative(d.get("retailer_feedback")):
        d["retailer_feedback"] = None
    if remarks:
        d["sm_remarks"] = remarks
    elif _looks_like_legacy_narrative(d.get("sm_remarks")):
        d["sm_remarks"] = None
    return d


def _persist_enriched_narratives(conn: sqlite3.Connection, rows: list[dict]) -> int:
    """Write engine narratives back to DB when they differ from stored dumps."""
    updated = 0
    for d in rows:
        visit_id = d.get("id")
        intel = d.get("visit_intel_json")
        if visit_id is None or not intel:
            continue
        fb = (d.get("retailer_feedback") or "").strip() or None
        sm = (d.get("sm_remarks") or "").strip() or None
        cur = conn.execute(
            "SELECT retailer_feedback, sm_remarks FROM dsr_market_visits WHERE id = ?",
            (visit_id,),
        ).fetchone()
        if not cur:
            continue
        old_fb = (cur[0] or "").strip() or None
        old_sm = (cur[1] or "").strip() or None
        if fb == old_fb and sm == old_sm:
            continue
        conn.execute(
            "UPDATE dsr_market_visits SET retailer_feedback = ?, sm_remarks = ? WHERE id = ?",
            (fb, sm, visit_id),
        )
        updated += 1
    if updated:
        conn.commit()
    return updated


def _truthy_flag(raw: str | None) -> bool:
    return (raw or "").strip().lower() in {"1", "true", "yes", "y", "on"}


def _norm_name(s: str | None) -> str:
    return " ".join((s or "").strip().lower().split())


def _norm_phone(s: str | None) -> str:
    return "".join(ch for ch in (s or "") if ch.isdigit())


def _address_tokens(s: str | None) -> set[str]:
    raw = _norm_name(s)
    return {t for t in raw.replace(",", " ").split() if len(t) > 2}


def _address_overlap(a: str | None, b: str | None) -> float:
    ta, tb = _address_tokens(a), _address_tokens(b)
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / float(max(len(ta), len(tb)))


def _open_visit_dates(conn: sqlite3.Connection, workspace_id: str, user_id: int | None) -> list[str]:
    if user_id is None:
        return []
    closed = {
        r[0]
        for r in conn.execute(
            "SELECT visit_date FROM dsr_day_closures WHERE workspace_id = ? AND user_id = ?",
            (workspace_id, user_id),
        ).fetchall()
    }
    rows = conn.execute(
        "SELECT DISTINCT visit_date FROM dsr_market_visits "
        "WHERE workspace_id = ? AND user_id = ? ORDER BY visit_date ASC",
        (workspace_id, user_id),
    ).fetchall()
    return [r[0] for r in rows if r[0] and r[0] not in closed]


def _master_db():
    from centralized_db_system.db import CentralizedDB

    return CentralizedDB()


def _day_is_closed(conn: sqlite3.Connection, workspace_id: str, user_id: int | None, visit_date: str) -> bool:
    if user_id is None or not visit_date:
        return False
    row = conn.execute(
        "SELECT 1 FROM dsr_day_closures WHERE workspace_id = ? AND user_id = ? AND visit_date = ?",
        (workspace_id, user_id, visit_date),
    ).fetchone()
    return row is not None


def _resolve_visit_narratives(data: dict, visit_intel_json: str | None) -> tuple[str | None, str | None]:
    """Generate HO narratives from structured intel when present; else keep client text."""
    from app.services.dsr_visit_narrative import generate_visit_narratives

    client_fb = (data.get("retailer_feedback") or "").strip() or None
    client_sm = (data.get("sm_remarks") or "").strip() or None

    if not visit_intel_json:
        return client_fb, client_sm

    narratives = generate_visit_narratives(
        visit_intel_json=visit_intel_json,
        retailer_name=(data.get("customer_name") or "").strip(),
        retailer_id=data.get("linked_retailer_id") or data.get("retailer_id"),
        visit_date=(data.get("visit_date") or "").strip(),
        customer_type=(data.get("customer_type") or "").strip() or None,
    )
    feedback = (narratives.get("retailer_feedback") or "").strip() or None
    remarks = (narratives.get("sm_remarks") or "").strip() or None
    # Prefer SM remarks typed by the executive (desktop + mobile).
    if client_sm and not _looks_like_legacy_narrative(client_sm):
        remarks = client_sm
    # Fall back to client feedback only when it is real prose — never keep pipe/chip dumps.
    if not feedback and client_fb and not _looks_like_legacy_narrative(client_fb):
        feedback = client_fb
    return feedback, remarks


@dsr_market_bp.route("/visits", methods=["POST"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def create_visit():
    workspace_id = get_workspace_id()
    user = _current_user()
    data = request.get_json(silent=True) or {}

    customer_name = (data.get("customer_name") or "").strip()
    visit_date = (data.get("visit_date") or "").strip()
    if not customer_name or not visit_date:
        return jsonify(
            {"success": False, "error": {"message": "customer_name and visit_date are required"}}
        ), 400

    is_draft = 1 if _truthy_flag(str(data.get("is_draft") if data.get("is_draft") is not None else "0")) else 0
    # Also treat explicit draft_party_kind as draft
    draft_party_kind = (data.get("draft_party_kind") or "").strip().lower() or None
    if draft_party_kind in {"retailer", "distributor"}:
        is_draft = 1
    elif draft_party_kind:
        draft_party_kind = None

    linked_distributor_id = data.get("linked_distributor_id")
    linked_retailer_id = data.get("linked_retailer_id")
    try:
        linked_distributor_id = int(linked_distributor_id) if linked_distributor_id not in (None, "") else None
    except (TypeError, ValueError):
        linked_distributor_id = None
    try:
        linked_retailer_id = int(linked_retailer_id) if linked_retailer_id not in (None, "") else None
    except (TypeError, ValueError):
        linked_retailer_id = None

    # Existing party visit is not a draft
    if linked_retailer_id or (linked_distributor_id and draft_party_kind is None and not is_draft):
        if linked_retailer_id:
            is_draft = 0
            draft_party_kind = None

    area_text = (data.get("area_text") or data.get("city_area") or "").strip() or None

    created_at = datetime.now(timezone.utc).isoformat()
    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_table(conn)
        uid = _user_id()
        open_dates = _open_visit_dates(conn, workspace_id, uid)
        if visit_date not in open_dates and len(open_dates) >= 2:
            return jsonify(
                {
                    "success": False,
                    "error": {
                        "message": (
                            "Maximum 2 open visit days allowed. "
                            f"Please Day Close one of: {', '.join(open_dates)} before starting {visit_date}."
                        ),
                        "code": "max_open_days",
                        "open_dates": open_dates,
                    },
                }
            ), 400

        # Re-open day if previously closed and user adds another visit
        if uid is not None:
            conn.execute(
                "DELETE FROM dsr_day_closures WHERE workspace_id = ? AND user_id = ? AND visit_date = ?",
                (workspace_id, uid, visit_date),
            )

        intel_raw = data.get("visit_intel_json")
        if isinstance(intel_raw, (dict, list)):
            visit_intel_json = json.dumps(intel_raw, ensure_ascii=False)
        else:
            visit_intel_json = (str(intel_raw).strip() if intel_raw is not None else "") or None

        # Generate narratives on Save from structured intel; store once for app + Excel.
        retailer_feedback, sm_remarks = _resolve_visit_narratives(data, visit_intel_json)

        cur = conn.execute(
            """
            INSERT INTO dsr_market_visits (
                workspace_id, user_id, username, visit_date, customer_name, location, owner_name, contact_nos,
                channel_type, customer_type, address, city_area, existing_or_new,
                order_lacs, bed, bath, tob, others, competitor_brands, branding_yn,
                retailer_feedback, sm_remarks, visit_intel_json,
                is_draft, draft_party_kind, linked_distributor_id, linked_retailer_id, area_text,
                created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                workspace_id,
                uid,
                user.get("username"),
                visit_date,
                customer_name,
                (data.get("location") or "").strip() or None,
                (data.get("owner_name") or "").strip() or None,
                (data.get("contact_nos") or "").strip() or None,
                (data.get("channel_type") or "").strip() or None,
                (data.get("customer_type") or "").strip() or None,
                (data.get("address") or "").strip() or None,
                (data.get("city_area") or area_text or "").strip() or None,
                (data.get("existing_or_new") or "").strip() or None,
                data.get("order_lacs"),
                (data.get("bed") or "").strip() or None,
                (data.get("bath") or "").strip() or None,
                (data.get("tob") or "").strip() or None,
                (data.get("others") or "").strip() or None,
                (data.get("competitor_brands") or "").strip() or None,
                (data.get("branding_yn") or "").strip() or None,
                retailer_feedback,
                sm_remarks,
                visit_intel_json,
                is_draft,
                draft_party_kind,
                linked_distributor_id,
                linked_retailer_id,
                area_text,
                created_at,
            ),
        )
        visit_id = int(cur.lastrowid)
        if draft_party_kind == "distributor":
            _upsert_approach_from_visit(
                conn,
                workspace_id=workspace_id,
                uid=uid,
                username=user.get("username"),
                data=data,
                visit_id=visit_id,
                visit_intel_json=visit_intel_json,
            )
        conn.commit()
        row = conn.execute("SELECT * FROM dsr_market_visits WHERE id = ?", (visit_id,)).fetchone()

    return jsonify({"success": True, "data": _row_to_dict(row)}), 201


@dsr_market_bp.route("/visits/<int:visit_id>", methods=["PATCH", "PUT"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def update_visit(visit_id: int):
    """Update visit questionnaire; regenerate narratives unless day is closed/locked."""
    workspace_id = get_workspace_id()
    data = request.get_json(silent=True) or {}
    force = _truthy_flag(str(data.get("force") if data.get("force") is not None else "0"))

    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_table(conn)
        uid = _user_id()
        row = conn.execute(
            "SELECT * FROM dsr_market_visits WHERE id = ? AND workspace_id = ?",
            (visit_id, workspace_id),
        ).fetchone()
        if not row:
            return jsonify({"success": False, "error": {"message": "Visit not found"}}), 404
        visit = dict(row)
        if uid is not None and visit.get("user_id") not in (None, uid):
            role = (_current_user().get("role") or "").strip().lower()
            if role not in {"admin", "hop_admin"}:
                return jsonify({"success": False, "error": {"message": "Not allowed"}}), 403

        visit_date = (data.get("visit_date") or visit.get("visit_date") or "").strip()
        owner_uid = visit.get("user_id") if visit.get("user_id") is not None else uid
        if _day_is_closed(conn, workspace_id, owner_uid, visit_date) and not force:
            return jsonify(
                {
                    "success": False,
                    "error": {
                        "message": (
                            f"Visit day {visit_date} is closed. "
                            "Re-open / unlock the day before editing, or pass force=1."
                        ),
                        "code": "day_closed",
                    },
                }
            ), 409

        customer_name = (data.get("customer_name") or visit.get("customer_name") or "").strip()
        if not customer_name:
            return jsonify(
                {"success": False, "error": {"message": "customer_name is required"}}
            ), 400

        intel_raw = data.get("visit_intel_json", visit.get("visit_intel_json"))
        if isinstance(intel_raw, (dict, list)):
            visit_intel_json = json.dumps(intel_raw, ensure_ascii=False)
        elif intel_raw is None:
            visit_intel_json = None
        else:
            visit_intel_json = (str(intel_raw).strip() or None)

        merged = {
            **visit,
            **{k: v for k, v in data.items() if v is not None},
            "customer_name": customer_name,
            "visit_date": visit_date,
            "visit_intel_json": visit_intel_json,
        }
        # Always regenerate narratives on update from current structured intel.
        retailer_feedback, sm_remarks = _resolve_visit_narratives(merged, visit_intel_json)

        area_text = (data.get("area_text") if "area_text" in data else visit.get("area_text"))
        if area_text is not None:
            area_text = (str(area_text).strip() or None)

        def _pick(key: str, strip: bool = True):
            if key in data:
                val = data.get(key)
                if val is None:
                    return None
                return (str(val).strip() or None) if strip else val
            return visit.get(key)

        conn.execute(
            """
            UPDATE dsr_market_visits SET
                visit_date = ?,
                customer_name = ?,
                location = ?,
                owner_name = ?,
                contact_nos = ?,
                channel_type = ?,
                customer_type = ?,
                address = ?,
                city_area = ?,
                existing_or_new = ?,
                order_lacs = ?,
                bed = ?,
                bath = ?,
                tob = ?,
                others = ?,
                competitor_brands = ?,
                branding_yn = ?,
                retailer_feedback = ?,
                sm_remarks = ?,
                visit_intel_json = ?,
                linked_distributor_id = ?,
                linked_retailer_id = ?,
                area_text = ?
            WHERE id = ? AND workspace_id = ?
            """,
            (
                visit_date,
                customer_name,
                _pick("location"),
                _pick("owner_name"),
                _pick("contact_nos"),
                _pick("channel_type"),
                _pick("customer_type"),
                _pick("address"),
                _pick("city_area") or area_text,
                _pick("existing_or_new"),
                data.get("order_lacs") if "order_lacs" in data else visit.get("order_lacs"),
                _pick("bed"),
                _pick("bath"),
                _pick("tob"),
                _pick("others"),
                _pick("competitor_brands"),
                _pick("branding_yn"),
                retailer_feedback,
                sm_remarks,
                visit_intel_json,
                data.get("linked_distributor_id")
                if "linked_distributor_id" in data
                else visit.get("linked_distributor_id"),
                data.get("linked_retailer_id")
                if "linked_retailer_id" in data
                else visit.get("linked_retailer_id"),
                area_text,
                visit_id,
                workspace_id,
            ),
        )
        conn.commit()
        updated = conn.execute(
            "SELECT * FROM dsr_market_visits WHERE id = ?", (visit_id,)
        ).fetchone()

    return jsonify({"success": True, "data": _row_to_dict(updated)})


@dsr_market_bp.route("/visits/<int:visit_id>", methods=["DELETE"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def delete_visit(visit_id: int):
    """Permanently delete a market visit owned by the caller (admins may delete any)."""
    workspace_id = get_workspace_id()

    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_table(conn)
        uid = _user_id()
        row = conn.execute(
            "SELECT * FROM dsr_market_visits WHERE id = ? AND workspace_id = ?",
            (visit_id, workspace_id),
        ).fetchone()
        if not row:
            return jsonify({"success": False, "error": {"message": "Visit not found"}}), 404
        visit = dict(row)
        if uid is not None and visit.get("user_id") not in (None, uid):
            role = (_current_user().get("role") or "").strip().lower()
            if role not in {"admin", "hop_admin"}:
                return jsonify({"success": False, "error": {"message": "Not allowed"}}), 403

        conn.execute(
            "DELETE FROM dsr_market_visits WHERE id = ? AND workspace_id = ?",
            (visit_id, workspace_id),
        )
        conn.commit()

    return jsonify({"success": True, "data": {"id": visit_id, "deleted": True}})


@dsr_market_bp.route("/visits", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def list_visits():
    workspace_id = get_workspace_id()
    from_date = (request.args.get("from") or request.args.get("from_date") or "").strip()
    to_date = (request.args.get("to") or request.args.get("to_date") or "").strip()
    visit_date = (request.args.get("date") or "").strip()

    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_table(conn)
        query = "SELECT * FROM dsr_market_visits WHERE workspace_id = ?"
        params: list = [workspace_id]
        if visit_date:
            query += " AND visit_date = ?"
            params.append(visit_date)
        else:
            if from_date:
                query += " AND visit_date >= ?"
                params.append(from_date)
            if to_date:
                query += " AND visit_date <= ?"
                params.append(to_date)
        uid = _user_id()
        # BD sees own visits; admin can pass all=1
        if request.args.get("all") != "1" and uid is not None:
            query += " AND (user_id = ? OR user_id IS NULL)"
            params.append(uid)
        query += " ORDER BY visit_date DESC, id DESC LIMIT ?"
        params.append(request.args.get("limit", 500, type=int) or 500)
        rows = conn.execute(query, tuple(params)).fetchall()
        data = [_row_to_dict(r) for r in rows]
        # Heal DB: replace stored pipe dumps with engine narratives.
        _persist_enriched_narratives(conn, data)

    return jsonify({"success": True, "data": data, "count": len(data)})


@dsr_market_bp.route("/report-tree", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def report_tree():
    """Year → month folder index for in-app report browsing (no Excel download)."""
    workspace_id = get_workspace_id()
    uid = _user_id()
    with sqlite3.connect(_db_path()) as conn:
        _ensure_table(conn)
        query = (
            "SELECT substr(visit_date, 1, 4) AS y, substr(visit_date, 6, 2) AS m, "
            "COUNT(*) AS visit_count, COUNT(DISTINCT visit_date) AS day_count "
            "FROM dsr_market_visits WHERE workspace_id = ? "
            "AND visit_date IS NOT NULL AND length(visit_date) >= 7"
        )
        params: list = [workspace_id]
        if request.args.get("all") != "1" and uid is not None:
            query += " AND (user_id = ? OR user_id IS NULL)"
            params.append(uid)
        query += " GROUP BY y, m ORDER BY y DESC, m DESC"
        rows = conn.execute(query, tuple(params)).fetchall()

    month_names = {
        "01": "January",
        "02": "February",
        "03": "March",
        "04": "April",
        "05": "May",
        "06": "June",
        "07": "July",
        "08": "August",
        "09": "September",
        "10": "October",
        "11": "November",
        "12": "December",
    }
    by_year: dict[str, dict] = {}
    for y, m, visit_count, day_count in rows:
        year = str(y or "").strip()
        month = str(m or "").strip().zfill(2)
        if not year.isdigit() or not month.isdigit():
            continue
        bucket = by_year.setdefault(
            year,
            {"year": int(year), "visit_count": 0, "months": []},
        )
        bucket["visit_count"] += int(visit_count or 0)
        bucket["months"].append(
            {
                "year": int(year),
                "month": int(month),
                "month_key": f"{year}-{month}",
                "label": f"{month_names.get(month, month)} {year}",
                "from_date": f"{year}-{month}-01",
                "to_date": _month_end(int(year), int(month)),
                "visit_count": int(visit_count or 0),
                "day_count": int(day_count or 0),
            }
        )

    years = sorted(by_year.values(), key=lambda item: item["year"], reverse=True)
    return jsonify({"success": True, "data": {"years": years}, "count": len(years)})


def _month_end(year: int, month: int) -> str:
    if month == 12:
        return f"{year}-12-31"
    nxt = datetime(year, month + 1, 1)
    last = nxt.toordinal() - 1
    d = datetime.fromordinal(last)
    return d.strftime("%Y-%m-%d")


def _excel_cell_text(value) -> str | float | int:
    """Normalize values so Excel cells stay readable (no runaway strings)."""
    if value is None:
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    # Keep newlines for wrap; collapse huge pipe dumps into line breaks.
    text = text.replace(" | ", "\n").replace(" |", "\n").replace("| ", "\n")
    while "\n\n\n" in text:
        text = text.replace("\n\n\n", "\n\n")
    # Soft cap — full detail stays in visit_intel_json on server.
    # Narratives can be multi-sentence HO prose; allow more room than pipe dumps.
    if len(text) > 2000:
        text = text[:1997].rstrip() + "…"
    return text


def _build_excel(rows: list[dict], sm_name: str, period_label: str, include_owner: bool) -> bytes:
    """Build DSR Excel. Note: `location` is app-only and must never appear in export."""
    from openpyxl.styles import Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "DSR"
    headers = _excel_headers(include_owner)
    last_col = len(headers)

    thin = Border(
        left=Side(style="thin", color="FFCCCCCC"),
        right=Side(style="thin", color="FFCCCCCC"),
        top=Side(style="thin", color="FFCCCCCC"),
        bottom=Side(style="thin", color="FFCCCCCC"),
    )
    header_fill = PatternFill("solid", fgColor="FF1B5E20")
    header_font = Font(bold=True, color="FFFFFFFF", size=10)
    title_font = Font(bold=True, size=14, color="FF1B5E20")
    label_font = Font(bold=True, size=10)
    wrap_top = Alignment(wrap_text=True, vertical="top", horizontal="left")
    wrap_center = Alignment(wrap_text=True, vertical="center", horizontal="center")

    # Title block (do not park labels in random far columns — breaks mobile Excel view).
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(8, last_col))
    ws["A1"] = f"DSR Report — {period_label}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(vertical="center", horizontal="left")

    ws["A2"] = "Name of the SM :"
    ws["A2"].font = label_font
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=min(4, last_col))
    ws["B2"] = sm_name or ""
    ws["B2"].alignment = Alignment(vertical="center")

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8
    ws.row_dimensions[5].height = 36

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = thin

    # Prefer non-draft visits for HO; if only drafts exist, still export them.
    export_rows = [r for r in rows if not r.get("is_draft")]
    if not export_rows:
        export_rows = rows

    last_date = None
    sr = 0
    excel_row = 6
    for item in export_rows:
        sr += 1
        visit_date = item.get("visit_date") or ""
        day_name = ""
        date_display = ""
        try:
            dt = datetime.strptime(str(visit_date)[:10], "%Y-%m-%d")
            day_name = dt.strftime("%A")
            date_display = dt.strftime("%d-%b-%Y")
        except ValueError:
            date_display = visit_date

        show_date = visit_date != last_date
        last_date = visit_date

        values: list = [
            sr,
            date_display if show_date else "",
            day_name if show_date else "",
            item.get("customer_name") or "",
        ]
        if include_owner:
            values.append(item.get("owner_name") or "")
        values.extend(
            [
                item.get("contact_nos") or "",
                item.get("channel_type") or "",
                item.get("customer_type") or "",
                item.get("address") or "",
                item.get("city_area") or "",
                item.get("existing_or_new") or "",
                item.get("order_lacs") if item.get("order_lacs") is not None else "",
                item.get("bed") or "",
                item.get("bath") or "",
                item.get("tob") or "",
                item.get("others") or "",
                item.get("competitor_brands") or "",
                item.get("branding_yn") or "",
                item.get("retailer_feedback") or "",
                item.get("sm_remarks") or "",
            ]
        )

        # Guard: never write more cells than headers (avoids "cell se bahar" layout).
        values = values[:last_col]
        while len(values) < last_col:
            values.append("")

        max_lines = 1
        for col, val in enumerate(values, start=1):
            cleaned = _excel_cell_text(val)
            cell = ws.cell(row=excel_row, column=col, value=cleaned)
            cell.border = thin
            # Sr / date / day / short flags centered; long text wraps.
            if col in {1, 2, 3} or (
                isinstance(cleaned, str) and cleaned.upper() in {"Y", "N", "YES", "NO"}
            ):
                cell.alignment = wrap_center
            else:
                cell.alignment = wrap_top
            if isinstance(cleaned, str) and cleaned:
                max_lines = max(max_lines, cleaned.count("\n") + 1)

        # Grow row for wrapped feedback / remarks.
        ws.row_dimensions[excel_row].height = min(120, max(18, 14 * max_lines))
        excel_row += 1

    widths = [6, 12, 11, 26]
    if include_owner:
        widths.append(18)
    widths.extend([13, 10, 11, 26, 16, 11, 10, 7, 7, 7, 9, 22, 10, 28, 28])
    for i, w in enumerate(widths[:last_col], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:{get_column_letter(last_col)}{max(5, excel_row - 1)}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@dsr_market_bp.route("/regenerate-narratives", methods=["POST"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def regenerate_narratives():
    """Backfill retailer_feedback / sm_remarks from visit_intel_json (deterministic)."""
    workspace_id = get_workspace_id()
    data = request.get_json(silent=True) or {}
    from_date = (data.get("from") or data.get("from_date") or request.args.get("from") or "").strip()
    to_date = (data.get("to") or data.get("to_date") or request.args.get("to") or "").strip()

    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_table(conn)
        query = "SELECT * FROM dsr_market_visits WHERE workspace_id = ?"
        params: list = [workspace_id]
        if from_date:
            query += " AND visit_date >= ?"
            params.append(from_date)
        if to_date:
            query += " AND visit_date <= ?"
            params.append(to_date)
        uid = _user_id()
        if request.args.get("all") != "1" and uid is not None:
            query += " AND (user_id = ? OR user_id IS NULL)"
            params.append(uid)
        raw_rows = conn.execute(query, tuple(params)).fetchall()
        enriched = [_enrich_row_narratives(dict(r)) for r in raw_rows]
        updated = _persist_enriched_narratives(conn, enriched)

    return jsonify(
        {
            "success": True,
            "data": {"scanned": len(enriched), "updated": updated},
            "message": f"Regenerated narratives for {updated} of {len(enriched)} visits",
        }
    )


@dsr_market_bp.route("/export", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def export_excel():
    workspace_id = get_workspace_id()
    from_date = (request.args.get("from") or request.args.get("from_date") or "").strip()
    to_date = (request.args.get("to") or request.args.get("to_date") or "").strip()
    if not from_date or not to_date:
        return jsonify(
            {"success": False, "error": {"message": "from and to dates are required (YYYY-MM-DD)"}}
        ), 400

    include_owner = _truthy_flag(
        request.args.get("include_owner") or request.args.get("includeOwner")
    )

    user = _current_user()
    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_table(conn)
        query = (
            "SELECT * FROM dsr_market_visits WHERE workspace_id = ? "
            "AND visit_date >= ? AND visit_date <= ?"
        )
        params: list = [workspace_id, from_date, to_date]
        uid = _user_id()
        if request.args.get("all") != "1" and uid is not None:
            query += " AND (user_id = ? OR user_id IS NULL)"
            params.append(uid)
        query += " ORDER BY visit_date ASC, id ASC"
        rows = [_enrich_row_narratives(dict(r)) for r in conn.execute(query, tuple(params)).fetchall()]
        _persist_enriched_narratives(conn, rows)

    try:
        start = datetime.strptime(from_date, "%Y-%m-%d")
        period = f"{month_name[start.month]} {start.year}"
    except ValueError:
        period = f"{from_date} to {to_date}"

    # Prefer profile full_name (Settings → My profile), not login id like bd_gt_north_head.
    sm_name = ""
    uid = _user_id()
    if uid is not None:
        try:
            from centralized_db_system.db import CentralizedDB

            profile = CentralizedDB(_db_path()).get_user_profile(uid) or {}
            sm_name = (profile.get("full_name") or "").strip()
            if not sm_name:
                email = (profile.get("email") or "").strip()
                if "@" in email:
                    local = email.split("@", 1)[0]
                    sm_name = " ".join(
                        p.capitalize() for p in local.replace(".", " ").replace("_", " ").split() if p
                    )
                else:
                    sm_name = (profile.get("username") or "").strip()
        except Exception:
            sm_name = ""
    if not sm_name:
        sm_name = (user.get("username") or "").strip()

    polish = _truthy_flag(request.args.get("polish") or request.args.get("ai"))
    polish_meta: dict = {}
    if polish:
        from app.services.dsr_notes_polish import polish_rows

        rows, polish_meta = polish_rows(rows)
        if polish_meta.get("errors") == ["missing_api_key"]:
            return jsonify(
                {
                    "success": False,
                    "error": {
                        "message": (
                            "GEMINI_API_KEY not set. Add free key from "
                            "https://aistudio.google.com/apikey on Render Environment, then redeploy."
                        )
                    },
                    "meta": polish_meta,
                }
            ), 503

    content = _build_excel(
        rows, sm_name=sm_name, period_label=period, include_owner=include_owner
    )
    filename = f"DSR_{from_date}_to_{to_date}.xlsx"
    if polish:
        filename = f"DSR_{from_date}_to_{to_date}_polished.xlsx"
    resp = send_file(
        io.BytesIO(content),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
    if polish and polish_meta:
        resp.headers["X-DSR-Polish"] = json.dumps(
            {
                "polished_count": polish_meta.get("polished_count"),
                "row_count": polish_meta.get("row_count"),
                "errors": polish_meta.get("errors") or [],
            }
        )
    return resp


@dsr_market_bp.route("/polish-notes", methods=["POST"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def polish_notes():
    """Test/preview: polish one visit's feedback + SM remarks via Gemini free API.

    Body (either):
      { "visit_id": 123 }
      or full visit fields + optional visit_intel_json / retailer_feedback / sm_remarks
    """
    from app.services.dsr_notes_polish import gemini_configured, polish_visit_notes

    if not gemini_configured():
        return jsonify(
            {
                "success": False,
                "error": {
                    "message": (
                        "GEMINI_API_KEY not set. Get a free key at "
                        "https://aistudio.google.com/apikey and set it on Render."
                    )
                },
                "data": {"gemini_configured": False},
            }
        ), 503

    payload = request.get_json(silent=True) or {}
    row: dict = {}
    visit_id = payload.get("visit_id")
    if visit_id is not None:
        workspace_id = get_workspace_id()
        with sqlite3.connect(_db_path()) as conn:
            conn.row_factory = sqlite3.Row
            _ensure_table(conn)
            found = conn.execute(
                "SELECT * FROM dsr_market_visits WHERE id = ? AND workspace_id = ?",
                (int(visit_id), workspace_id),
            ).fetchone()
        if not found:
            return jsonify({"success": False, "error": {"message": "Visit not found"}}), 404
        row = dict(found)
    else:
        row = {
            "customer_name": payload.get("customer_name"),
            "retailer_feedback": payload.get("retailer_feedback"),
            "sm_remarks": payload.get("sm_remarks"),
            "customer_type": payload.get("customer_type"),
            "visit_intel_json": payload.get("visit_intel_json")
            or payload.get("intel")
            or {},
        }

    notes, err = polish_visit_notes(row)
    if err and err not in ("empty",):
        soft = err.startswith(("missing_api_key", "invalid_api_key", "quota_exceeded"))
        status = 503 if soft else 502
        return jsonify(
            {
                "success": False,
                "error": {"message": f"Gemini polish failed: {err}"},
                "data": {
                    "gemini_configured": True,
                    "original": {
                        "retailer_feedback": row.get("retailer_feedback"),
                        "sm_remarks": row.get("sm_remarks"),
                    },
                    "polished": notes,
                },
            }
        ), status

    return jsonify(
        {
            "success": True,
            "data": {
                "gemini_configured": True,
                "original": {
                    "retailer_feedback": row.get("retailer_feedback"),
                    "sm_remarks": row.get("sm_remarks"),
                },
                "polished": notes,
                "error": err,
            },
        }
    )


@dsr_market_bp.route("/polish-status", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def polish_status():
    from app.services.dsr_notes_polish import gemini_configured

    return jsonify(
        {
            "success": True,
            "data": {
                "gemini_configured": gemini_configured(),
                "hint": (
                    None
                    if gemini_configured()
                    else "Set GEMINI_API_KEY from https://aistudio.google.com/apikey"
                ),
            },
        }
    )


@dsr_market_bp.route("/competitor-brands", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def list_competitor_brands():
    workspace_id = get_workspace_id()
    with sqlite3.connect(_db_path()) as conn:
        _ensure_table(conn)
        brands = _merged_competitor_brands(conn, workspace_id)
    return jsonify(
        {
            "success": True,
            "data": brands,
            "defaults": list(DEFAULT_COMPETITOR_BRANDS),
            "count": len(brands),
        }
    )


@dsr_market_bp.route("/competitor-brands", methods=["POST"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def add_competitor_brand():
    workspace_id = get_workspace_id()
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or data.get("brand_name") or "").strip()
    if not name:
        return jsonify({"success": False, "error": {"message": "name is required"}}), 400
    key = _brand_key(name)
    if not key:
        return jsonify({"success": False, "error": {"message": "name is required"}}), 400

    created_at = datetime.now(timezone.utc).isoformat()
    with sqlite3.connect(_db_path()) as conn:
        _ensure_table(conn)
        existing = conn.execute(
            "SELECT brand_name FROM dsr_competitor_brands "
            "WHERE workspace_id = ? AND brand_key = ?",
            (workspace_id, key),
        ).fetchone()
        if existing is None and key not in {_brand_key(b) for b in DEFAULT_COMPETITOR_BRANDS}:
            conn.execute(
                """
                INSERT INTO dsr_competitor_brands (
                    workspace_id, brand_name, brand_key, created_by, created_at
                ) VALUES (?, ?, ?, ?, ?)
                """,
                (workspace_id, name, key, _user_id(), created_at),
            )
            conn.commit()
        brands = _merged_competitor_brands(conn, workspace_id)

    return jsonify({"success": True, "data": brands, "count": len(brands)}), 201


@dsr_market_bp.route("/main-categories", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def list_main_categories():
    workspace_id = get_workspace_id()
    with sqlite3.connect(_db_path()) as conn:
        _ensure_table(conn)
        categories = _merged_main_categories(conn, workspace_id)
    return jsonify(
        {
            "success": True,
            "data": categories,
            "defaults": list(DEFAULT_MAIN_CATEGORIES),
            "count": len(categories),
        }
    )


@dsr_market_bp.route("/main-categories", methods=["POST"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def add_main_category():
    workspace_id = get_workspace_id()
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or data.get("category_name") or "").strip()
    if not name:
        return jsonify({"success": False, "error": {"message": "name is required"}}), 400
    key = _brand_key(name)
    if not key:
        return jsonify({"success": False, "error": {"message": "name is required"}}), 400

    created_at = datetime.now(timezone.utc).isoformat()
    with sqlite3.connect(_db_path()) as conn:
        _ensure_table(conn)
        existing = conn.execute(
            "SELECT category_name FROM dsr_main_categories "
            "WHERE workspace_id = ? AND category_key = ?",
            (workspace_id, key),
        ).fetchone()
        if existing is None and key not in {_brand_key(b) for b in DEFAULT_MAIN_CATEGORIES}:
            conn.execute(
                """
                INSERT INTO dsr_main_categories (
                    workspace_id, category_name, category_key, created_by, created_at
                ) VALUES (?, ?, ?, ?, ?)
                """,
                (workspace_id, name, key, _user_id(), created_at),
            )
            conn.commit()
        categories = _merged_main_categories(conn, workspace_id)

    return jsonify({"success": True, "data": categories, "count": len(categories)}), 201


@dsr_market_bp.route("/low-stock-reasons", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def list_low_stock_reasons():
    workspace_id = get_workspace_id()
    with sqlite3.connect(_db_path()) as conn:
        _ensure_table(conn)
        reasons = _merged_low_stock_reasons(conn, workspace_id)
    return jsonify(
        {
            "success": True,
            "data": reasons,
            "defaults": list(DEFAULT_LOW_STOCK_REASONS),
            "count": len(reasons),
        }
    )


@dsr_market_bp.route("/low-stock-reasons", methods=["POST"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def add_low_stock_reason():
    workspace_id = get_workspace_id()
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or data.get("reason_name") or "").strip()
    if not name:
        return jsonify({"success": False, "error": {"message": "name is required"}}), 400
    key = _brand_key(name)
    if not key:
        return jsonify({"success": False, "error": {"message": "name is required"}}), 400
    if key == _brand_key("Other"):
        return jsonify(
            {"success": False, "error": {"message": "Use Other free-text instead"}}
        ), 400

    created_at = datetime.now(timezone.utc).isoformat()
    with sqlite3.connect(_db_path()) as conn:
        _ensure_table(conn)
        existing = conn.execute(
            "SELECT reason_name FROM dsr_low_stock_reasons "
            "WHERE workspace_id = ? AND reason_key = ?",
            (workspace_id, key),
        ).fetchone()
        if existing is None and key not in {_brand_key(b) for b in DEFAULT_LOW_STOCK_REASONS}:
            conn.execute(
                """
                INSERT INTO dsr_low_stock_reasons (
                    workspace_id, reason_name, reason_key, created_by, created_at
                ) VALUES (?, ?, ?, ?, ?)
                """,
                (workspace_id, name, key, _user_id(), created_at),
            )
            conn.commit()
        reasons = _merged_low_stock_reasons(conn, workspace_id)

    return jsonify({"success": True, "data": reasons, "count": len(reasons)}), 201


@dsr_market_bp.route("/visit-issues", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def list_visit_issues():
    workspace_id = get_workspace_id()
    with sqlite3.connect(_db_path()) as conn:
        _ensure_table(conn)
        issues = _merged_visit_issues(conn, workspace_id)
    return jsonify(
        {
            "success": True,
            "data": issues,
            "defaults": list(DEFAULT_VISIT_ISSUES),
            "count": len(issues),
        }
    )


@dsr_market_bp.route("/visit-issues", methods=["POST"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def add_visit_issue():
    workspace_id = get_workspace_id()
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or data.get("issue_name") or "").strip()
    if not name:
        return jsonify({"success": False, "error": {"message": "name is required"}}), 400
    key = _brand_key(name)
    if not key:
        return jsonify({"success": False, "error": {"message": "name is required"}}), 400
    if key == _brand_key("Other"):
        return jsonify(
            {"success": False, "error": {"message": "Use Other free-text instead"}}
        ), 400

    created_at = datetime.now(timezone.utc).isoformat()
    with sqlite3.connect(_db_path()) as conn:
        _ensure_table(conn)
        existing = conn.execute(
            "SELECT issue_name FROM dsr_visit_issues "
            "WHERE workspace_id = ? AND issue_key = ?",
            (workspace_id, key),
        ).fetchone()
        if existing is None and key not in {_brand_key(b) for b in DEFAULT_VISIT_ISSUES}:
            conn.execute(
                """
                INSERT INTO dsr_visit_issues (
                    workspace_id, issue_name, issue_key, created_by, created_at
                ) VALUES (?, ?, ?, ?, ?)
                """,
                (workspace_id, name, key, _user_id(), created_at),
            )
            conn.commit()
        issues = _merged_visit_issues(conn, workspace_id)

    return jsonify({"success": True, "data": issues, "count": len(issues)}), 201


@dsr_market_bp.route("/placement-categories", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def list_placement_categories():
    workspace_id = get_workspace_id()
    with sqlite3.connect(_db_path()) as conn:
        _ensure_table(conn)
        categories = _merged_placement_categories(conn, workspace_id)
    return jsonify(
        {
            "success": True,
            "data": categories,
            "defaults": list(DEFAULT_PLACEMENT_CATEGORIES),
            "count": len(categories),
        }
    )


@dsr_market_bp.route("/placement-categories", methods=["POST"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def add_placement_category():
    workspace_id = get_workspace_id()
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or data.get("category_name") or "").strip()
    if not name:
        return jsonify({"success": False, "error": {"message": "name is required"}}), 400
    key = _brand_key(name)
    if not key:
        return jsonify({"success": False, "error": {"message": "name is required"}}), 400
    if key == _brand_key("Other"):
        return jsonify(
            {"success": False, "error": {"message": "Use Other free-text instead"}}
        ), 400

    created_at = datetime.now(timezone.utc).isoformat()
    with sqlite3.connect(_db_path()) as conn:
        _ensure_table(conn)
        existing = conn.execute(
            "SELECT category_name FROM dsr_placement_categories "
            "WHERE workspace_id = ? AND category_key = ?",
            (workspace_id, key),
        ).fetchone()
        if existing is None and key not in {
            _brand_key(b) for b in DEFAULT_PLACEMENT_CATEGORIES
        }:
            conn.execute(
                """
                INSERT INTO dsr_placement_categories (
                    workspace_id, category_name, category_key, created_by, created_at
                ) VALUES (?, ?, ?, ?, ?)
                """,
                (workspace_id, name, key, _user_id(), created_at),
            )
            conn.commit()
        categories = _merged_placement_categories(conn, workspace_id)

    return jsonify({"success": True, "data": categories, "count": len(categories)}), 201


@dsr_market_bp.route("/open-days", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def open_days():
    workspace_id = get_workspace_id()
    uid = _user_id()
    with sqlite3.connect(_db_path()) as conn:
        _ensure_table(conn)
        dates = _open_visit_dates(conn, workspace_id, uid)
        draft_count = 0
        if uid is not None:
            draft_count = conn.execute(
                "SELECT COUNT(*) FROM dsr_market_visits "
                "WHERE workspace_id = ? AND user_id = ? AND is_draft = 1",
                (workspace_id, uid),
            ).fetchone()[0]
    return jsonify(
        {
            "success": True,
            "data": {
                "open_dates": dates,
                "count": len(dates),
                "max_open_days": 2,
                "draft_count": draft_count,
            },
        }
    )


@dsr_market_bp.route("/drafts", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def list_drafts():
    workspace_id = get_workspace_id()
    uid = _user_id()
    visit_date = (request.args.get("date") or "").strip()
    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_table(conn)
        query = "SELECT * FROM dsr_market_visits WHERE workspace_id = ? AND is_draft = 1"
        params: list = [workspace_id]
        if uid is not None and request.args.get("all") != "1":
            query += " AND user_id = ?"
            params.append(uid)
        if visit_date:
            query += " AND visit_date = ?"
            params.append(visit_date)
        query += " ORDER BY visit_date DESC, id DESC"
        rows = conn.execute(query, tuple(params)).fetchall()
    return jsonify({"success": True, "data": [_row_to_dict(r) for r in rows], "count": len(rows)})


def _score_party_candidate(payload: dict, party: dict, kind: str) -> tuple[int, list[str]]:
    score = 0
    reasons: list[str] = []
    p_name = _norm_name(payload.get("name"))
    c_name = _norm_name(party.get("name") or party.get("firm_name"))
    if p_name and c_name:
        if p_name == c_name:
            score += 50
            reasons.append("Exact firm name")
        elif p_name in c_name or c_name in p_name:
            score += 30
            reasons.append("Similar firm name")

    p_city = _norm_name(payload.get("city") or payload.get("location") or payload.get("city_area"))
    c_city = _norm_name(party.get("city") or party.get("location") or party.get("territory"))
    if p_city and c_city and (p_city == c_city or p_city in c_city or c_city in p_city):
        score += 20
        reasons.append("City / area match")

    p_phone = _norm_phone(payload.get("phone") or payload.get("contact_nos") or payload.get("phone_number"))
    c_phone = _norm_phone(
        party.get("phone_number") or party.get("phone") or party.get("phone_number_2")
    )
    if p_phone and c_phone and (p_phone == c_phone or p_phone[-10:] == c_phone[-10:]):
        score += 25
        reasons.append("Phone match")

    overlap = _address_overlap(payload.get("address"), party.get("address"))
    if overlap >= 0.5:
        score += 15
        reasons.append("Address strongly similar")
    elif overlap >= 0.25:
        score += 8
        reasons.append("Address partly similar")

    if kind == "retailer":
        try:
            want_dist = int(payload.get("distributor_id") or payload.get("linked_distributor_id") or 0)
        except (TypeError, ValueError):
            want_dist = 0
        try:
            got_dist = int(party.get("distributor_id") or 0)
        except (TypeError, ValueError):
            got_dist = 0
        if want_dist and got_dist and want_dist == got_dist:
            score += 15
            reasons.append("Same distributor")

    return score, reasons


@dsr_market_bp.route("/party-match", methods=["POST"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def party_match():
    workspace_id = get_workspace_id()
    data = request.get_json(silent=True) or {}
    kind = (data.get("kind") or data.get("draft_party_kind") or "retailer").strip().lower()
    if kind not in {"retailer", "distributor"}:
        return jsonify({"success": False, "error": {"message": "kind must be retailer or distributor"}}), 400

    db = _master_db()
    if kind == "retailer":
        parties = db.list_master_retailers(limit=800, workspace_id=workspace_id) or []
    else:
        parties = db.list_master_distributors(limit=800, workspace_id=workspace_id) or []

    scored = []
    for party in parties:
        score, reasons = _score_party_candidate(data, party, kind)
        if score >= 30:
            scored.append(
                {
                    "id": party.get("id"),
                    "kind": kind,
                    "name": party.get("name") or party.get("firm_name"),
                    "city": party.get("city") or party.get("location"),
                    "phone": party.get("phone_number") or party.get("phone"),
                    "address": party.get("address"),
                    "distributor_id": party.get("distributor_id"),
                    "score": score,
                    "reasons": reasons,
                    "needs_user_help": score < 70,
                }
            )
    scored.sort(key=lambda x: x["score"], reverse=True)
    top = scored[:10]
    ambiguous = len(top) > 1 and top[0]["score"] < 85
    return jsonify(
        {
            "success": True,
            "data": {
                "candidates": top,
                "ambiguous": ambiguous or any(c.get("needs_user_help") for c in top[:3]),
                "message": (
                    "Possible matches found — please confirm which party, or create new."
                    if top
                    else "No close match — safe to create new party."
                ),
            },
        }
    )


@dsr_market_bp.route("/drafts/<int:visit_id>/resolve", methods=["POST"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def resolve_draft(visit_id: int):
    workspace_id = get_workspace_id()
    uid = _user_id()
    data = request.get_json(silent=True) or {}
    force_create = _truthy_flag(str(data.get("force_create") or ""))
    link_party_id = data.get("link_party_id")
    try:
        link_party_id = int(link_party_id) if link_party_id not in (None, "") else None
    except (TypeError, ValueError):
        link_party_id = None

    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_table(conn)
        row = conn.execute(
            "SELECT * FROM dsr_market_visits WHERE id = ? AND workspace_id = ?",
            (visit_id, workspace_id),
        ).fetchone()
        if row is None:
            return jsonify({"success": False, "error": {"message": "Visit not found"}}), 404
        visit = dict(row)
        if uid is not None and visit.get("user_id") not in (None, uid) and request.args.get("all") != "1":
            return jsonify({"success": False, "error": {"message": "Not your visit"}}), 403
        if not visit.get("is_draft"):
            return jsonify({"success": False, "error": {"message": "Visit is not a draft"}}), 400

        kind = (data.get("kind") or visit.get("draft_party_kind") or "retailer").strip().lower()
        if kind not in {"retailer", "distributor"}:
            kind = "retailer"

        name = (data.get("name") or data.get("customer_name") or visit.get("customer_name") or "").strip()
        if not name:
            return jsonify({"success": False, "error": {"message": "name is required"}}), 400

        owner = (data.get("owner_name") or data.get("contact_person") or visit.get("owner_name") or "").strip() or None
        phone = (data.get("phone") or data.get("contact_nos") or visit.get("contact_nos") or "").strip() or None
        address = (data.get("address") or visit.get("address") or "").strip() or None
        city = (
            data.get("city")
            or data.get("location")
            or data.get("city_area")
            or visit.get("city_area")
            or visit.get("area_text")
            or ""
        ).strip() or None
        distributor_id = data.get("distributor_id") or visit.get("linked_distributor_id")
        try:
            distributor_id = int(distributor_id) if distributor_id not in (None, "") else None
        except (TypeError, ValueError):
            distributor_id = None

        match_payload = {
            "name": name,
            "city": city,
            "location": city,
            "phone": phone,
            "address": address,
            "distributor_id": distributor_id,
        }

        db = _master_db()
        if link_party_id is None and not force_create:
            if kind == "retailer":
                parties = db.list_master_retailers(limit=800, workspace_id=workspace_id) or []
            else:
                parties = db.list_master_distributors(limit=800, workspace_id=workspace_id) or []
            scored = []
            for party in parties:
                score, reasons = _score_party_candidate(match_payload, party, kind)
                if score >= 30:
                    scored.append(
                        {
                            "id": party.get("id"),
                            "kind": kind,
                            "name": party.get("name") or party.get("firm_name"),
                            "city": party.get("city") or party.get("location"),
                            "phone": party.get("phone_number") or party.get("phone"),
                            "address": party.get("address"),
                            "score": score,
                            "reasons": reasons,
                            "needs_user_help": score < 70,
                        }
                    )
            scored.sort(key=lambda x: x["score"], reverse=True)
            top = scored[:10]
            if top:
                return jsonify(
                    {
                        "success": False,
                        "error": {
                            "message": "Possible duplicate — confirm existing party or force create.",
                            "code": "possible_duplicate",
                        },
                        "data": {"candidates": top, "ambiguous": True},
                    }
                ), 409

        party_id = link_party_id
        if party_id is None:
            if kind == "retailer":
                party_id = db.add_master_retailer(
                    name=name,
                    distributor_id=distributor_id,
                    location=city,
                    phone_number=phone,
                    address=address,
                    contact_person=owner,
                    workspace_id=workspace_id,
                )
            else:
                party_id = db.add_master_distributor(
                    name=name,
                    firm_name=name,
                    location=city,
                    address=address,
                    phone_number=phone,
                    workspace_id=workspace_id,
                )

        linked_retailer_id = party_id if kind == "retailer" else visit.get("linked_retailer_id")
        linked_distributor_id = (
            party_id if kind == "distributor" else (distributor_id or visit.get("linked_distributor_id"))
        )

        conn.execute(
            """
            UPDATE dsr_market_visits SET
                customer_name = ?,
                owner_name = ?,
                contact_nos = ?,
                address = ?,
                city_area = ?,
                location = COALESCE(?, location),
                is_draft = 0,
                draft_party_kind = NULL,
                linked_distributor_id = ?,
                linked_retailer_id = ?,
                existing_or_new = 'Existing'
            WHERE id = ?
            """,
            (
                name,
                owner,
                phone,
                address,
                city,
                city,
                linked_distributor_id,
                linked_retailer_id,
                visit_id,
            ),
        )
        conn.commit()
        updated = conn.execute(
            "SELECT * FROM dsr_market_visits WHERE id = ?", (visit_id,)
        ).fetchone()

    return jsonify(
        {
            "success": True,
            "data": {
                "visit": _row_to_dict(updated),
                "party_id": party_id,
                "party_kind": kind,
            },
            "message": "Draft resolved and party master updated",
        }
    )


@dsr_market_bp.route("/day-close", methods=["POST"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def day_close():
    workspace_id = get_workspace_id()
    uid = _user_id()
    if uid is None:
        return jsonify({"success": False, "error": {"message": "User id required"}}), 400
    data = request.get_json(silent=True) or {}
    visit_date = (data.get("date") or data.get("visit_date") or "").strip()
    if not visit_date:
        return jsonify({"success": False, "error": {"message": "date is required (YYYY-MM-DD)"}}), 400

    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_table(conn)
        open_drafts = conn.execute(
            "SELECT id, customer_name FROM dsr_market_visits "
            "WHERE workspace_id = ? AND user_id = ? AND visit_date = ? AND is_draft = 1",
            (workspace_id, uid, visit_date),
        ).fetchall()
        if open_drafts:
            return jsonify(
                {
                    "success": False,
                    "error": {
                        "message": (
                            f"{len(open_drafts)} draft party(ies) still open for {visit_date}. "
                            "Resolve drafts first, then Day Close."
                        ),
                        "code": "drafts_pending",
                    },
                    "data": {"drafts": [_row_to_dict(r) for r in open_drafts]},
                }
            ), 400

        closed_at = datetime.now(timezone.utc).isoformat()
        conn.execute(
            """
            INSERT INTO dsr_day_closures (workspace_id, user_id, visit_date, closed_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(workspace_id, user_id, visit_date) DO UPDATE SET closed_at = excluded.closed_at
            """,
            (workspace_id, uid, visit_date, closed_at),
        )
        conn.commit()
        open_dates = _open_visit_dates(conn, workspace_id, uid)

    return jsonify(
        {
            "success": True,
            "data": {
                "closed_date": visit_date,
                "closed_at": closed_at,
                "open_dates": open_dates,
            },
            "message": f"Day closed for {visit_date}",
        }
    )


def _parse_notes(raw) -> list:
    if raw is None or raw == "":
        return []
    if isinstance(raw, list):
        return raw
    try:
        parsed = json.loads(raw) if isinstance(raw, str) else raw
        return parsed if isinstance(parsed, list) else []
    except (TypeError, ValueError, json.JSONDecodeError):
        return []


def _approach_row_to_dict(row: sqlite3.Row) -> dict:
    d = dict(row)
    d["notes"] = _parse_notes(d.pop("notes_json", None))
    cats = d.get("main_categories")
    if isinstance(cats, str) and cats.strip().startswith("["):
        try:
            d["main_categories_list"] = json.loads(cats)
        except (TypeError, ValueError, json.JSONDecodeError):
            d["main_categories_list"] = [c.strip() for c in cats.split(",") if c.strip()]
    elif isinstance(cats, str) and cats.strip():
        d["main_categories_list"] = [c.strip() for c in cats.split(",") if c.strip()]
    else:
        d["main_categories_list"] = []
    return d


def _categories_to_store(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, list):
        cleaned = [str(x).strip() for x in value if str(x).strip()]
        return json.dumps(cleaned, ensure_ascii=False) if cleaned else None
    text = str(value).strip()
    return text or None


def _upsert_approach_from_visit(
    conn: sqlite3.Connection,
    *,
    workspace_id: str,
    uid,
    username,
    data: dict,
    visit_id: int,
    visit_intel_json: str | None,
) -> int | None:
    """Mirror New-distributor visit fields into Approach Distributor list."""
    firm = (data.get("customer_name") or "").strip()
    if not firm:
        return None

    monthly_ht = None
    main_categories = None
    if visit_intel_json:
        try:
            intel = json.loads(visit_intel_json)
            if isinstance(intel, dict):
                monthly_ht = (str(intel.get("annual_ht_business") or intel.get("monthly_ht_business") or "").strip() or None)
                cats = intel.get("main_categories")
                main_categories = _categories_to_store(cats)
        except (TypeError, ValueError, json.JSONDecodeError):
            pass
    if main_categories is None:
        parts = []
        if _truthy_flag(str(data.get("bed") or "")):
            parts.append("Bedsheets")
        if _truthy_flag(str(data.get("bath") or "")):
            parts.append("Towels")
        if _truthy_flag(str(data.get("tob") or "")):
            parts.append("Comforters / Blankets / Pillows")
        if _truthy_flag(str(data.get("others") or "")):
            parts.append("Other")
        main_categories = _categories_to_store(parts)

    owner_name = (data.get("owner_name") or "").strip() or None
    contact_nos = (data.get("contact_nos") or "").strip() or None
    city_area = (data.get("city_area") or data.get("area_text") or "").strip() or None
    location = (data.get("location") or "").strip() or None
    address = (data.get("address") or "").strip() or None
    monthly_ht = monthly_ht or (
        str(data.get("annual_ht") or data.get("monthly_ht") or "").strip() or None
    )
    channel_type = (data.get("channel_type") or "AWD").strip() or "AWD"
    existing_or_new = (data.get("existing_or_new") or "New").strip() or "New"
    customer_type = (data.get("customer_type") or "").strip() or None
    now = datetime.now(timezone.utc).isoformat()

    existing = conn.execute(
        """
        SELECT id FROM dsr_approach_distributors
        WHERE workspace_id = ? AND lower(firm_name) = lower(?)
        ORDER BY id DESC LIMIT 1
        """,
        (workspace_id, firm),
    ).fetchone()

    if existing:
        conn.execute(
            """
            UPDATE dsr_approach_distributors SET
                owner_name = COALESCE(?, owner_name),
                contact_nos = COALESCE(?, contact_nos),
                city_area = COALESCE(?, city_area),
                location = COALESCE(?, location),
                address = COALESCE(?, address),
                monthly_ht = COALESCE(?, monthly_ht),
                main_categories = COALESCE(?, main_categories),
                channel_type = COALESCE(?, channel_type),
                existing_or_new = COALESCE(?, existing_or_new),
                customer_type = COALESCE(?, customer_type),
                source_visit_id = ?,
                updated_at = ?,
                user_id = COALESCE(user_id, ?),
                username = COALESCE(username, ?)
            WHERE id = ?
            """,
            (
                owner_name,
                contact_nos,
                city_area,
                location,
                address,
                monthly_ht,
                main_categories,
                channel_type,
                existing_or_new,
                customer_type,
                visit_id,
                now,
                uid,
                username,
                int(existing["id"]),
            ),
        )
        return int(existing["id"])

    cur = conn.execute(
        """
        INSERT INTO dsr_approach_distributors (
            workspace_id, user_id, username, firm_name,
            owner_name, contact_nos, city_area, location, address,
            monthly_ht, main_categories, channel_type, existing_or_new, customer_type,
            source_visit_id, notes_json, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            workspace_id,
            uid,
            username,
            firm,
            owner_name,
            contact_nos,
            city_area,
            location,
            address,
            monthly_ht,
            main_categories,
            channel_type,
            existing_or_new,
            customer_type,
            visit_id,
            "[]",
            now,
            now,
        ),
    )
    return int(cur.lastrowid)


@dsr_market_bp.route("/approach-distributors", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def list_approach_distributors():
    workspace_id = get_workspace_id()
    q = (request.args.get("q") or "").strip().lower()
    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_table(conn)
        rows = conn.execute(
            """
            SELECT * FROM dsr_approach_distributors
            WHERE workspace_id = ?
            ORDER BY updated_at DESC, id DESC
            """,
            (workspace_id,),
        ).fetchall()
    items = [_approach_row_to_dict(r) for r in rows]
    if q:
        items = [
            x
            for x in items
            if q in (x.get("firm_name") or "").lower()
            or q in (x.get("owner_name") or "").lower()
            or q in (x.get("city_area") or "").lower()
            or q in (x.get("contact_nos") or "").lower()
            or q in (x.get("location") or "").lower()
        ]
    return jsonify({"success": True, "data": items, "count": len(items)})


@dsr_market_bp.route("/approach-distributors", methods=["POST"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def create_approach_distributor():
    workspace_id = get_workspace_id()
    user = _current_user()
    data = request.get_json(silent=True) or {}
    firm = (data.get("firm_name") or data.get("customer_name") or data.get("name") or "").strip()
    if not firm:
        return jsonify({"success": False, "error": {"message": "firm_name is required"}}), 400

    now = datetime.now(timezone.utc).isoformat()
    main_categories = _categories_to_store(
        data.get("main_categories_list")
        if data.get("main_categories_list") is not None
        else data.get("main_categories")
    )
    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_table(conn)
        # Upsert by firm name so Customers form + Market Visit stay one record.
        existing = conn.execute(
            """
            SELECT id FROM dsr_approach_distributors
            WHERE workspace_id = ? AND lower(firm_name) = lower(?)
            ORDER BY id DESC LIMIT 1
            """,
            (workspace_id, firm),
        ).fetchone()
        fields = (
            (data.get("owner_name") or "").strip() or None,
            (data.get("contact_nos") or data.get("phone") or "").strip() or None,
            (data.get("city_area") or data.get("city") or "").strip() or None,
            (data.get("location") or "").strip() or None,
            (data.get("address") or "").strip() or None,
            (str(data.get("annual_ht") or data.get("monthly_ht") or "").strip() or None),
            main_categories,
            (data.get("channel_type") or "AWD").strip() or "AWD",
            (data.get("existing_or_new") or "New").strip() or "New",
            (data.get("customer_type") or "").strip() or None,
        )
        if existing:
            conn.execute(
                """
                UPDATE dsr_approach_distributors SET
                    owner_name = COALESCE(?, owner_name),
                    contact_nos = COALESCE(?, contact_nos),
                    city_area = COALESCE(?, city_area),
                    location = COALESCE(?, location),
                    address = COALESCE(?, address),
                    monthly_ht = COALESCE(?, monthly_ht),
                    main_categories = COALESCE(?, main_categories),
                    channel_type = COALESCE(?, channel_type),
                    existing_or_new = COALESCE(?, existing_or_new),
                    customer_type = COALESCE(?, customer_type),
                    updated_at = ?
                WHERE id = ?
                """,
                (*fields, now, int(existing["id"])),
            )
            row_id = int(existing["id"])
        else:
            cur = conn.execute(
                """
                INSERT INTO dsr_approach_distributors (
                    workspace_id, user_id, username, firm_name,
                    owner_name, contact_nos, city_area, location, address,
                    monthly_ht, main_categories, channel_type, existing_or_new, customer_type,
                    source_visit_id, notes_json, created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    workspace_id,
                    _user_id(),
                    user.get("username"),
                    firm,
                    *fields,
                    data.get("source_visit_id"),
                    "[]",
                    now,
                    now,
                ),
            )
            row_id = int(cur.lastrowid)
        conn.commit()
        row = conn.execute(
            "SELECT * FROM dsr_approach_distributors WHERE id = ?", (row_id,)
        ).fetchone()
    return jsonify({"success": True, "data": _approach_row_to_dict(row)}), 201


@dsr_market_bp.route("/approach-distributors/<int:approach_id>", methods=["GET"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def get_approach_distributor(approach_id: int):
    workspace_id = get_workspace_id()
    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_table(conn)
        row = conn.execute(
            "SELECT * FROM dsr_approach_distributors WHERE id = ? AND workspace_id = ?",
            (approach_id, workspace_id),
        ).fetchone()
    if not row:
        return jsonify({"success": False, "error": {"message": "Not found"}}), 404
    return jsonify({"success": True, "data": _approach_row_to_dict(row)})


@dsr_market_bp.route("/approach-distributors/<int:approach_id>/notes", methods=["POST"])
@require_jwt_auth
@require_role("admin", "sales_executive")
def add_approach_distributor_note(approach_id: int):
    """Append a free-text note. Linking to other entities comes later."""
    workspace_id = get_workspace_id()
    user = _current_user()
    data = request.get_json(silent=True) or {}
    text = (data.get("text") or data.get("note") or "").strip()
    if not text:
        return jsonify({"success": False, "error": {"message": "note text is required"}}), 400

    now = datetime.now(timezone.utc).isoformat()
    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_table(conn)
        row = conn.execute(
            "SELECT * FROM dsr_approach_distributors WHERE id = ? AND workspace_id = ?",
            (approach_id, workspace_id),
        ).fetchone()
        if not row:
            return jsonify({"success": False, "error": {"message": "Not found"}}), 404
        notes = _parse_notes(row["notes_json"])
        notes.append(
            {
                "id": f"n-{int(datetime.now(timezone.utc).timestamp() * 1000)}",
                "text": text,
                "created_at": now,
                "username": user.get("username"),
                "user_id": _user_id(),
                # link_type / link_id reserved for later
            }
        )
        conn.execute(
            """
            UPDATE dsr_approach_distributors
            SET notes_json = ?, updated_at = ?
            WHERE id = ?
            """,
            (json.dumps(notes, ensure_ascii=False), now, approach_id),
        )
        conn.commit()
        updated = conn.execute(
            "SELECT * FROM dsr_approach_distributors WHERE id = ?", (approach_id,)
        ).fetchone()
    return jsonify({"success": True, "data": _approach_row_to_dict(updated)})
