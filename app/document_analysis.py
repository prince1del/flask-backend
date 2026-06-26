import csv
import os
import re
from pathlib import Path
from typing import Any

from rapidfuzz import fuzz

from openpyxl import load_workbook


def _fuzzy_similarity(left: str, right: str) -> float:
    left_text = normalize_name(left)
    right_text = normalize_name(right)
    if not left_text or not right_text:
        return 0.0
    if left_text == right_text:
        return 1.0
    return fuzz.ratio(left_text, right_text) / 100.0


def normalize_name(value: str) -> str:
    return " ".join(part.lower() for part in value.replace("-", " ").split() if part)


def match_person_names(left: str, right: str, aliases: dict[str, list[str]] | None = None) -> bool:
    aliases = aliases or {}
    normalized_left = normalize_name(left)
    normalized_right = normalize_name(right)

    if not normalized_left or not normalized_right:
        return False

    if normalized_left == normalized_right:
        return True

    left_tokens = normalized_left.split()
    right_tokens = normalized_right.split()
    left_set = set(left_tokens)
    right_set = set(right_tokens)

    if left_set and right_set and (left_set & right_set):
        return True

    left_initials = "".join(token[0] for token in left_tokens if token)
    right_initials = "".join(token[0] for token in right_tokens if token)
    if left_initials and right_initials and (left_initials.startswith(right_initials) or right_initials.startswith(left_initials)):
        return True

    if _fuzzy_similarity(left, right) >= 0.6:
        return True

    alias_lookup = {normalize_name(item): True for values in aliases.values() for item in values}
    if normalize_name(left) in alias_lookup and normalize_name(right) in alias_lookup:
        return True

    if len(left_tokens) <= 2 and len(right_tokens) <= 2:
        return bool(left_set & right_set)

    return False


def _read_csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _read_excel_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    data_rows = []
    for row in rows[1:]:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        data_rows.append({headers[index]: row[index] for index in range(min(len(headers), len(row)))})
    return data_rows


def _read_pdf_rows(path: Path) -> list[dict[str, Any]]:
    text = path.read_text(encoding="utf-8", errors="ignore")
    payload: dict[str, Any] = {}
    lines = [line.strip() for line in text.splitlines() if line.strip()]

    for line in lines:
        if ":" not in line:
            continue
        key, value = line.split(":", 1)
        normalized_key = key.strip().lower().replace(" ", "_")
        field_map = {
            "product": "product",
            "quantity": "quantity",
            "rate": "rate",
            "gst": "gst",
            "discount": "discount",
            "client_name": "client_name",
            "invoice_amount": "invoice_amount",
            "client": "client_name",
            "invoice": "invoice_amount",
            "invoiceamt": "invoice_amount",
            "invoiceamount": "invoice_amount",
            "total_amount": "invoice_amount",
        }
        payload[field_map.get(normalized_key, normalized_key)] = value.strip()

    if not payload:
        return []
    return [payload]


def _normalize_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(value)
    text = str(value).strip()
    if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text):
        return text
    return text.lower()


def _extract_document_data(path: Path) -> dict[str, Any]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows = _read_csv_rows(path)
    elif suffix in {".xlsx", ".xlsm"}:
        rows = _read_excel_rows(path)
    elif suffix == ".pdf":
        rows = _read_pdf_rows(path)
    else:
        rows = []

    if not rows:
        return {"file": str(path), "rows": []}
    return {"file": str(path), "rows": rows}


def analyze_documents(paths: list[Path | str], aliases: dict[str, list[str]] | None = None) -> dict[str, Any]:
    documents = [_extract_document_data(Path(path)) for path in paths]
    mismatches: list[dict[str, Any]] = []
    person_match_count = 0

    if len(documents) < 2:
        return {
            "documents": documents,
            "summary": {"mismatch_count": 0, "person_match_count": 0, "status": "insufficient-data"},
            "mismatches": [],
        }

    reference = documents[0]["rows"][0] if documents[0]["rows"] else {}
    for document in documents[1:]:
        if not document["rows"]:
            continue
        row = document["rows"][0]
        for field in ["product", "quantity", "rate", "gst", "discount", "client_name", "invoice_amount"]:
            reference_value = reference.get(field)
            current_value = row.get(field)
            if field == "client_name":
                matched = match_person_names(str(reference_value or ""), str(current_value or ""), aliases)
                if not matched:
                    mismatches.append({"field": field, "source": reference.get(field), "target": row.get(field), "message": "Client name mismatch"})
                else:
                    person_match_count += 1
                    mismatches.append({
                        "field": field,
                        "source": reference.get(field),
                        "target": row.get(field),
                        "message": "Client name matched as alias or similar name",
                    })
                continue

            if _normalize_value(reference_value) != _normalize_value(current_value):
                mismatches.append({
                    "field": field,
                    "source": reference_value,
                    "target": current_value,
                    "message": f"{field} mismatch",
                })

    effective_mismatches = [
        item for item in mismatches if item.get("message") != "Client name matched as alias or similar name"
    ]

    return {
        "documents": documents,
        "summary": {
            "mismatch_count": len(effective_mismatches),
            "person_match_count": person_match_count,
            "status": "ok" if not effective_mismatches else "mismatches-found",
        },
        "mismatches": mismatches,
    }
