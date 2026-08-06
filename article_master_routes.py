"""
Article Master — Flask Routes (per-user isolation)

Each logged-in user sees and edits only their own catalog.
"""

import io
import os
import sqlite3
import tempfile
from pathlib import Path

import pandas as pd
from flask import Blueprint, current_app, jsonify, request, send_file

import article_master_db as amdb
import article_master_parser as amparser
from app.routes.auth import get_workspace_id, require_jwt_auth

article_master_bp = Blueprint("article_master", __name__, url_prefix="/api/v1/article-master")

DEFAULT_KEY_FIELDS = ["brand", "size"]

# Canonical Article Master layout — locked to user's preferred export
# (Desktop Article_Master_All.xlsx, Aug 2026). All downloads use this order.
# Blank/spacer columns from hand-edited Excels are ignored.
CANONICAL_ARTICLE_MASTER_COLUMNS = [
    "Category",
    "Product",
    "Brand",
    "Size",
    "TC",
    "Units",
    "BS Size",
    "Pillow Size",
    "Color",
    "Pillow Stitching Style",
    "Print Style",
    "Blend",
    "Packing",
    "Bale Pack Size",
    "MRP",
    "AWD Mark up on Exmill",
    "Ex-Mill",
    "Proposed Customer Discount",
    "Retailer Margin",
    "PTR",
]

# Legacy name kept for older call sites; every category maps to the same layout.
ORIGINAL_TEMPLATES = {
    "All": list(CANONICAL_ARTICLE_MASTER_COLUMNS),
    "Bed": list(CANONICAL_ARTICLE_MASTER_COLUMNS),
    "TOB": list(CANONICAL_ARTICLE_MASTER_COLUMNS),
    "Bath": list(CANONICAL_ARTICLE_MASTER_COLUMNS),
    "Pillow": list(CANONICAL_ARTICLE_MASTER_COLUMNS),
}

# Extra header aliases when resolving a canonical export column from stored data.
EXPORT_COLUMN_ALIASES = {
    "bs size": ["bs size", "bedset size (cms)", "bedset size", "bed set size"],
    "pillow size": ["pillow size", "pillow size (cms)"],
    "bale pack size": ["bale pack size", "bale size", "bale pack sizes", "pack sizes", "pack size"],
    "ex-mill": ["ex-mill", "exmill price", "ex mill", "ex-mill per pcs", "ex mill per pcs"],
    "awd mark up on exmill": [
        "awd mark up on exmill", "awd markup on exmill", "awd md", "awd mu",
        "distributor mark up", "mark up on exmill",
    ],
    "retailer margin": ["retailer margin", "retail mark down", "retailer md", "retailer markdown"],
    "print style": ["print style", "print/dyed/weave", "print dyed weave"],
    "color": ["color", "colour", "shade"],
    "proposed customer discount": [
        "proposed customer discount", "perceived", "perceive", "perceived margin",
    ],
    "perceived": [
        "proposed customer discount", "perceived", "perceive", "perceived margin",
    ],
}


def _resolve_export_value(article, column_name):
    """
    Resolve a canonical export column from core fields + extra_attributes.
    """
    if amparser.is_excluded_extra_column(column_name):
        return None

    # Direct core fields for Category / Product
    lower = str(column_name or "").strip().lower()
    if lower == "category":
        return article.get("category")
    if lower == "product":
        return article.get("product_type")
    if lower == "size":
        return amparser.size_display_name(article.get("size"))

    core_field = amparser.resolve_core_field_for_name(column_name)
    if core_field:
        return article.get(core_field)

    extra = amparser.strip_excluded_extra_attributes(article.get("extra_attributes") or {})
    aliases = EXPORT_COLUMN_ALIASES.get(lower, [lower])
    # Always try the exact column name first
    alias_list = [column_name.strip()] + [a for a in aliases if a != lower]
    # Also include the normalized lower forms against extra keys
    extra_lower = {str(k).strip().lower(): v for k, v in extra.items()}
    for alias in alias_list:
        if alias in extra:
            val = extra[alias]
            break
        hit = extra_lower.get(str(alias).strip().lower())
        if hit is not None and hit != "":
            val = hit
            break
    else:
        return None

    if lower in {"awd mark up on exmill", "retailer margin", "perceived", "proposed customer discount"}:
        return amparser.format_percent_display(val)
    return val


def _sanitize_for_json(value):
    """Convert pandas/numpy scalars so jsonify never crashes on upload results."""
    if isinstance(value, dict):
        return {k: _sanitize_for_json(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_sanitize_for_json(v) for v in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            pass
    return value


def _cleanup_temp_file(path):
    try:
        if path and os.path.exists(path):
            os.remove(path)
    except OSError:
        pass


def _get_db_connection():
    db_path = current_app.config.get("DATABASE_PATH", "centralized_db.sqlite3")
    conn = sqlite3.connect(db_path)
    amdb.ensure_schema(conn)
    return conn


def _price_change_meta(existing_val, upload_val):
    try:
        old = float(existing_val)
        new = float(upload_val)
        delta = new - old
        if abs(delta) < 0.005:
            return {"direction": "same", "delta": 0, "pct": 0}
        pct = (delta / old * 100) if old else None
        return {
            "direction": "increase" if delta > 0 else "decrease",
            "delta": round(delta, 2),
            "pct": round(pct, 1) if pct is not None else None,
        }
    except (TypeError, ValueError):
        return {"direction": "changed", "delta": None, "pct": None}


def _build_price_revision_summary(price_comparisons):
    parts = []
    increases = 0
    decreases = 0
    for row in price_comparisons:
        if row.get("status") != "mismatch":
            continue
        meta = row.get("change") or {}
        field = str(row.get("field", "")).upper().replace("_", "-")
        direction = meta.get("direction")
        if direction == "increase":
            increases += 1
            pct = meta.get("pct")
            parts.append(f"{field} ↑" + (f" {pct}%" if pct is not None else ""))
        elif direction == "decrease":
            decreases += 1
            pct = meta.get("pct")
            parts.append(f"{field} ↓" + (f" {pct}%" if pct is not None else ""))
    if not parts:
        return "Seasonal price revision detected."
    trend = "increase" if increases >= decreases else "decrease"
    return f"Seasonal price {trend}: " + ", ".join(parts)


def _get_current_user_id():
    user = getattr(request, "user", None)
    if isinstance(user, dict) and user.get("user_id") is not None:
        return int(user["user_id"])
    raise RuntimeError("Authentication required")


def _get_changed_by():
    user = getattr(request, "user", None)
    if isinstance(user, dict) and user.get("username"):
        return str(user["username"])
    return "user"


def _build_upload_conflict_payload(upload_index, article, classification, key_fields):
    existing = classification["existing"]
    core_fields = amdb._article_core_fields(article)
    extra = article.get("extra_attributes") or {}

    field_comparisons = []
    for field in key_fields:
        file_val = amparser.extract_key_field_value(field, core_fields, extra)
        cand_core = amdb._article_core_fields(existing)
        cand_extra = existing.get("extra_attributes") or {}
        master_val = amparser.extract_key_field_value(field, cand_core, cand_extra)
        field_l = field.lower()
        if field_l == "brand":
            if amparser.brands_match_fuzzy(file_val, master_val):
                status = "match"
            elif not file_val and not master_val:
                status = "both_empty"
            else:
                status = "mismatch"
        else:
            n_file = amparser.normalize_key_part_value(field, file_val)
            n_master = amparser.normalize_key_part_value(field, master_val)
            if n_file and n_master:
                status = "match" if n_file == n_master else "mismatch"
            elif not n_file and not n_master:
                status = "both_empty"
            elif not n_file:
                status = "missing_in_file"
            else:
                status = "missing_in_master"
        field_comparisons.append({
            "field": field,
            "upload_value": file_val if file_val not in (None, "") else None,
            "existing_value": master_val if master_val not in (None, "") else None,
            "status": status,
        })

    price_comparisons = []
    for field in amdb.TRACKED_PRICE_FIELDS:
        upload_val = article.get(field)
        existing_val = existing.get(field)
        change = _price_change_meta(existing_val, upload_val)
        price_comparisons.append({
            "field": field,
            "upload_value": upload_val,
            "existing_value": existing_val,
            "status": "match" if amdb._values_equal(field, existing_val, upload_val) else "mismatch",
            "change": change,
        })

    price_diffs = classification.get("price_diffs") or []
    key_differs = existing["item_key"] != article["item_key"]
    revision_summary = _build_price_revision_summary(price_comparisons)
    parts = []
    if classification.get("duplicate_ids"):
        parts.append(f"{len(classification['duplicate_ids'])} duplicate row(s) in Article Master (e.g. Blumen/Bluman)")
    if price_diffs:
        parts.append("price revision on " + ", ".join(price_diffs))
    if key_differs:
        parts.append("item key differs (brand spelling normalized on replace)")
    if classification.get("conflict_reason") == "duplicate_entries_in_master" and not price_diffs:
        issue_summary = (
            "Duplicate entries for the same product in Article Master. "
            "Replace will keep one row and remove duplicates."
        )
    elif price_diffs:
        issue_summary = revision_summary
    else:
        issue_summary = "; ".join(parts) if parts else "Conflict with existing article"

    return {
        "upload_index": upload_index,
        "upload_item_key": article["item_key"],
        "existing_id": existing["id"],
        "existing_item_key": existing["item_key"],
        "brand": article.get("brand"),
        "size": article.get("size"),
        "category": article["category"],
        "product_type": article.get("product_type"),
        "price_diffs": price_diffs,
        "conflict_reason": classification.get("conflict_reason"),
        "can_create_new": key_differs,
        "field_comparisons": field_comparisons,
        "price_comparisons": price_comparisons,
        "issue_summary": issue_summary,
        "recommended_action": "replace",
        "duplicate_ids": classification.get("duplicate_ids") or [],
    }


def _build_upload_success_message(created, updated, skipped):
    if created > 0:
        suffix = f" ({updated} updated)" if updated else ""
        return f"Successfully added {created} item(s){suffix}."
    if updated > 0:
        return f"Successfully updated {updated} item(s)."
    if skipped > 0:
        return "0 items updated — items already available in Article Master."
    return "Upload complete — no changes made."


@article_master_bp.route("/upload", methods=["POST"])
@require_jwt_auth
def upload_article_sheet():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    user_id = _get_current_user_id()
    workspace_id = get_workspace_id()
    # confirmed_category:
    #   "" / missing  → preview only (no save)
    #   "AUTO"        → save with per-row category detection (mixed sheets OK)
    #   "Bed"/etc     → force EVERY row into that one category
    confirmed_category = (request.form.get("confirmed_category") or "").strip()
    season_tag = amparser.normalize_season_tag(request.form.get("season_tag"))
    suggested_season = amparser.suggest_season_tag_from_filename(file.filename)
    conn = _get_db_connection()

    suffix = Path(file.filename or "upload.xlsx").suffix or ".xlsx"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp_path = tmp.name
    tmp.close()
    file.save(tmp_path)

    try:
        with pd.ExcelFile(tmp_path) as xl:
            sheet_name = xl.sheet_names[0]

        amdb.ensure_default_categories(conn, user_id, workspace_id=workspace_id)
        categories = amdb.get_all_categories(conn, user_id)
        key_fields_lookup = {c["category_name"]: c["key_fields"] for c in categories}

        force = None
        if confirmed_category and confirmed_category.upper() != "AUTO":
            force = confirmed_category
            if force == "TOB Pillow":
                force = "Pillow"

        articles, suggested_category, is_new_category, needs_review, category_breakdown = (
            amparser.parse_article_sheet(
                tmp_path,
                sheet_name,
                key_fields_lookup,
                DEFAULT_KEY_FIELDS,
                forced_category=force,
                source_filename=file.filename,
            )
        )
        articles = amdb.apply_brand_aliases_to_articles(
            conn, user_id, articles, key_fields_lookup, DEFAULT_KEY_FIELDS,
        )

        if not confirmed_category:
            breakdown_text = ", ".join(
                f"{cat}: {count}" for cat, count in sorted(category_breakdown.items())
            )
            return jsonify({
                "status": "confirmation_required",
                "message": (
                    "Each row will be saved under its own category (mixed sheets OK). "
                    "Confirm: AUTO (recommended) or force one category (Bed/Bath/TOB/Pillow)."
                ),
                "detected_category": suggested_category,
                "category_breakdown": category_breakdown,
                "breakdown_text": breakdown_text,
                "suggested_key_fields": DEFAULT_KEY_FIELDS,
                "sample_articles": _sanitize_for_json(articles[:5]),
                "article_count": len(articles),
                "suggested_season_tag": suggested_season,
                "season_tag": season_tag or suggested_season,
            }), 200

        effective_season = season_tag or suggested_season
        if effective_season:
            for article in articles:
                article["season_tag"] = effective_season

        if is_new_category:
            unknown = [
                cat for cat in category_breakdown
                if cat not in key_fields_lookup
            ]
            return jsonify({
                "error": (
                    f"{'This category is' if len(unknown) == 1 else 'These categories are'} not configured: "
                    f"{', '.join(unknown)}. Please create them via confirm-new-category first."
                ),
            }), 400

        created, updated, skipped = 0, 0, 0
        created_by_category = {}
        changed_summary = []
        conflicts = []
        changed_by = _get_changed_by()

        import json as json_module
        conflict_resolutions = {}
        raw_resolutions = (request.form.get("conflict_resolutions") or "").strip()
        if raw_resolutions:
            try:
                parsed = json_module.loads(raw_resolutions)
                if isinstance(parsed, dict):
                    conflict_resolutions = {str(k): str(v) for k, v in parsed.items()}
            except json_module.JSONDecodeError:
                return jsonify({"error": "Invalid conflict_resolutions JSON"}), 400

        for idx, article in enumerate(articles):
            cat = article["category"]
            key_fields = key_fields_lookup.get(cat, DEFAULT_KEY_FIELDS)
            classification = amdb.classify_upload_article(conn, user_id, article, key_fields)

            if classification["action"] == "skip":
                skipped += 1
                continue

            if classification["action"] == "create":
                _, was_created, changed_fields = amdb.upsert_article(
                    conn, user_id, article,
                    source_filename=file.filename,
                    workspace_id=workspace_id,
                    changed_by=changed_by,
                )
                if was_created:
                    created += 1
                    created_by_category[cat] = created_by_category.get(cat, 0) + 1
                elif changed_fields:
                    updated += 1
                    changed_summary.append({"item_key": article["item_key"], "changed": changed_fields})
                continue

            resolution = conflict_resolutions.get(str(idx))
            if not resolution:
                conflicts.append(
                    _build_upload_conflict_payload(idx, article, classification, key_fields)
                )
                continue

            existing = classification["existing"]
            if resolution == "replace":
                updated_row, removed_dupes = amdb.replace_article_and_consolidate(
                    conn, user_id, existing["id"], article,
                    key_fields,
                    source_filename=file.filename,
                    workspace_id=workspace_id,
                    changed_by=changed_by,
                )
                updated += 1
                if removed_dupes:
                    changed_summary.append({
                        "item_key": article["item_key"],
                        "changed": ["consolidated_duplicates"],
                        "removed_duplicates": removed_dupes,
                    })
            elif resolution == "create_new":
                if not amdb.get_article_by_item_key(conn, user_id, article["item_key"]):
                    amdb.insert_article(
                        conn, user_id, article,
                        source_filename=file.filename,
                        workspace_id=workspace_id,
                    )
                    created += 1
                    created_by_category[cat] = created_by_category.get(cat, 0) + 1
                else:
                    skipped += 1
            elif resolution == "skip":
                skipped += 1

        if conflicts:
            return jsonify({
                "status": "price_mismatch_confirmation_required",
                "message": (
                    f"{len(conflicts)} item(s) need review — seasonal price changes and/or duplicate rows "
                    f"(e.g. Blumen vs Bluman). Use Replace all to apply new prices and merge duplicates."
                ),
                "conflicts": _sanitize_for_json(conflicts),
                "category": suggested_category if not force else force,
                "category_breakdown": category_breakdown,
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "created_by_category": created_by_category,
                "total_parsed": len(articles) + len(needs_review),
                "needs_manual_review": _sanitize_for_json(needs_review),
            }), 200

        return jsonify({
            "status": "success",
            "message": _build_upload_success_message(created, updated, skipped),
            "category": suggested_category if not force else force,
            "category_breakdown": category_breakdown,
            "created_by_category": created_by_category,
            "total_parsed": len(articles) + len(needs_review),
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "changed_details": changed_summary,
            "needs_manual_review": _sanitize_for_json(needs_review),
            "duplicate_groups_remaining": len(
                amdb.find_duplicate_groups(conn, user_id, key_fields_lookup, DEFAULT_KEY_FIELDS)
            ),
        }), 200

    except Exception as exc:
        return jsonify({"error": str(exc)}), 500

    finally:
        conn.close()
        _cleanup_temp_file(tmp_path)


@article_master_bp.route("/duplicates", methods=["GET"])
@require_jwt_auth
def list_duplicate_articles():
    try:
        user_id = _get_current_user_id()
        conn = _get_db_connection()
        categories = amdb.get_all_categories(conn, user_id)
        key_fields_lookup = {c["category_name"]: c["key_fields"] for c in categories}
        amdb.normalize_catalog_brand_names(conn, user_id, key_fields_lookup, DEFAULT_KEY_FIELDS)
        groups = amdb.find_duplicate_groups(conn, user_id, key_fields_lookup, DEFAULT_KEY_FIELDS)
        conn.close()
        return jsonify({
            "groups": _sanitize_for_json(groups),
            "group_count": len(groups),
            "duplicate_row_count": sum(len(g["articles"]) for g in groups),
        }), 200
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 401
    except Exception as e:
        current_app.logger.exception("list_duplicate_articles failed")
        return jsonify({"error": str(e)}), 500


@article_master_bp.route("/merge-duplicates", methods=["POST"])
@require_jwt_auth
def merge_duplicate_articles():
    conn = None
    try:
        data = request.get_json() or {}
        merges = data.get("merges")
        auto = data.get("auto", False)
        user_id = _get_current_user_id()
        conn = _get_db_connection()
        changed_by = _get_changed_by()
        categories = amdb.get_all_categories(conn, user_id)
        key_fields_lookup = {c["category_name"]: c["key_fields"] for c in categories}

        if auto:
            groups = amdb.find_duplicate_groups(conn, user_id, key_fields_lookup, DEFAULT_KEY_FIELDS)
            merges = []
            for group in groups:
                article_ids = [a["id"] for a in group["articles"]]
                keep_id = group["suggested_keep_id"]
                remove_ids = [aid for aid in article_ids if aid != keep_id]
                merges.append({
                    "keep_id": keep_id,
                    "remove_ids": remove_ids,
                    "price_from_id": group["suggested_price_from_id"],
                })

        if not merges:
            normalized = amdb.normalize_catalog_brand_names(
                conn, user_id, key_fields_lookup, DEFAULT_KEY_FIELDS,
            )
            conn.close()
            conn = None
            msg = "No duplicate articles found."
            if normalized:
                msg += f" Renamed {normalized} row(s) to canonical brand spelling."
            return jsonify({
                "status": "success",
                "merged_groups": 0,
                "removed_rows": 0,
                "brands_normalized": normalized,
                "message": msg,
            }), 200

        merged_groups = 0
        removed_rows = 0
        for spec in merges:
            keep_id = spec.get("keep_id")
            remove_ids = spec.get("remove_ids") or []
            price_from_id = spec.get("price_from_id")
            if not keep_id or not remove_ids:
                continue
            amdb.merge_articles(
                conn, user_id, int(keep_id), remove_ids,
                price_from_id=price_from_id, changed_by=changed_by,
            )
            merged_groups += 1
            removed_rows += len(remove_ids)

        normalized = amdb.normalize_catalog_brand_names(
            conn, user_id, key_fields_lookup, DEFAULT_KEY_FIELDS,
        )
        conn.close()
        conn = None
        msg = f"Merged {merged_groups} duplicate group(s), removed {removed_rows} row(s)."
        if normalized:
            msg += f" Renamed {normalized} row(s) to canonical brand spelling."
        return jsonify({
            "status": "success",
            "merged_groups": merged_groups,
            "removed_rows": removed_rows,
            "brands_normalized": normalized,
            "message": msg,
        }), 200
    except ValueError as e:
        if conn:
            conn.close()
        return jsonify({"error": str(e)}), 400
    except RuntimeError as e:
        if conn:
            conn.close()
        return jsonify({"error": str(e)}), 401
    except Exception as e:
        if conn:
            conn.close()
        current_app.logger.exception("merge_duplicate_articles failed")
        return jsonify({"error": str(e)}), 500


@article_master_bp.route("/<int:article_id>", methods=["DELETE"])
@require_jwt_auth
def delete_one_article(article_id):
    user_id = _get_current_user_id()
    conn = _get_db_connection()
    try:
        amdb.delete_article(conn, user_id, article_id)
    except ValueError as e:
        conn.close()
        return jsonify({"error": str(e)}), 404
    conn.close()
    return jsonify({"status": "success", "deleted_id": article_id}), 200


@article_master_bp.route("/delete-selected", methods=["POST"])
@require_jwt_auth
def delete_selected_articles():
    data = request.get_json(silent=True) or {}
    raw_ids = data.get("ids") or data.get("article_ids") or []
    if not isinstance(raw_ids, list):
        return jsonify({"error": "ids must be a list"}), 400
    user_id = _get_current_user_id()
    conn = _get_db_connection()
    deleted = amdb.delete_articles_by_ids(conn, user_id, raw_ids)
    conn.close()
    return jsonify({"status": "success", "deleted": deleted}), 200


@article_master_bp.route("/delete-all", methods=["POST"])
@require_jwt_auth
def delete_all_user_articles():
    data = request.get_json(silent=True) or {}
    category = (data.get("category") or request.args.get("category") or "All").strip() or "All"
    user_id = _get_current_user_id()
    conn = _get_db_connection()
    deleted = amdb.delete_all_articles(conn, user_id, category=category)
    conn.close()
    return jsonify({
        "status": "success",
        "deleted": deleted,
        "category": category,
    }), 200


@article_master_bp.route("/confirm-new-category", methods=["POST"])
@require_jwt_auth
def confirm_new_category():
    data = request.get_json() or {}
    user_id = _get_current_user_id()
    workspace_id = get_workspace_id()
    conn = _get_db_connection()

    category_name = data.get("category_name")
    key_fields = data.get("key_fields", DEFAULT_KEY_FIELDS)

    if not category_name:
        conn.close()
        return jsonify({"error": "category_name is required"}), 400

    category = amdb.create_category(
        conn, user_id, category_name, key_fields,
        is_confirmed=True, workspace_id=workspace_id,
    )
    conn.close()
    return jsonify({"status": "success", "category": category}), 200


@article_master_bp.route("/list", methods=["GET"])
@require_jwt_auth
def list_articles():
    user_id = _get_current_user_id()
    conn = _get_db_connection()
    amdb.ensure_default_brand_aliases(conn, user_id)
    amdb.ensure_default_categories(conn, user_id, workspace_id=get_workspace_id())
    amdb.repair_product_types_from_size(conn, user_id)
    category = request.args.get("category")

    if category and category != "All":
        articles = amdb.get_articles_by_category(conn, user_id, category)
    else:
        articles = amdb.get_all_articles(conn, user_id)

    conn.close()
    return jsonify({"articles": articles, "count": len(articles)}), 200


@article_master_bp.route("/brand-aliases", methods=["GET"])
@require_jwt_auth
def list_brand_aliases():
    user_id = _get_current_user_id()
    conn = _get_db_connection()
    amdb.ensure_default_brand_aliases(conn, user_id)
    aliases = amdb.list_brand_aliases(conn, user_id)
    conn.close()
    return jsonify({"aliases": aliases}), 200


@article_master_bp.route("/brand-aliases", methods=["POST"])
@require_jwt_auth
def create_brand_alias():
    data = request.get_json() or {}
    alias = data.get("alias")
    canonical_brand = data.get("canonical_brand")
    user_id = _get_current_user_id()
    conn = _get_db_connection()
    try:
        amdb.upsert_brand_alias(conn, user_id, alias, canonical_brand)
    except ValueError as e:
        conn.close()
        return jsonify({"error": str(e)}), 400
    conn.close()
    return jsonify({"status": "success"}), 200


@article_master_bp.route("/<int:article_id>/edit", methods=["PATCH"])
@require_jwt_auth
def edit_article(article_id):
    data = request.get_json() or {}
    field = data.get("field")
    new_value = data.get("value")

    if not field or new_value is None:
        return jsonify({"error": "field and value are required"}), 400

    user_id = _get_current_user_id()
    conn = _get_db_connection()

    try:
        updated = amdb.manual_edit_article(
            conn, user_id, article_id, field, new_value,
            changed_by=_get_changed_by(),
        )
    except ValueError as e:
        conn.close()
        return jsonify({"error": str(e)}), 400

    conn.close()
    return jsonify({"status": "success", "article": updated}), 200


@article_master_bp.route("/<int:article_id>/edit-full", methods=["PATCH"])
@require_jwt_auth
def edit_article_full(article_id):
    data = request.get_json() or {}
    updates = data.get("updates") or {}
    if not isinstance(updates, dict) or not updates:
        return jsonify({"error": "updates object with at least one field is required"}), 400

    user_id = _get_current_user_id()
    conn = _get_db_connection()
    try:
        cols = ", ".join(amdb.ARTICLE_MASTER_COLUMNS)
        row = conn.execute(
            f"SELECT {cols} FROM article_master WHERE id = ? AND user_id = ?",
            (article_id, user_id),
        ).fetchone()
        if row is None:
            return jsonify({"error": "Article not found"}), 404
        article = amdb._row_to_article_dict(row)
        categories = amdb.get_all_categories(conn, user_id)
        key_fields_lookup = {c["category_name"]: c["key_fields"] for c in categories}
        key_fields = key_fields_lookup.get(article["category"], DEFAULT_KEY_FIELDS)

        updated, changed_fields = amdb.update_article_full(
            conn, user_id, article_id, updates, key_fields,
            changed_by=_get_changed_by(),
        )
    except ValueError as e:
        conn.close()
        return jsonify({"error": str(e)}), 400

    conn.close()
    return jsonify({
        "status": "success",
        "article": updated,
        "changed_fields": changed_fields,
    }), 200


@article_master_bp.route("/<int:article_id>/price-history", methods=["GET"])
@require_jwt_auth
def price_history(article_id):
    user_id = _get_current_user_id()
    conn = _get_db_connection()
    try:
        history = amdb.get_price_history(conn, article_id, user_id)
    except ValueError as e:
        conn.close()
        return jsonify({"error": str(e)}), 404
    conn.close()
    return jsonify({"history": history}), 200


@article_master_bp.route("/<int:article_id>/price-seasons", methods=["GET"])
@require_jwt_auth
def price_seasons(article_id):
    """Last 3 season price snapshots (missing seasons omitted)."""
    user_id = _get_current_user_id()
    conn = _get_db_connection()
    try:
        payload = amdb.get_season_prices_last_n(conn, article_id, user_id, limit=3)
    except ValueError as e:
        conn.close()
        return jsonify({"error": str(e)}), 404
    conn.close()
    return jsonify(_sanitize_for_json(payload)), 200


@article_master_bp.route("/download", methods=["GET"])
@require_jwt_auth
def download_articles():
    """
    Category-wise (or 'All') export in canonical Article Master column order.
    Size uses taught full names. Bed drops Color; Bath keeps Color (from Shade).
    When season price history exists, MRP/Ex-Mill/PTR expand into season columns
    with older seasons Excel-grouped (latest season stays visible).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    user_id = _get_current_user_id()
    conn = _get_db_connection()
    category = request.args.get("category", "All")

    if category != "All":
        articles = amdb.get_articles_by_category(conn, user_id, category)
    else:
        articles = amdb.get_all_articles(conn, user_id)

    if not articles:
        conn.close()
        return jsonify({"error": "No articles found for this selection"}), 404

    # Collect last-3 seasons across the selection (union, sorted).
    season_payloads = {}
    all_seasons = set()
    for a in articles:
        payload = amdb.get_season_prices_last_n(conn, a["id"], user_id, limit=3)
        season_payloads[a["id"]] = payload
        for s in payload.get("seasons") or []:
            all_seasons.add(s)
    conn.close()

    seasons_sorted = sorted(all_seasons, key=amparser.season_rank)
    use_season_cols = len(seasons_sorted) >= 1

    attr_cols = [
        c for c in CANONICAL_ARTICLE_MASTER_COLUMNS
        if c not in {"MRP", "Ex-Mill", "PTR"}
    ]
    # Split attr: before prices vs after (margins sit with latest or after seasons)
    before_price = []
    after_price = []
    hit_bale = False
    for c in attr_cols:
        if c == "Bale Pack Size":
            before_price.append(c)
            hit_bale = True
            continue
        if not hit_bale:
            before_price.append(c)
        else:
            after_price.append(c)

    money_fields = [("MRP", "mrp"), ("Ex-Mill", "ex_mill_price"), ("PTR", "ptr")]
    columns = list(before_price)
    if use_season_cols:
        for label, _key in money_fields:
            for s in seasons_sorted:
                columns.append(f"{label} ({s})")
    else:
        columns.extend(["MRP", "Ex-Mill", "PTR"])
    columns.extend(after_price)

    rows = []
    for a in articles:
        row = {col: _resolve_export_value(a, col) for col in before_price + after_price}
        if use_season_cols:
            payload = season_payloads.get(a["id"]) or {}
            by_field = payload.get("rows") or {}
            for label, key in money_fields:
                by_season = by_field.get(key) or {}
                for s in seasons_sorted:
                    val = by_season.get(s)
                    if val is None and amparser.normalize_season_tag(a.get("season_tag")) == s:
                        # fallback to article core if snapshot missing
                        val = a.get(key)
                    if label == "MRP" and val is not None:
                        try:
                            row[f"{label} ({s})"] = int(round(float(val)))
                        except (TypeError, ValueError):
                            row[f"{label} ({s})"] = val
                    elif val is not None:
                        try:
                            row[f"{label} ({s})"] = round(float(val), 2)
                        except (TypeError, ValueError):
                            row[f"{label} ({s})"] = val
                    else:
                        row[f"{label} ({s})"] = None
        else:
            for label, key in money_fields:
                val = a.get(key)
                if label == "MRP" and val is not None:
                    try:
                        row[label] = int(round(float(val)))
                    except (TypeError, ValueError):
                        row[label] = val
                else:
                    row[label] = val
        rows.append(row)

    wb = Workbook()
    ws = wb.active
    ws.title = "Article Master"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for c_idx, h in enumerate(columns, 1):
        cell = ws.cell(1, c_idx, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    for r_idx, row in enumerate(rows, 2):
        for c_idx, h in enumerate(columns, 1):
            val = row.get(h)
            cell = ws.cell(r_idx, c_idx, val)
            if h.startswith("MRP") and isinstance(val, int):
                cell.number_format = "0"
            elif (h.startswith("Ex-Mill") or h.startswith("PTR") or h in {"Ex-Mill", "PTR"}) and isinstance(val, (int, float)):
                cell.number_format = "0.00"

    for col in ws.columns:
        letter = col[0].column_letter
        width = max(len(str(cell.value or "")) for cell in col[: min(80, len(col))])
        ws.column_dimensions[letter].width = min(max(width + 2, 10), 36)

    # Group older season columns; leave latest season visible (summary on right).
    if use_season_cols and len(seasons_sorted) >= 2:
        latest = seasons_sorted[-1]
        ws.sheet_properties.outlinePr.summaryRight = True
        ws.sheet_properties.outlinePr.applyStyles = True
        for label, _key in money_fields:
            old_idxs = [
                i for i, h in enumerate(columns, 1)
                if h.startswith(f"{label} (") and not h.endswith(f"({latest})")
            ]
            if old_idxs:
                start = get_column_letter(min(old_idxs))
                end = get_column_letter(max(old_idxs))
                ws.column_dimensions.group(start, end, outline_level=1, hidden=False)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Article_Master_{category.replace(' ', '_')}.xlsx"
    return send_file(
        output, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
