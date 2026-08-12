"""TOB AW-26 Article Master preview — bedsheet-style season columns."""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SRC = Path(
    r"G:\My Drive\2026-2027\Oder Management\AW26 order\TOB"
    r"\AW-26 TOB Revised Booking Sheet_23.06.2026.xlsx"
)
OUT = Path(r"E:\test files\Article_Master_TOB_AW26_Preview.xlsx")
OUT_REPO = Path(r"E:\centralized-db-system\Output\Article_Master_TOB_AW26_Preview.xlsx")

SEASON = "AW-26"
MONEY = ["MRP", "Ex-Mill", "PTR"]
PCT = ["AWD Mark up on Exmill", "Proposed Customer Discount", "Retailer Margin"]

ATTR_COLS = [
    "Category",
    "Product",
    "Brand",
    "Size",
    "TC",
    "Units",
    "BS Size",
    "Pillow Size",
    "Color",
    "Pillow Stitching Style",
    "Print Style",
    "Blend",
    "Packing",
    "Bale Pack Size",
]

SECTION_RE = re.compile(
    r"collection|awd name|location|^product$",
    re.IGNORECASE,
)
PLY_MAP = {
    "SNL-PLY": "Single Ply",
    "SNL PLY": "Single Ply",
    "1-PLY": "1 Ply",
    "1 PLY": "1 Ply",
    "2-PLY": "2 Ply",
    "2 PLY": "2 Ply",
    "3-PLY": "3 Ply",
    "3 PLY": "3 Ply",
}


def blank(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and not v.strip():
        return True
    if isinstance(v, str) and v.strip() in {"-", "—", "–", "NA", "N/A", "null", "None"}:
        return True
    return False


def _norm_ws(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())


def normalize_product(product) -> str | None:
    if blank(product):
        return None
    text = _norm_ws(product)
    if re.sub(r"[^a-z0-9]+", " ", text.lower()).strip() == "bed in a bga":
        return "Bed in a Bag"
    return text


def normalize_quality(quality) -> str | None:
    if blank(quality):
        return None
    text = _norm_ws(quality)
    text = re.sub(r"polyster", "Polyester", text, flags=re.IGNORECASE)
    text = re.sub(r"\bpolyester\b", "Polyester", text, flags=re.IGNORECASE)
    return text


def normalize_ply(ply) -> str | None:
    if blank(ply):
        return None
    raw = _norm_ws(ply)
    mapped = PLY_MAP.get(raw.upper())
    if mapped:
        return mapped
    # soft: "3-ply" already covered via upper; keep readable
    m = re.match(r"^(\d+)\s*-?\s*ply$", raw, re.IGNORECASE)
    if m:
        return f"{m.group(1)} Ply"
    if re.match(r"^snl", raw, re.IGNORECASE):
        return "Single Ply"
    return raw


def normalize_weight(weight):
    if blank(weight):
        return None
    if isinstance(weight, (int, float)):
        if float(weight).is_integer():
            return str(int(weight))
        return str(weight)
    return _norm_ws(weight)


def build_blend(quality, ply, weight) -> str | None:
    q = normalize_quality(quality)
    p = normalize_ply(ply)
    w = normalize_weight(weight)
    head_parts = [x for x in (q, p) if x]
    if not head_parts and not w:
        return None
    head = " ".join(head_parts) if head_parts else None
    if head and w:
        return f"{head}, {w}"
    return head or w


def fmt_money(v):
    if blank(v):
        return None
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return v


def fmt_mrp(v):
    if blank(v):
        return None
    try:
        f = float(v)
        return int(f) if f.is_integer() else round(f, 2)
    except (TypeError, ValueError):
        return v


def fmt_pct(v):
    if blank(v):
        return None
    try:
        f = float(v)
        # sheet stores 0.4 meaning 40%
        if 0 < abs(f) <= 1:
            f = f * 100
        return round(f, 2)
    except (TypeError, ValueError):
        return v


def is_section_row(product, brand) -> bool:
    if blank(brand) and not blank(product):
        return True
    if blank(product):
        return False
    return bool(SECTION_RE.search(_norm_ws(product)))


def parse_tob(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header_r = None
    for r in range(1, min(20, (ws.max_row or 0) + 1)):
        vals = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, 20)]
        if "brand" in vals and "product" in vals and "size" in vals:
            header_r = r
            break
    if not header_r:
        raise RuntimeError("TOB header row not found")

    cmap = {}
    for c in range(1, (ws.max_column or 0) + 1):
        h = ws.cell(header_r, c).value
        if h is None:
            continue
        key = re.sub(r"\s+", " ", str(h).strip().lower())
        cmap[key] = c

    def col(*names):
        for n in names:
            if n in cmap:
                return cmap[n]
        for n in names:
            for k, idx in cmap.items():
                if n in k:
                    return idx
        return None

    c_prod = col("product")
    c_brand = col("brand")
    c_size = col("size")
    c_quality = col("quality")
    c_ply = col("ply")
    c_print = col("print/dyed/weave", "print dyed weave")
    c_weight = col("weight in gram", "weight")
    c_bale = col("bale pack size", "bale pack sizes", "pack sizes")
    c_mrp = col("mrp")
    c_ptr = col("ptr")
    c_ex = col("ex-mill", "ex mill")
    c_ret = col("retail mark down", "retailer margin")
    c_awd = col("awd md", "awd mu", "awd mark up")

    rows = []
    for r in range(header_r + 1, (ws.max_row or 0) + 1):
        product = ws.cell(r, c_prod).value if c_prod else None
        brand = ws.cell(r, c_brand).value if c_brand else None
        size = ws.cell(r, c_size).value if c_size else None
        if is_section_row(product, brand):
            continue
        if blank(brand) and blank(size):
            continue

        product_n = normalize_product(product)
        brand_n = None if blank(brand) else _norm_ws(brand)
        size_n = None if blank(size) else _norm_ws(size)
        print_style = None
        if c_print and not blank(ws.cell(r, c_print).value):
            print_style = _norm_ws(ws.cell(r, c_print).value)
        blend = build_blend(
            ws.cell(r, c_quality).value if c_quality else None,
            ws.cell(r, c_ply).value if c_ply else None,
            ws.cell(r, c_weight).value if c_weight else None,
        )
        bale = ws.cell(r, c_bale).value if c_bale else None
        if not blank(bale) and isinstance(bale, float) and bale.is_integer():
            bale = int(bale)

        rows.append(
            {
                "_src_row": r,
                "Category": "TOB",
                "Product": product_n,
                "Brand": brand_n,
                "Size": size_n,
                "TC": None,
                "Units": None,
                "BS Size": None,
                "Pillow Size": None,
                "Color": None,
                "Pillow Stitching Style": None,
                "Print Style": print_style,
                "Blend": blend,
                "Packing": None,
                "Bale Pack Size": bale,
                "MRP": ws.cell(r, c_mrp).value if c_mrp else None,
                "Ex-Mill": ws.cell(r, c_ex).value if c_ex else None,
                "PTR": ws.cell(r, c_ptr).value if c_ptr else None,
                "AWD Mark up on Exmill": ws.cell(r, c_awd).value if c_awd else None,
                "Retailer Margin": ws.cell(r, c_ret).value if c_ret else None,
                "Proposed Customer Discount": None,
            }
        )
    return rows


def merge_rows(rows: list[dict]) -> list[dict]:
    store: dict[str, dict] = {}
    for r in sorted(rows, key=lambda x: x["_src_row"]):
        b = "" if blank(r.get("Brand")) else str(r["Brand"]).strip().upper()
        s = "" if blank(r.get("Size")) else str(r["Size"]).strip().upper()
        c = "" if blank(r.get("Color")) else str(r["Color"]).strip().upper()
        p = "" if blank(r.get("Product")) else str(r["Product"]).strip().upper()
        key = f"{b}|{s}|{c}|{p}"
        if key not in store:
            store[key] = {
                "attrs": {col: None for col in ATTR_COLS},
                "money": {f: None for f in MONEY},
                "pct": {col: None for col in PCT},
            }
        cur = store[key]
        for col in ATTR_COLS:
            new = r.get(col)
            if blank(new):
                continue
            cur["attrs"][col] = new.strip() if isinstance(new, str) else new
        for f in MONEY:
            if not blank(r.get(f)):
                cur["money"][f] = r.get(f)
        for col in PCT:
            if not blank(r.get(col)):
                cur["pct"][col] = r.get(col)

    out = []
    for cur in store.values():
        row = dict(cur["attrs"])
        for f in MONEY:
            val = cur["money"][f]
            col = f"{f} ({SEASON})"
            row[col] = fmt_mrp(val) if f == "MRP" else fmt_money(val)
        for col in PCT:
            row[col] = fmt_pct(cur["pct"][col])
        out.append(row)

    out.sort(
        key=lambda r: (
            str(r.get("Product") or "").upper(),
            str(r.get("Brand") or "").upper(),
            str(r.get("Size") or "").upper(),
        )
    )
    return out


def write_excel(rows: list[dict], path: Path) -> None:
    columns = (
        ATTR_COLS
        + [f"{f} ({SEASON})" for f in MONEY]
        + PCT
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Article Master TOB"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for i, h in enumerate(columns, 1):
        cell = ws.cell(1, i, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    for r_i, row in enumerate(rows, 2):
        for c_i, h in enumerate(columns, 1):
            ws.cell(r_i, c_i, row.get(h))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"
    for i, h in enumerate(columns, 1):
        width = 14
        if h in {"Product", "Brand", "Blend"}:
            width = 28
        if h.startswith("MRP") or h.startswith("Ex-Mill") or h.startswith("PTR"):
            width = 14
        ws.column_dimensions[get_column_letter(i)].width = width

    notes = wb.create_sheet("Locks", 1)
    notes["A1"] = "TOB AW-26 teaching locks (preview only — not yet in live parser)"
    for i, line in enumerate(
        [
            "Category: TOB (incl. Bed in a Bag / BIAB)",
            "Identity: Brand + Size + Product + Color (Color empty — option counts ignored)",
            "Size: as-is",
            "Quality + Ply + Weight → Blend (e.g. 100% Polyester Single Ply, 950 / …, 1kg)",
            "Polyster→Polyester; SNL-PLY→Single Ply; 1-Ply→1 Ply; …",
            "Print/Dyed/Weave → Print Style as-is",
            "Brand: as-is (trim)",
            "Product: trim; Bed in a Bga → Bed in a Bag",
            "Retail Mark down → Retailer Margin; AWD MD → AWD Markup",
            "Bale Pack Size → Pack Sizes (column Bale Pack Size)",
            "Drop: MOQ, Booking Qnty, Dyed/Printed Option, Print Colorways, Delivery Months",
            "Prices: MRP / Ex-Mill / PTR (AW-26); margins latest-style",
        ],
        3,
    ):
        notes[f"A{i}"] = line
    notes.column_dimensions["A"].width = 100

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main():
    raw = parse_tob(SRC)
    merged = merge_rows(raw)
    write_excel(merged, OUT)
    write_excel(merged, OUT_REPO)
    print(f"raw={len(raw)} merged={len(merged)}")
    print(f"wrote {OUT}")
    print(f"wrote {OUT_REPO}")
    # sample blends
    for row in merged[:5]:
        print(
            row.get("Product"),
            "|",
            row.get("Brand"),
            "|",
            row.get("Size"),
            "|",
            row.get("Blend"),
            "|",
            row.get("Print Style"),
            "|",
            row.get(f"MRP ({SEASON})"),
        )


if __name__ == "__main__":
    main()
