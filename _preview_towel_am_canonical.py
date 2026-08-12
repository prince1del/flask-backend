"""Full Article Master preview: Bed/TOB from DB + AW-25 + AW-26 Towel merged."""
from __future__ import annotations

import json
import re
import sqlite3
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import article_master_db as amdb
import article_master_parser as amparser

SRC_AW25 = Path(r"E:\test files\AW-25 Towel Booking Sheet.xlsx")
SRC_AW26 = Path(
    r"G:\My Drive\2026-2027\Oder Management\AW26 order\Towel"
    r"\AW-26 Towel Phase-2 Booking Sheet.xlsx"
)
OUT = Path(r"E:\test files\Article_Master_All_with_Towel_AW25_AW26_Preview.xlsx")
OUT_REPO = Path(
    r"E:\centralized-db-system\Output\Article_Master_All_with_Towel_AW25_AW26_Preview.xlsx"
)
DB_PATH = Path(r"E:\centralized-db-system\centralized_db.sqlite3")
USER_ID = 2

CANONICAL_ARTICLE_MASTER_COLUMNS = [
    "Category",
    "Product",
    "Brand",
    "Size",
    "TC",
    "Units",
    "BS Size",
    "Pillow Size",
    "Pillow Stitching Style",
    "Print Style",
    "Blend",
    "Packing",
    "Bale Pack Size",
    "MRP",
    "AWD Mark up on Exmill",
    "Ex-Mill",
    "Proposed Customer Discount",
    "Retailer Margin",
    "PTR",
]

EXPORT_COLUMN_ALIASES = {
    "bs size": ["bs size", "bedset size (cms)", "bedset size", "bed set size"],
    "pillow size": ["pillow size", "pillow size (cms)"],
    "bale pack size": ["bale pack size", "bale size", "bale pack sizes", "pack sizes", "pack size"],
    "ex-mill": ["ex-mill", "exmill price", "ex mill", "ex-mill per pcs", "ex mill per pcs"],
    "awd mark up on exmill": [
        "awd mark up on exmill", "awd markup on exmill", "awd md", "awd mu",
        "distributor mark up", "mark up on exmill",
    ],
    "retailer margin": ["retailer margin", "retail mark down", "retailer md", "retailer markdown"],
    "print style": ["print style", "print/dyed/weave", "print dyed weave"],
    "proposed customer discount": [
        "proposed customer discount", "perceived", "perceive", "perceived margin",
    ],
}

COLUMNS = []
for col in CANONICAL_ARTICLE_MASTER_COLUMNS:
    COLUMNS.append(col)
    if col == "Pillow Size":
        COLUMNS.append("Color")

SIZE_CODE_TO_DISPLAY = {
    "40X60": "Hand Towel",
    "75X150": "Bath Towel",
    "60X120": "Ladies Towel",
    "90X180": "Pool Towel",
    "R4": "Towel Set",
    "30X30": "Face Towel",
    "72X144": "Bath Towel",
    "50X70": "Bath Mat",
    "50X100": "Gym Towel",
    "91X100": "91x100",
    "L": "Large",
    "XL": "Extra Large",
    "XXL": "Double Extra Large",
}

SIZE_SORT_RANK = {
    "Face Towel": 10,
    "Face Towel Set of 3": 11,
    "Hand Towel": 20,
    "Hand Towel Set of 2": 21,
    "Ladies Towel": 30,
    "Bath Towel": 40,
    "Bath Mat": 45,
    "Pool Towel": 50,
    "Towel Set": 60,
    "Gym Towel": 70,
    "91x100": 75,
    "Large": 80,
    "Extra Large": 81,
    "Double Extra Large": 82,
}

PVC_RE = re.compile(r"\(?\s*PVC\s*bag\s*Pkg\.?\s*\)?", re.IGNORECASE)
SEASON_RANK = {"AW-25": 1, "AW-26": 2}


def _resolve_export_value(article, column_name):
    if amparser.is_excluded_extra_column(column_name):
        return None
    lower = str(column_name or "").strip().lower()
    if lower == "category":
        return article.get("category")
    if lower == "product":
        return article.get("product_type")
    if lower == "size":
        return amparser.size_display_name(article.get("size"))
    core_field = amparser.resolve_core_field_for_name(column_name)
    if core_field:
        return article.get(core_field)
    extra = amparser.strip_excluded_extra_attributes(article.get("extra_attributes") or {})
    aliases = EXPORT_COLUMN_ALIASES.get(lower, [lower])
    alias_list = [column_name.strip()] + [a for a in aliases if a != lower]
    extra_lower = {str(k).strip().lower(): v for k, v in extra.items()}
    for alias in alias_list:
        if alias in extra:
            val = extra[alias]
            break
        hit = extra_lower.get(str(alias).strip().lower())
        if hit is not None and hit != "":
            val = hit
            break
    else:
        return None
    if lower in {
        "awd mark up on exmill",
        "retailer margin",
        "perceived",
        "proposed customer discount",
    }:
        return amparser.format_percent_display(val)
    return val


def _norm_ws(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())


def normalize_brand(brand) -> str:
    text = _norm_ws(brand)
    upper = text.upper()
    if upper == "BD WHITE":
        return "BD White"
    if upper == "GYM TOWEL":
        return "Gym Towel"
    if upper == "HUK A BUK":
        return "Huk A Buk"
    if upper == "LEPORD":
        return "Leopard"
    if upper == "ECO STRIPE":
        return "Eco Stripe"
    return text


def normalize_product(product) -> str:
    text = _norm_ws(product)
    key = re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()
    if key.startswith("bathmat"):
        return "Bathmat Antiskid"
    return text


def normalize_size(size) -> str:
    raw = _norm_ws(size)
    if not raw:
        return ""
    key = re.sub(r"\s+", "", raw).upper()
    if (
        re.fullmatch(r"40X60\(2PC\)|40X60\(SETOF2\)|40X60SETOF2", key)
        or ("SET OF 2" in raw.upper() and "40" in key and "60" in key)
        or re.search(r"40X60.*2PC|40X60\(2", key)
    ):
        return "Hand Towel Set of 2"
    if (
        re.fullmatch(r"30X30\(3PC\)|30X30\(SETOF3\)", key)
        or ("SET OF 3" in raw.upper() and "30" in key)
        or re.search(r"30X30.*3PC|30X30\(3", key)
    ):
        return "Face Towel Set of 3"
    compact = re.sub(r"[^0-9A-Z]", "", key)
    for code, display in SIZE_CODE_TO_DISPLAY.items():
        if compact == re.sub(r"[^0-9A-Z]", "", code) or key == code:
            return display
    return raw


def normalize_color_and_packing(shade, description) -> tuple[str, str | None, bool]:
    text = _norm_ws(shade)
    desc = _norm_ws(description)
    blob = f"{text} {desc}"
    has_pvc = bool(PVC_RE.search(blob))
    has_pkg_l = bool(
        re.search(r"\(\s*L\s*\)", blob, re.IGNORECASE)
        or (re.search(r"\bPkg\b", blob, re.IGNORECASE) and not has_pvc)
    )
    packing = "PVC bag Pkg" if has_pvc else None
    text = PVC_RE.sub("", text)
    text = re.sub(r"\(\s*L\s*\)", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\(\s*Pkg\.?\s*\)", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip(" -")
    if not text:
        return "", packing, has_pkg_l and not has_pvc
    if re.fullmatch(r"jacquarad|jacquard", text, flags=re.IGNORECASE):
        return "Jacquard", packing, has_pkg_l and not has_pvc
    m2 = re.fullmatch(
        r"(?:Assorted|Asorted|Assortede|Asst\.?|Asst)\s*[- ]*0*(\d+)",
        text,
        flags=re.IGNORECASE,
    )
    if m2:
        return f"Assorted {int(m2.group(1)):02d}", packing, has_pkg_l and not has_pvc
    if text.upper() == "WHITE":
        return "White", packing, has_pkg_l and not has_pvc
    return text, packing, has_pkg_l and not has_pvc


def format_pct(val) -> str | None:
    if val is None or val == "":
        return None
    try:
        f = float(val)
    except (TypeError, ValueError):
        return str(val)
    if 0 <= f <= 1:
        return f"{(f * 100):g}%"
    return f"{f:g}%"


def round_money(val):
    if val is None or val == "":
        return None
    try:
        f = float(val)
    except (TypeError, ValueError):
        return val
    if abs(f - round(f)) < 1e-9:
        return int(round(f))
    return round(f, 2)


def detect_header_and_map(ws):
    """Return (header_row_idx_1based, col_map name->1based index)."""
    for r in range(1, min(15, (ws.max_row or 0) + 1)):
        vals = {}
        for c in range(1, min(25, (ws.max_column or 0) + 1)):
            v = ws.cell(r, c).value
            if v is None:
                continue
            key = _norm_ws(v).lower()
            vals[key] = c
        if "brand" in vals and "size" in vals and ("mrp" in vals or "product" in vals):
            return r, vals
    raise ValueError("Could not find towel header row")


def _col(cmap, *names):
    for n in names:
        if n in cmap:
            return cmap[n]
    return None


def parse_towel_sheet(path: Path, season: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_r, cmap = detect_header_and_map(ws)
    c_prod = _col(cmap, "product")
    c_brand = _col(cmap, "brand")
    c_shade = _col(cmap, "shade")
    c_desc = _col(cmap, "description")
    c_size = _col(cmap, "size")
    c_bale = _col(cmap, "pack sizes", "bale pack sizes", "bale pack size", "bale size")
    c_awd = _col(cmap, "awd mu", "awd md")
    c_ret = _col(cmap, "retailer md", "retailer margin")
    c_mrp = _col(cmap, "mrp")
    c_ptr = _col(cmap, "ptr")
    c_ex = _col(cmap, "ex-mill per pcs", "ex mill per pcs", "ex-mill", "ex mill")

    rows = []
    for r in range(header_r + 1, (ws.max_row or 0) + 1):
        product = ws.cell(r, c_prod).value if c_prod else None
        brand = ws.cell(r, c_brand).value if c_brand else None
        shade = ws.cell(r, c_shade).value if c_shade else None
        desc = ws.cell(r, c_desc).value if c_desc else None
        size = ws.cell(r, c_size).value if c_size else None
        bale = ws.cell(r, c_bale).value if c_bale else None
        awd = ws.cell(r, c_awd).value if c_awd else None
        retailer = ws.cell(r, c_ret).value if c_ret else None
        mrp = ws.cell(r, c_mrp).value if c_mrp else None
        ptr = ws.cell(r, c_ptr).value if c_ptr else None
        ex_mill = ws.cell(r, c_ex).value if c_ex else None

        if not any([product, brand, size, mrp]):
            continue

        color, packing, is_pkg_only = normalize_color_and_packing(shade, desc)
        rows.append(
            {
                "Category": "Bath",
                "Product": normalize_product(product),
                "Brand": normalize_brand(brand),
                "Size": normalize_size(size),
                "TC": None,
                "Units": None,
                "BS Size": None,
                "Pillow Size": None,
                "Pillow Stitching Style": None,
                "Color": color,
                "Print Style": None,
                "Blend": None,
                "Packing": packing,
                "Bale Pack Size": bale,
                "MRP": round_money(mrp),
                "AWD Mark up on Exmill": format_pct(awd),
                "Ex-Mill": round_money(ex_mill),
                "Proposed Customer Discount": None,
                "Retailer Margin": format_pct(retailer),
                "PTR": round_money(ptr),
                "_src_row": r,
                "_season": season,
                "_is_pkg_only": is_pkg_only,
                "_has_pvc": packing == "PVC bag Pkg",
                "_source": path.name,
            }
        )
    return rows


def dedupe_pkg_within_season(rows):
    by_key = {}
    for row in rows:
        key = (row["Brand"], row["Size"], row["Color"], row["Product"], row["_season"])
        by_key.setdefault(key, []).append(row)

    kept, dropped = [], []
    for group in by_key.values():
        if len(group) == 1:
            kept.append(group[0])
            continue
        pvc = [g for g in group if g["_has_pvc"]]
        non_pkg = [g for g in group if not g["_is_pkg_only"]]
        pool = pvc or non_pkg or group
        winner = max(pool, key=lambda g: g["_src_row"])
        kept.append(winner)
        for g in group:
            if g is not winner:
                dropped.append(g)
    return kept, dropped


def merge_seasons(rows_aw25, rows_aw26):
    """Same Brand+Size+Color+Product → keep latest season (AW-26)."""
    by_key = {}
    for row in rows_aw25 + rows_aw26:
        key = (row["Brand"], row["Size"], row["Color"], row["Product"])
        prev = by_key.get(key)
        if prev is None:
            by_key[key] = row
            continue
        if SEASON_RANK.get(row["_season"], 0) >= SEASON_RANK.get(prev["_season"], 0):
            # Prefer PVC packing if latest has it, else keep latest wholesale
            by_key[key] = row
        elif row["_has_pvc"] and not prev["_has_pvc"]:
            by_key[key] = row

    merged = list(by_key.values())

    def sort_key(row):
        return (
            str(row["Brand"] or "").upper(),
            SIZE_SORT_RANK.get(row["Size"], 999),
            str(row["Size"] or ""),
            str(row["Color"] or ""),
            str(row["Product"] or ""),
            row.get("_src_row", 0),
        )

    merged.sort(key=sort_key)
    return merged


def resolve_db_row(article: dict) -> dict:
    row = {}
    for col in COLUMNS:
        if col == "Color":
            extra = article.get("extra_attributes") or {}
            if isinstance(extra, str):
                try:
                    extra = json.loads(extra)
                except Exception:
                    extra = {}
            extra_l = {str(k).strip().lower(): v for k, v in (extra or {}).items()}
            row[col] = extra_l.get("color") or extra_l.get("shade")
            continue
        row[col] = _resolve_export_value(article, col)
    return row


def build_db_rows(user_id: int):
    conn = sqlite3.connect(DB_PATH)
    amdb.ensure_schema(conn)
    articles = amdb.get_all_articles(conn, user_id, active_only=True)
    conn.close()
    cat_rank = {"Bed": 0, "TOB": 1, "TOB Pillow": 2, "Bath": 3}
    articles = [a for a in articles if a.get("category") != "Bath"]
    articles.sort(
        key=lambda a: (
            cat_rank.get(a.get("category"), 9),
            str(a.get("brand") or "").upper(),
            amparser.bed_size_sort_rank(a.get("size")),
            str(a.get("size") or ""),
        )
    )
    return [resolve_db_row(a) for a in articles]


def write_workbook(all_rows, towel_rows, dropped_pkg, db_count, stats):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Article Master"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )

    # Add Season column at end for towel transparency
    out_cols = list(COLUMNS) + ["Season"]

    for c, name in enumerate(out_cols, 1):
        cell = ws.cell(1, c, name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin

    for r_i, row in enumerate(all_rows, 2):
        for c, name in enumerate(out_cols, 1):
            if name == "Season":
                val = row.get("_season") if row.get("Category") == "Bath" else None
            else:
                val = row.get(name)
            cell = ws.cell(r_i, c, val)
            cell.border = thin

    widths = {
        "Category": 10,
        "Product": 18,
        "Brand": 22,
        "Size": 26,
        "Color": 14,
        "Packing": 14,
        "Bale Pack Size": 14,
        "MRP": 10,
        "AWD Mark up on Exmill": 14,
        "Ex-Mill": 12,
        "Retailer Margin": 12,
        "PTR": 10,
        "Season": 10,
        "Pillow Stitching Style": 16,
    }
    for c, name in enumerate(out_cols, 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(name, 12)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(out_cols))}{len(all_rows) + 1}"
    ws.freeze_panes = "A2"

    bath = wb.create_sheet("Bath Towel merged")
    for c, name in enumerate(out_cols, 1):
        cell = bath.cell(1, c, name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
    for r_i, row in enumerate(towel_rows, 2):
        for c, name in enumerate(out_cols, 1):
            val = row.get("_season") if name == "Season" else row.get(name)
            cell = bath.cell(r_i, c, val)
            cell.border = thin
    bath.auto_filter.ref = f"A1:{get_column_letter(len(out_cols))}{len(towel_rows) + 1}"
    bath.freeze_panes = "A2"
    for c, name in enumerate(out_cols, 1):
        bath.column_dimensions[get_column_letter(c)].width = widths.get(name, 12)

    notes = wb.create_sheet("Notes")
    notes["A1"] = "Full AM: Bed/TOB (DB) + AW-25 + AW-26 Towel merge — NOT uploaded"
    notes["A1"].font = Font(bold=True)
    bullets = [
        f"DB Bed/TOB: {db_count}",
        f"AW-25 rows in: {stats['aw25_in']} → after Pkg dedupe: {stats['aw25_kept']}",
        f"AW-26 rows in: {stats['aw26_in']} → after Pkg dedupe: {stats['aw26_kept']}",
        f"Bath merged (latest season wins): {len(towel_rows)}",
        f"Total Article Master: {len(all_rows)}",
        f"Pkg/(L) dropped: {len(dropped_pkg)}",
        "Luxury Living Hand Towel comes from AW-26 (not in AW-25 source)",
        f"Sources: {SRC_AW25.name} + {SRC_AW26.name}",
    ]
    for i, line in enumerate(bullets, 3):
        notes[f"A{i}"] = f"• {line}"
    notes["A12"] = "Luxury Living rows in merge:"
    lux = [r for r in towel_rows if str(r.get("Brand") or "") == "Luxury Living"]
    for i, r in enumerate(lux, 13):
        notes[f"A{i}"] = (
            f"  {r['Size']} | {r['Color']} | MRP={r['MRP']} | Season={r['_season']}"
        )
    notes.column_dimensions["A"].width = 110

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT_REPO.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    wb.save(OUT_REPO)
    print(f"Wrote {OUT}")
    print(f"DB={db_count} Bath={len(towel_rows)} Total={len(all_rows)}")
    print("Luxury Living:")
    for r in lux:
        print(f"  {r['Size']} | {r['Color']} | MRP={r['MRP']} | {r['_season']}")


def main():
    db_rows = build_db_rows(USER_ID)

    aw25_raw = parse_towel_sheet(SRC_AW25, "AW-25")
    aw26_raw = parse_towel_sheet(SRC_AW26, "AW-26")
    aw25_kept, drop25 = dedupe_pkg_within_season(aw25_raw)
    aw26_kept, drop26 = dedupe_pkg_within_season(aw26_raw)
    dropped = drop25 + drop26
    towel_rows = merge_seasons(aw25_kept, aw26_kept)

    stats = {
        "aw25_in": len(aw25_raw),
        "aw25_kept": len(aw25_kept),
        "aw26_in": len(aw26_raw),
        "aw26_kept": len(aw26_kept),
    }

    # For Article Master sheet, attach season on bath rows then strip helpers later
    all_rows = db_rows + towel_rows
    write_workbook(all_rows, towel_rows, dropped, len(db_rows), stats)


if __name__ == "__main__":
    main()
