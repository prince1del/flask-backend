"""Review Choice Corner / DCA Order / kag.xlsx under locked FO rules."""
from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import subprocess

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

BASE = Path(r"g:\My Drive\2026-2027\Oder Management\AW26 order\Bedsheet")
OUT = Path(r"E:\test files\FO_Review_Choice_DCA_KAG_v3.xlsx")

FILES = [
    ("Choice Corner.xlsx", "Choice Corner"),
    ("DCA Order.xlsx", "DCA"),
    ("kag.xlsx", "KAG"),
]


def fnum(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def find_header_map(ws):
    """Return {normalized_header: col} from row 2 (or nearby)."""
    headers = {}
    for r in range(1, 5):
        for c in range(1, (ws.max_column or 1) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip():
                key = v.strip().lower()
                headers.setdefault(key, (r, c))
    return headers


def _col_positive_count(ws, col):
    n = 0
    for r in range(3, (ws.max_row or 3) + 1):
        if not ws.cell(r, 2).value:
            continue
        v = fnum(ws.cell(r, col).value)
        if v is not None and v > 0:
            n += 1
    return n


def _header_at(ws, col, row=2):
    v = ws.cell(row, col).value
    return str(v).strip().lower() if isinstance(v, str) else ""


def detect_qty_col(ws, headers):
    """
    Prefer header Qnty/Qty (NOT Qnty Per Color / Qnty pre Design / Value).
    If that header column is empty but the next column has numbers, use shifted col.
    Never treat a Value column as Qty.
    """
    skip = {
        "qnty per color",
        "qnty pre design",
        "qnty per design",
        "value",
        "awd value",
    }
    qty_header_cols = []
    for c in range(1, (ws.max_column or 1) + 1):
        h = _header_at(ws, c)
        if not h or h in skip or "value" in h:
            continue
        if h in {"qnty", "qty", "quantity", "additional quantity"}:
            qty_header_cols.append(c)

    if not qty_header_cols:
        # fallback: any header that is exactly qnty-like from headers map
        for h, (_r, c) in headers.items():
            if h in {"qnty", "qty", "quantity"} and "value" not in h:
                qty_header_cols.append(c)

    if not qty_header_cols:
        return None, "not_found", None

    header_col = qty_header_cols[0]
    header_filled = _col_positive_count(ws, header_col)

    # Do not shift into a Value-header column
    right_col = header_col + 1
    right_header = _header_at(ws, right_col)
    right_is_value = ("value" in right_header) if right_header else False
    right_filled = 0 if right_is_value else _col_positive_count(ws, right_col)

    if header_filled > 0:
        return header_col, "header", header_col
    if right_filled > 0:
        return right_col, "shifted_right_of_Qnty_header", header_col
    return header_col, "header_empty", header_col


def detect_value_col(ws, qty_col):
    # header Value
    for c in range(1, (ws.max_column or 1) + 1):
        h = _header_at(ws, c)
        if h and "value" in h:
            return c, "header"
    # else column right of qty if numeric present (and not another known header)
    if qty_col:
        c = qty_col + 1
        h = _header_at(ws, c)
        if h and h not in {"",} and "qnty" in h:
            return None, "missing"
        filled = _col_positive_count(ws, c)
        if filled:
            return c, "right_of_qty"
    return None, "missing"


def detect_bales_col(headers):
    for h, (r, c) in headers.items():
        if "bale" in h and "size" not in h and "pack" not in h:
            return c
    return None


def parse_file(path: Path, distributor_hint: str):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = find_header_map(ws)
    qty_col, qty_how, qty_header_col = detect_qty_col(ws, headers)
    value_col, value_how = detect_value_col(ws, qty_col)
    bales_col = detect_bales_col(headers)

    # find bale size / exmill cols
    bs_col = headers.get("bale size", (None, 11))[1]
    ex_col = None
    # Prefer exact ExMill Price — never "AWD Mark up on Exmill"
    for prefer in ("exmill price", "ex-mill price", "ex mill price", "exmill"):
        if prefer in headers:
            ex_col = headers[prefer][1]
            break
    if ex_col is None:
        for h, (_r, c) in headers.items():
            compact = h.replace(" ", "").replace("-", "")
            if compact in {"exmillprice", "exmill"} or (
                compact.startswith("exmill") and "markup" not in compact and "mark" not in compact
            ):
                ex_col = c
                break
    if ex_col is None:
        ex_col = 24

    lines = []
    for r in range(3, (ws.max_row or 3) + 1):
        brand = ws.cell(r, 2).value
        if not brand:
            continue
        bs = fnum(ws.cell(r, bs_col).value) if bs_col else None
        ex = fnum(ws.cell(r, ex_col).value) if ex_col else None
        q = fnum(ws.cell(r, qty_col).value) if qty_col else None
        b = fnum(ws.cell(r, bales_col).value) if bales_col else None
        sheet_v = fnum(ws.cell(r, value_col).value) if value_col else None

        if (q is None or q <= 0) and (b is None or b <= 0):
            continue

        notes = []
        bale_status = ""
        # Rule 1 / 2
        if q is not None and q > 0 and b is not None:
            expected_b = q / bs if bs else None
            if expected_b is not None and abs(b - expected_b) >= 0.01:
                bale_status = "MISMATCH"
                notes.append(f"sheet bales={b}, expected={round(expected_b, 2)}")
            else:
                bale_status = "OK"
        elif b is not None and (q is None or q <= 0):
            q = b * bs if bs else None
            bale_status = "ONLY_BALES"
            notes.append("qty auto = bales * bale_size")
        elif q is not None and q > 0 and b is None:
            bale_status = "NO_BALES_COL"

        value = (q * ex) if (q is not None and ex is not None) else None
        if sheet_v is not None and value is not None and abs(sheet_v - value) >= 1:
            value_status = "SHEET_VALUE_DIFFERS"
            notes.append(f"sheet value={round(sheet_v, 2)}, qty*ex={round(value, 2)}")
        elif sheet_v is not None:
            value_status = "SHEET_MATCHES_QTY"
        else:
            value_status = "FROM_QTY"

        # informational: clean bale multiple?
        clean = None
        if q is not None and bs:
            clean = abs(q / bs - round(q / bs)) < 0.01

        lines.append(
            {
                "brand": str(brand).strip(),
                "tc": ws.cell(r, 3).value,
                "size": str(ws.cell(r, 4).value or ""),
                "product": ws.cell(r, 9).value,
                "bs": bs,
                "ex": ex,
                "qty": q,
                "bales": b,
                "value": value,
                "sheet_value": sheet_v,
                "bale_status": bale_status,
                "value_status": value_status,
                "clean_bale_multiple": clean,
                "notes": "; ".join(notes),
            }
        )

    meta = {
        "file": path.name,
        "distributor_hint": distributor_hint,
        "sheet": ws.title,
        "qty_col": qty_col,
        "qty_how": qty_how,
        "qty_header_col": qty_header_col,
        "value_col": value_col,
        "value_how": value_how,
        "bales_col": bales_col,
        "bs_col": bs_col,
        "ex_col": ex_col,
        "noise_note": "Ignore Qnty Per Color / Qnty pre Design — not order qty",
    }
    return meta, lines


def main():
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill("solid", "0F766E")
    bad = PatternFill("solid", "FEE2E2")
    warn = PatternFill("solid", "FEF3C7")
    ok = PatternFill("solid", "DCFCE7")
    nf = Font(bold=True)
    tfill = PatternFill("solid", "CCFBF1")

    wb = Workbook()
    overview = wb.active
    overview.title = "Overview"
    overview.append(
        [
            "File",
            "Distributor (from filename)",
            "Sheet fields used",
            "Qty detection",
            "Lines with qty",
            "Total Qty",
            "Total Value (=Qty*ExMill)",
            "Bale mismatches",
            "Sheet value differs",
            "Notes",
        ]
    )
    for cell in overview[1]:
        cell.font = hf
        cell.fill = hfill
        cell.border = thin

    all_results = []
    for fname, hint in FILES:
        path = BASE / fname
        meta, lines = parse_file(path, hint)
        all_results.append((meta, lines))

        q = sum(x["qty"] or 0 for x in lines)
        v = sum(x["value"] or 0 for x in lines)
        mm = sum(1 for x in lines if x["bale_status"] == "MISMATCH")
        vd = sum(1 for x in lines if x["value_status"] == "SHEET_VALUE_DIFFERS")

        fields = []
        if meta["qty_col"]:
            fields.append(f"Qty col {get_column_letter(meta['qty_col'])}")
        if meta["bales_col"]:
            fields.append(f"Bales col {get_column_letter(meta['bales_col'])}")
        else:
            fields.append("No bales col")
        if meta["value_col"]:
            fields.append(f"Sheet value col {get_column_letter(meta['value_col'])} ({meta['value_how']})")
        else:
            fields.append("No sheet value col")

        det = f"col {get_column_letter(meta['qty_col'])} ({meta['qty_how']})"
        if meta["qty_how"] == "shifted_right_of_Qnty_header":
            det += f" — Qnty header was {get_column_letter(meta['qty_header_col'])} but empty"

        notes = [meta["noise_note"]]
        if meta["qty_how"] == "shifted_right_of_Qnty_header":
            notes.append("WARNING: distributor filled qty one column right of Qnty header")
        if not meta["bales_col"]:
            notes.append("Rule1 bale check N/A (no bales column)")

        overview.append(
            [
                meta["file"],
                meta["distributor_hint"],
                "; ".join(fields),
                det,
                len(lines),
                round(q, 2),
                round(v, 2),
                mm,
                vd,
                " | ".join(notes),
            ]
        )
        if meta["qty_how"] == "shifted_right_of_Qnty_header":
            for c in range(1, 11):
                overview.cell(overview.max_row, c).fill = warn

    for i, w in enumerate([28, 18, 45, 45, 12, 12, 18, 14, 16, 70], 1):
        overview.column_dimensions[get_column_letter(i)].width = w

    # Rules sheet
    rs = wb.create_sheet("Rules Applied", 1)
    rs.append(["#", "Rule"])
    rs.append(["1", "If Qty + Bales both present: focus Qty; check bales == Qty/BaleSize; mismatch -> highlight (no silent fix)"])
    rs.append(["2", "If only Bales: auto Qty = Bales * Bale Size"])
    rs.append(["3", "Value always = Qty * ExMill"])
    rs.append(["Detect", "Order Qty header = Qnty/Qty — NOT Qnty Per Color / Qnty pre Design"])
    rs.append(["Detect", "If Qnty header column empty but next column has numbers -> use shifted column (flag warning)"])
    rs.append(["Distributor", "From filename (Choice Corner / DCA Order / kag)"])
    rs.column_dimensions["A"].width = 12
    rs.column_dimensions["B"].width = 110

    for meta, lines in all_results:
        name = meta["distributor_hint"]
        # Brandwise
        ws = wb.create_sheet(f"{name} Brandwise"[:31])
        ws.append(["File", meta["file"]])
        ws.append(["Distributor hint", meta["distributor_hint"]])
        ws.append(
            [
                "Qty detection",
                f"col {get_column_letter(meta['qty_col'])} ({meta['qty_how']})",
            ]
        )
        ws.append(["Brand", "Qty", "Bales (sheet)", "Value (=Qty*ExMill)", "Bale check"])
        for cell in ws[4]:
            cell.font = hf
            cell.fill = hfill
        club = defaultdict(lambda: {"q": 0.0, "b": 0.0, "v": 0.0, "has_b": False, "mm": 0})
        for x in lines:
            club[x["brand"]]["q"] += x["qty"] or 0
            club[x["brand"]]["v"] += x["value"] or 0
            if x["bales"] is not None:
                club[x["brand"]]["b"] += x["bales"]
                club[x["brand"]]["has_b"] = True
            if x["bale_status"] == "MISMATCH":
                club[x["brand"]]["mm"] += 1
        tq = tv = tb = 0.0
        any_b = False
        for brand in sorted(club):
            c = club[brand]
            check = "MISMATCH" if c["mm"] else ("" if not c["has_b"] else "OK")
            ws.append(
                [
                    brand,
                    round(c["q"], 2),
                    round(c["b"], 2) if c["has_b"] else "",
                    round(c["v"], 2),
                    check,
                ]
            )
            if check == "MISMATCH":
                for col in range(1, 6):
                    ws.cell(ws.max_row, col).fill = bad
            tq += c["q"]
            tv += c["v"]
            if c["has_b"]:
                tb += c["b"]
                any_b = True
        ws.append(["TOTAL", round(tq, 2), round(tb, 2) if any_b else "", round(tv, 2), ""])
        for c in ws[ws.max_row]:
            c.font = nf
            c.fill = tfill
        for col, w in zip("ABCDE", [28, 12, 14, 18, 12]):
            ws.column_dimensions[col].width = w

        # Lines
        ws = wb.create_sheet(f"{name} Lines"[:31])
        ws.append(
            [
                "Brand",
                "TC",
                "Size",
                "Bale Size",
                "ExMill",
                "Qty",
                "Bales",
                "Value(=Qty*Ex)",
                "Sheet value",
                "Bale check",
                "Value check",
                "Clean bale multiple?",
                "Notes",
            ]
        )
        for cell in ws[1]:
            cell.font = hf
            cell.fill = hfill
        for x in lines:
            ws.append(
                [
                    x["brand"],
                    x["tc"],
                    x["size"],
                    x["bs"],
                    round(x["ex"], 2) if x["ex"] is not None else None,
                    round(x["qty"], 2) if x["qty"] is not None else None,
                    x["bales"] if x["bales"] is not None else "",
                    round(x["value"], 2) if x["value"] is not None else None,
                    round(x["sheet_value"], 2) if x["sheet_value"] is not None else "",
                    x["bale_status"],
                    x["value_status"],
                    "YES" if x["clean_bale_multiple"] else ("NO" if x["clean_bale_multiple"] is False else ""),
                    x["notes"],
                ]
            )
            if x["bale_status"] == "MISMATCH":
                ws.cell(ws.max_row, 10).fill = bad
            if x["value_status"] == "SHEET_VALUE_DIFFERS":
                ws.cell(ws.max_row, 11).fill = bad
            if x["clean_bale_multiple"] is False:
                ws.cell(ws.max_row, 12).fill = warn
        for i, w in enumerate([22, 22, 10, 10, 10, 10, 10, 14, 12, 14, 18, 16, 40], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print("Saved", OUT)
    print()
    for meta, lines in all_results:
        q = sum(x["qty"] or 0 for x in lines)
        v = sum(x["value"] or 0 for x in lines)
        print(
            meta["file"],
            "| dist",
            meta["distributor_hint"],
            "| qty_col",
            get_column_letter(meta["qty_col"]),
            meta["qty_how"],
            "| lines",
            len(lines),
            "| qty",
            round(q, 2),
            "| value",
            round(v, 2),
            "| bales_col",
            meta["bales_col"],
        )

    subprocess.Popen(["cmd", "/c", "start", "", str(OUT)], shell=False)


if __name__ == "__main__":
    main()
