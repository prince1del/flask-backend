"""Towel FINAL merged AM — same season-column layout as bedsheet FINAL merge."""
from __future__ import annotations

import json
import re
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import article_master_db as amdb
import article_master_parser as amparser

SRC_AW25 = Path(r"E:\test files\AW-25 Towel Booking Sheet.xlsx")
SRC_AW26 = Path(
    r"G:\My Drive\2026-2027\Oder Management\AW26 order\Towel"
    r"\AW-26 Towel Phase-2 Booking Sheet.xlsx"
)
OUT = Path(r"E:\test files\Article_Master_FINAL_Merged_Bed_Towel.xlsx")
OUT_REPO = Path(r"E:\centralized-db-system\Output\Article_Master_FINAL_Merged_Bed_Towel.xlsx")
DB_PATH = Path(r"E:\centralized-db-system\centralized_db.sqlite3")
USER_ID = 2

# Bedsheet style: SS-25 / SS-26 / AW-26 from Bed DB.
# Towel training adds AW-25. Latest visible = AW-26.
SEASONS = ["SS-25", "SS-26", "AW-25", "AW-26"]
LATEST_SEASON = "AW-26"
MONEY_SEASONS = ["MRP", "Ex-Mill", "PTR"]
PCT_LATEST = ["AWD Mark up on Exmill", "Proposed Customer Discount", "Retailer Margin"]

ATTR_COLS = [
    "Category",
    "Product",
    "Brand",
    "Size",
    "TC",
    "Units",
    "BS Size",
    "Pillow Size",
    "Color",
    "Pillow Stitching Style",
    "Print Style",
    "Blend",
    "Packing",
    "Bale Pack Size",
]

SIZE_CODE_TO_DISPLAY = {
    "40X60": "Hand Towel",
    "75X150": "Bath Towel",
    "60X120": "Ladies Towel",
    "90X180": "Pool Towel",
    "R4": "Towel Set",
    "30X30": "Face Towel",
    "72X144": "Bath Towel",
    "50X70": "Bath Mat",
    "50X100": "Gym Towel",
    "91X100": "91x100",
    "L": "Large",
    "XL": "Extra Large",
    "XXL": "Double Extra Large",
}

SIZE_SORT_RANK = {
    "Face Towel": 10,
    "Face Towel Set of 3": 11,
    "Hand Towel": 20,
    "Hand Towel Set of 2": 21,
    "Ladies Towel": 30,
    "Bath Towel": 40,
    "Bath Mat": 45,
    "Pool Towel": 50,
    "Towel Set": 60,
    "Gym Towel": 70,
    "91x100": 75,
    "Large": 80,
    "Extra Large": 81,
    "Double Extra Large": 82,
}

PVC_RE = re.compile(r"\(?\s*PVC\s*bag\s*Pkg\.?\s*\)?", re.IGNORECASE)
STICKY = {"Color", "Packing", "Print Style", "Blend", "Pillow Size", "Pillow Stitching Style"}


def blank(v):
    return amparser.is_blank_attr_value(v)


def _norm_ws(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())


def normalize_brand(brand) -> str | None:
    if blank(brand):
        return None
    text = _norm_ws(brand)
    upper = text.upper()
    if upper == "BD WHITE":
        return "BD White"
    if upper == "GYM TOWEL":
        return "Gym Towel"
    if upper == "HUK A BUK":
        return "Huk A Buk"
    if upper == "LEPORD":
        return "Leopard"
    if upper == "ECO STRIPE":
        return "Eco Stripe"
    return text


def normalize_product(product) -> str | None:
    if blank(product):
        return None
    text = _norm_ws(product)
    key = re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()
    if key.startswith("bathmat"):
        return "Bathmat Antiskid"
    return text


def normalize_size(size) -> str | None:
    if blank(size):
        return None
    raw = _norm_ws(size)
    key = re.sub(r"\s+", "", raw).upper()
    if (
        re.fullmatch(r"40X60\(2PC\)|40X60\(SETOF2\)|40X60SETOF2", key)
        or ("SET OF 2" in raw.upper() and "40" in key and "60" in key)
        or re.search(r"40X60.*2PC|40X60\(2", key)
    ):
        return "Hand Towel Set of 2"
    if (
        re.fullmatch(r"30X30\(3PC\)|30X30\(SETOF3\)", key)
        or ("SET OF 3" in raw.upper() and "30" in key)
        or re.search(r"30X30.*3PC|30X30\(3", key)
    ):
        return "Face Towel Set of 3"
    compact = re.sub(r"[^0-9A-Z]", "", key)
    for code, display in SIZE_CODE_TO_DISPLAY.items():
        if compact == re.sub(r"[^0-9A-Z]", "", code) or key == code:
            return display
    return raw


def normalize_color_and_packing(shade, description):
    text = _norm_ws(shade)
    desc = _norm_ws(description)
    blob = f"{text} {desc}"
    has_pvc = bool(PVC_RE.search(blob))
    has_pkg_l = bool(
        re.search(r"\(\s*L\s*\)", blob, re.IGNORECASE)
        or (re.search(r"\bPkg\b", blob, re.IGNORECASE) and not has_pvc)
    )
    packing = "PVC bag Pkg" if has_pvc else None
    text = PVC_RE.sub("", text)
    text = re.sub(r"\(\s*L\s*\)", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\(\s*Pkg\.?\s*\)", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip(" -")
    color = ""
    if text:
        if re.fullmatch(r"jacquarad|jacquard", text, flags=re.IGNORECASE):
            color = "Jacquard"
        else:
            m2 = re.fullmatch(
                r"(?:Assorted|Asorted|Assortede|Asst\.?|Asst)\s*[- ]*0*(\d+)",
                text,
                flags=re.IGNORECASE,
            )
            if m2:
                color = f"Assorted {int(m2.group(1)):02d}"
            elif text.upper() == "WHITE":
                color = "White"
            else:
                color = text
    return color or None, packing, has_pkg_l and not has_pvc, has_pvc


def to_float(v):
    if blank(v):
        return None
    if isinstance(v, str):
        s = v.strip().replace("%", "").replace(",", "")
        return float(s) if s else None
    return float(v)


def fmt_pct(v):
    num = to_float(v)
    if num is None:
        return None
    if abs(num) <= 1:
        return f"{int(round(num * 100))}%"
    return f"{int(round(num))}%"


def fmt_mrp(v):
    num = to_float(v)
    return None if num is None else int(round(num))


def fmt_money(v):
    num = to_float(v)
    return None if num is None else round(num + 0.0, 2)


def detect_header_and_map(ws):
    for r in range(1, min(15, (ws.max_row or 0) + 1)):
        vals = {}
        for c in range(1, min(25, (ws.max_column or 0) + 1)):
            v = ws.cell(r, c).value
            if v is None:
                continue
            vals[_norm_ws(v).lower()] = c
        if "brand" in vals and "size" in vals:
            return r, vals
    raise ValueError("header not found")


def _col(cmap, *names):
    for n in names:
        if n in cmap:
            return cmap[n]
    return None


def parse_towel_sheet(path: Path, season: str):
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_r, cmap = detect_header_and_map(ws)
    c_prod = _col(cmap, "product")
    c_brand = _col(cmap, "brand")
    c_shade = _col(cmap, "shade")
    c_desc = _col(cmap, "description")
    c_size = _col(cmap, "size")
    c_bale = _col(cmap, "pack sizes", "bale pack sizes", "bale pack size", "bale size")
    c_awd = _col(cmap, "awd mu", "awd md")
    c_ret = _col(cmap, "retailer md", "retailer margin")
    c_mrp = _col(cmap, "mrp")
    c_ptr = _col(cmap, "ptr")
    c_ex = _col(cmap, "ex-mill per pcs", "ex mill per pcs", "ex-mill", "ex mill")

    rows = []
    for r in range(header_r + 1, (ws.max_row or 0) + 1):
        product = ws.cell(r, c_prod).value if c_prod else None
        brand = ws.cell(r, c_brand).value if c_brand else None
        shade = ws.cell(r, c_shade).value if c_shade else None
        desc = ws.cell(r, c_desc).value if c_desc else None
        size = ws.cell(r, c_size).value if c_size else None
        bale = ws.cell(r, c_bale).value if c_bale else None
        awd = ws.cell(r, c_awd).value if c_awd else None
        retailer = ws.cell(r, c_ret).value if c_ret else None
        mrp = ws.cell(r, c_mrp).value if c_mrp else None
        ptr = ws.cell(r, c_ptr).value if c_ptr else None
        ex_mill = ws.cell(r, c_ex).value if c_ex else None
        if not any([product, brand, size, mrp]):
            continue

        color, packing, is_pkg_only, has_pvc = normalize_color_and_packing(shade, desc)
        brand_n = normalize_brand(brand)
        size_n = normalize_size(size)
        product_n = normalize_product(product)
        rows.append(
            {
                "_season": season,
                "_src_row": r,
                "_is_pkg_only": is_pkg_only,
                "_has_pvc": has_pvc,
                "Category": "Bath",
                "Product": product_n,
                "Brand": brand_n,
                "Size": size_n,
                "TC": None,
                "Units": None,
                "BS Size": None,
                "Pillow Size": None,
                "Color": color,
                "Pillow Stitching Style": None,
                "Print Style": None,
                "Blend": None,
                "Packing": packing,
                "Bale Pack Size": bale,
                "MRP": mrp,
                "Ex-Mill": ex_mill,
                "PTR": ptr,
                "AWD Mark up on Exmill": awd,
                "Retailer Margin": retailer,
                "Proposed Customer Discount": None,
            }
        )
    return rows


def dedupe_pkg(rows):
    by_key = {}
    for row in rows:
        key = (
            row.get("Brand"),
            row.get("Size"),
            row.get("Color"),
            row.get("Product"),
            row.get("_season"),
        )
        by_key.setdefault(key, []).append(row)
    kept, dropped = [], []
    for group in by_key.values():
        if len(group) == 1:
            kept.append(group[0])
            continue
        pvc = [g for g in group if g["_has_pvc"]]
        non_pkg = [g for g in group if not g["_is_pkg_only"]]
        pool = pvc or non_pkg or group
        winner = max(pool, key=lambda g: g["_src_row"])
        kept.append(winner)
        for g in group:
            if g is not winner:
                dropped.append(g)
    return kept, dropped


def merge_towel(rows):
    """Bedsheet-style merge: attrs sticky latest; prices per season."""
    store = {}
    order = {"AW-25": 0, "AW-26": 1}
    rows_sorted = sorted(rows, key=lambda r: (order.get(r["_season"], 9), r["_src_row"]))

    for r in rows_sorted:
        b = "" if blank(r.get("Brand")) else str(r["Brand"]).strip().upper()
        s = "" if blank(r.get("Size")) else str(r["Size"]).strip().upper()
        c = "" if blank(r.get("Color")) else str(r["Color"]).strip().upper()
        p = "" if blank(r.get("Product")) else str(r["Product"]).strip().upper()
        key = f"{b}|{s}|{c}|{p}"
        if key not in store:
            store[key] = {
                "attrs": {col: None for col in ATTR_COLS},
                "prices": {season: {f: None for f in MONEY_SEASONS} for season in SEASONS},
                "pct": {col: None for col in PCT_LATEST},
            }
        cur = store[key]
        season = r["_season"]

        for col in ATTR_COLS:
            new = r.get(col)
            if blank(new):
                continue
            if isinstance(new, str):
                new = new.strip()
            cur["attrs"][col] = new

        for f in MONEY_SEASONS:
            if not blank(r.get(f)):
                cur["prices"][season][f] = r.get(f)

        for col in PCT_LATEST:
            if not blank(r.get(col)):
                cur["pct"][col] = r.get(col)

    out = []
    for key, cur in store.items():
        row = dict(cur["attrs"])
        for season in SEASONS:
            for f in MONEY_SEASONS:
                val = cur["prices"][season][f]
                col = f"{f} ({season})"
                row[col] = fmt_mrp(val) if f == "MRP" else fmt_money(val)
        for col in PCT_LATEST:
            row[col] = fmt_pct(cur["pct"][col])
        out.append(row)

    def sort_key(row):
        return (
            str(row.get("Brand") or "").upper(),
            SIZE_SORT_RANK.get(row.get("Size") or "", 999),
            str(row.get("Size") or ""),
            str(row.get("Color") or ""),
            str(row.get("Product") or ""),
        )

    out.sort(key=sort_key)
    return out


def build_columns(merged_rows):
    money_cols = []
    for f in MONEY_SEASONS:
        for season in SEASONS:
            col = f"{f} ({season})"
            if any(not blank(r.get(col)) for r in merged_rows):
                money_cols.append(col)
    return ATTR_COLS + money_cols + PCT_LATEST


def apply_old_season_grouping(ws, columns):
    ws.sheet_properties.outlinePr.summaryRight = True
    ws.sheet_properties.outlinePr.applyStyles = True
    for f in MONEY_SEASONS:
        old_idxs = [
            i
            for i, h in enumerate(columns, 1)
            if h.startswith(f"{f} (") and not h.endswith(f"({LATEST_SEASON})")
        ]
        if not old_idxs:
            continue
        start = get_column_letter(min(old_idxs))
        end = get_column_letter(max(old_idxs))
        ws.column_dimensions.group(start, end, outline_level=1, hidden=False)


def load_bed_tob_rows():
    """Bed/TOB from DB with full season history (SS-25 / SS-26 / AW-26) like bedsheet FINAL."""
    conn = sqlite3.connect(DB_PATH)
    amdb.ensure_schema(conn)
    articles = amdb.get_all_articles(conn, USER_ID, active_only=True)
    rows = []
    for a in articles:
        if a.get("category") == "Bath":
            continue
        extra = a.get("extra_attributes") or {}
        if isinstance(extra, str):
            try:
                extra = json.loads(extra)
            except Exception:
                extra = {}
        extra_l = {str(k).strip().lower(): v for k, v in (extra or {}).items()}
        size_disp = amparser.size_display_name(a.get("size")) or a.get("size")
        row = {col: None for col in ATTR_COLS}
        row["Category"] = a.get("category")
        row["Product"] = a.get("product_type")
        row["Brand"] = a.get("brand")
        row["Size"] = size_disp
        row["TC"] = extra_l.get("tc")
        for k in (
            "Units",
            "BS Size",
            "Pillow Size",
            "Pillow Stitching Style",
            "Print Style",
            "Blend",
            "Packing",
        ):
            row[k] = extra_l.get(k.lower()) or extra.get(k)
        for ek, ev in (extra or {}).items():
            if ek in row and blank(row[ek]):
                row[ek] = ev
        row["Bale Pack Size"] = a.get("bale_pack_size")

        # Pull SS-25 / SS-26 / AW-26 (and any other) from season table — same as AM download
        payload = amdb.get_season_prices_last_n(conn, a["id"], USER_ID, limit=5)
        by_field = payload.get("rows") or {}
        cur_tag = amparser.normalize_season_tag(a.get("season_tag"))
        for season in SEASONS:
            for label, key in (("MRP", "mrp"), ("Ex-Mill", "ex_mill_price"), ("PTR", "ptr")):
                val = (by_field.get(key) or {}).get(season)
                if val is None and cur_tag == season:
                    val = a.get(key)
                col = f"{label} ({season})"
                row[col] = fmt_mrp(val) if label == "MRP" else fmt_money(val)

        row["AWD Mark up on Exmill"] = fmt_pct(
            extra_l.get("awd mu")
            or extra_l.get("awd mark up on exmill")
            or extra.get("AWD Mark up on Exmill")
        )
        row["Retailer Margin"] = fmt_pct(
            extra_l.get("retailer md")
            or extra_l.get("retailer margin")
            or extra.get("Retailer Margin")
        )
        row["Proposed Customer Discount"] = fmt_pct(
            extra_l.get("proposed customer discount")
            or extra_l.get("perceived")
            or extra.get("Proposed Customer Discount")
            or extra.get("Perceived")
        )
        rows.append(row)
    conn.close()
    return amparser.sort_articles_for_display(
        [
            {
                **r,
                "brand": r.get("Brand"),
                "size": r.get("Size"),
                "product_type": r.get("Product"),
            }
            for r in rows
        ]
    )


def main():
    aw25, drop25 = dedupe_pkg(parse_towel_sheet(SRC_AW25, "AW-25"))
    aw26, drop26 = dedupe_pkg(parse_towel_sheet(SRC_AW26, "AW-26"))
    bath = merge_towel(aw25 + aw26)
    bed = load_bed_tob_rows()

    # One sheet: Bed/TOB then Bath (bedsheet-style season cols throughout)
    all_rows = bed + bath
    columns = build_columns(all_rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Article Master"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for c, h in enumerate(columns, 1):
        cell = ws.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    for r_idx, row in enumerate(all_rows, 2):
        for c_idx, h in enumerate(columns, 1):
            val = row.get(h)
            cell = ws.cell(r_idx, c_idx, val)
            if h.startswith("MRP ") and isinstance(val, int):
                cell.number_format = "0"
            elif (h.startswith("Ex-Mill ") or h.startswith("PTR ")) and isinstance(
                val, (int, float)
            ):
                cell.number_format = "0.00"

    for col in ws.columns:
        letter = col[0].column_letter
        width = max(len(str(cell.value or "")) for cell in col[: min(80, len(col))])
        ws.column_dimensions[letter].width = min(max(width + 2, 12), 36)

    apply_old_season_grouping(ws, columns)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # Bath-only sheet (pure towel review like bedsheet FINAL)
    ws_b = wb.create_sheet("Bath Towel")
    bath_cols = build_columns(bath)
    for c, h in enumerate(bath_cols, 1):
        cell = ws_b.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    for r_idx, row in enumerate(bath, 2):
        for c_idx, h in enumerate(bath_cols, 1):
            val = row.get(h)
            cell = ws_b.cell(r_idx, c_idx, val)
            if h.startswith("MRP ") and isinstance(val, int):
                cell.number_format = "0"
            elif (h.startswith("Ex-Mill ") or h.startswith("PTR ")) and isinstance(
                val, (int, float)
            ):
                cell.number_format = "0.00"
    for col in ws_b.columns:
        letter = col[0].column_letter
        width = max(len(str(cell.value or "")) for cell in col[:80])
        ws_b.column_dimensions[letter].width = min(max(width + 2, 12), 36)
    apply_old_season_grouping(ws_b, bath_cols)
    ws_b.auto_filter.ref = ws_b.dimensions
    ws_b.freeze_panes = "A2"

    rules = wb.create_sheet("0_Rules_Applied", 0)
    lines = [
        "Towel FINAL merged Article Master — same layout as bedsheet FINAL (review only, not uploaded)",
        "",
        "Identity: Brand + Size + Color + Product",
        "Size display: Hand Towel / Face Towel / Bath Towel / Ladies / Pool / Towel Set / …",
        "Color from Shade; Assorted NN zero-padded; Jacquarad→Jacquard; WHITE→White",
        "Pkg/(L) ignored — keep PVC / latest sibling",
        "PVC (Shade or Description) → Packing = PVC bag Pkg",
        "Bathmat → Bathmat Antiskid; Lepord→Leopard; Eco Stripe; BD White; Gym Towel",
        "Pack Sizes ≡ Bale Pack Size; Description / SL NO / Colours / AWD order ignored",
        "Prices: MRP/Ex-Mill/PTR per season — Bed: SS-25/SS-26/AW-26; Towel: AW-25/AW-26",
        "OLD seasons Excel-grouped; AW-26 always visible (same as bedsheet FINAL)",
        "Margins: latest non-blank wins (shown as %)",
        f"Dropped Pkg rows: {len(drop25) + len(drop26)}",
        f"Bath SKUs merged: {len(bath)} | Bed/TOB from DB: {len(bed)}",
        f"Sources: {SRC_AW25.name} + {SRC_AW26.name}",
    ]
    for i, line in enumerate(lines, 1):
        rules.cell(i, 1, line)
    rules.column_dimensions["A"].width = 110

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT_REPO.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    wb.save(OUT_REPO)
    print(f"saved {OUT}")
    print(f"Bath={len(bath)} Bed/TOB={len(bed)} cols={columns}")
    # Luxury Living sample
    for r in bath:
        if r.get("Brand") == "Luxury Living":
            print(
                r.get("Size"),
                r.get("Color"),
                "MRP25=",
                r.get("MRP (AW-25)"),
                "MRP26=",
                r.get("MRP (AW-26)"),
            )


if __name__ == "__main__":
    main()
